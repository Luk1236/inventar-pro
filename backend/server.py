from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, Response, Query, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, field_validator
from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta, timezone
from passlib.context import CryptContext
from jose import JWTError, jwt
import os
import re
import shutil
import logging
import uuid
import asyncio
from apscheduler.schedulers.asyncio import AsyncIOScheduler
import json
import secrets
import csv
import io
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from websocket_handler import manager, websocket_endpoint

# Configure logging (must be early for use throughout file)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
# M5 — Rotating file handler: 10 MB per file, keep 5 files → max 50 MB on disk
_log_dir = os.path.join(os.path.dirname(__file__), "logs")
os.makedirs(_log_dir, exist_ok=True)
from logging.handlers import RotatingFileHandler as _RotatingFileHandler
_file_handler = _RotatingFileHandler(
    os.path.join(_log_dir, "app.log"),
    maxBytes=10 * 1024 * 1024,
    backupCount=5,
    encoding="utf-8",
)
_file_handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
logging.getLogger().addHandler(_file_handler)
logger = logging.getLogger(__name__)

# Cross-platform sync will be implemented in next version
# For now, we'll add basic cross-platform support

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# ===========================================
# SECURITY CONFIGURATION
# ===========================================

# Rate Limiter - Protects against brute force and DoS
limiter = Limiter(key_func=get_remote_address)

security = HTTPBearer()
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

SECRET_KEY = os.environ.get("SECRET_KEY")
_known_weak_keys = {
    "change_this_to_a_secure_random_string_in_production",
    "inventory_management_secret_key_2025",
    "changeme",
    "",
}
if not SECRET_KEY or SECRET_KEY in _known_weak_keys:
    raise RuntimeError(
        "⛔ CRITICAL: SECRET_KEY ist nicht gesetzt oder verwendet einen unsicheren Standardwert.\n"
        "Bitte einen sicheren Key in der .env setzen:\n"
        "  SECRET_KEY=" + secrets.token_hex(32) + "\n"
        "Server wird nicht gestartet."
    )

ALGORITHM = "HS256"

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
DB_NAME = os.environ.get('DB_NAME', 'inventory_db')
client = AsyncIOMotorClient(mongo_url)
db = client[DB_NAME]

# Backup directory
BACKUP_DIR = Path(os.environ.get("BACKUP_DIR", str(Path(__file__).parent / "backups")))
BACKUP_DIR.mkdir(exist_ok=True)
BACKUP_FILE = BACKUP_DIR / "latest_backup.json"
MAX_EXPORT_ROWS = 5000   # K2: Hard cap for all export endpoints (CSV/PDF/Excel)
MAX_BACKUP_SIZE_MB = 500  # H2: Backup download size limit
MAX_IMPORT_ROWS = 1000   # F10: Hard cap for bulk article import
DASHBOARD_CACHE_TTL = 30  # F9: Dashboard stats cached for 30 seconds

# F9: In-memory dashboard cache — avoids 17 count_documents() calls on every page load
_dashboard_cache: dict = {}
_dashboard_cache_ts: float = 0.0

# Scheduler for daily backups
scheduler = AsyncIOScheduler()

# Email configuration
ADMIN_EMAIL = os.getenv("ADMIN_EMAIL", "")
SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
NOTIFICATION_EMAIL = os.getenv("NOTIFICATION_EMAIL", "")

# Token Configuration
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.environ.get("ACCESS_TOKEN_EXPIRE_MINUTES", "15"))
REFRESH_TOKEN_EXPIRE_DAYS = int(os.environ.get("REFRESH_TOKEN_EXPIRE_DAYS", "7"))

app = FastAPI(title="Inventory Management System", version="1.0.0")
api_router = APIRouter(prefix="/api")

# M4 — Health check (used by load balancers, Docker HEALTHCHECK, monitoring)
@app.get("/health", tags=["health"])
async def health_check():
    try:
        await db.command("ping")
        return {"status": "ok", "db": "connected", "version": "1.0.0"}
    except Exception:
        from fastapi.responses import JSONResponse
        return JSONResponse(
            {"status": "degraded", "db": "unreachable", "version": "1.0.0"},
            status_code=503
        )

# ===========================================
# RBAC - Role-Based Access Control
# ===========================================

class Permission:
    # Article permissions
    VIEW_ARTICLES = "view_articles"
    CREATE_ARTICLE = "create_article"
    EDIT_ARTICLE = "edit_article"
    DELETE_ARTICLE = "delete_article"
    
    # Stock permissions
    VIEW_STOCK = "view_stock"
    EDIT_STOCK = "edit_stock"
    
    # Event permissions
    VIEW_EVENTS = "view_events"
    CREATE_EVENT = "create_event"
    EDIT_EVENT = "edit_event"
    DELETE_EVENT = "delete_event"
    
    # Customer permissions
    VIEW_CUSTOMERS = "view_customers"
    CREATE_CUSTOMER = "create_customer"
    EDIT_CUSTOMER = "edit_customer"
    
    # Invoice permissions
    CREATE_INVOICE = "create_invoice"
    VIEW_INVOICES = "view_invoices"

    # Admin permissions
    MANAGE_USERS = "manage_users"
    VIEW_REPORTS = "view_reports"
    ADMIN_ACCESS = "admin_access"
    BACKUP_DATABASE = "backup_database"

# Role permission mappings
ROLE_PERMISSIONS = {
    "admin": [
        Permission.VIEW_ARTICLES, Permission.CREATE_ARTICLE, Permission.EDIT_ARTICLE, Permission.DELETE_ARTICLE,
        Permission.VIEW_STOCK, Permission.EDIT_STOCK,
        Permission.VIEW_EVENTS, Permission.CREATE_EVENT, Permission.EDIT_EVENT, Permission.DELETE_EVENT,
        Permission.VIEW_CUSTOMERS, Permission.CREATE_CUSTOMER, Permission.EDIT_CUSTOMER,
        Permission.MANAGE_USERS, Permission.VIEW_REPORTS, Permission.ADMIN_ACCESS, Permission.BACKUP_DATABASE,
        Permission.CREATE_INVOICE, Permission.VIEW_INVOICES,
    ],
    "lager": [
        Permission.VIEW_ARTICLES, Permission.CREATE_ARTICLE, Permission.EDIT_ARTICLE,
        Permission.VIEW_STOCK, Permission.EDIT_STOCK,
        Permission.VIEW_EVENTS, Permission.CREATE_EVENT, Permission.EDIT_EVENT,
        Permission.VIEW_CUSTOMERS, Permission.CREATE_CUSTOMER,
        Permission.VIEW_REPORTS, Permission.CREATE_INVOICE, Permission.VIEW_INVOICES,
    ],
    "techniker": [
        Permission.VIEW_ARTICLES, Permission.EDIT_ARTICLE,
        Permission.VIEW_STOCK,
        Permission.VIEW_EVENTS,
        Permission.VIEW_CUSTOMERS
    ],
    # F12: New fine-grained roles
    "viewer": [
        # Read-only access across all main entities
        Permission.VIEW_ARTICLES,
        Permission.VIEW_STOCK,
        Permission.VIEW_EVENTS,
        Permission.VIEW_CUSTOMERS,
        Permission.VIEW_REPORTS,
    ],
    "fahrer": [
        # Drivers: see events + articles (for loading/unloading), no customer/financial data
        Permission.VIEW_ARTICLES,
        Permission.VIEW_STOCK,
        Permission.VIEW_EVENTS,
    ],
}

def has_permission(user_role: str, permission: str) -> bool:
    """Check if a role has a specific permission"""
    return permission in ROLE_PERMISSIONS.get(user_role, [])

def require_permission(permission: str):
    """Dependency to check user permission"""
    async def permission_checker(current_user: "User" = Depends(get_current_user)):
        if not has_permission(current_user.role, permission):
            raise HTTPException(
                status_code=403, 
                detail="Keine Berechtigung für diese Aktion"
            )
        return current_user
    return permission_checker

# Models
class UserRole(str):
    ADMIN = "admin"
    LAGER = "lager"
    TECHNIKER = "techniker"
    VIEWER = "viewer"   # F12: read-only across all entities
    FAHRER = "fahrer"   # F12: drivers — events + articles only

class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    username: str
    role: str
    is_active: bool = True
    is_approved: bool = False  # New field for admin approval
    profile_image: Optional[str] = None  # Base64 profile image
    created_at: datetime = Field(default_factory=datetime.utcnow)

class UserCreate(BaseModel):
    email: EmailStr
    username: str = Field(..., min_length=3, max_length=50)
    password: str = Field(..., min_length=8, max_length=128)
    role: str = UserRole.LAGER

class UserLogin(BaseModel):
    username: str = Field(..., min_length=1, max_length=50)
    password: str = Field(..., min_length=1, max_length=128)

class Token(BaseModel):
    access_token: str
    refresh_token: str  # NEW: Refresh token
    token_type: str
    user: User
    expires_in: int  # NEW: Seconds until access token expires

class RefreshTokenRequest(BaseModel):
    refresh_token: str

# L1 — Base64 image validation helpers
_MAX_IMAGE_BYTES = 10 * 1024 * 1024   # 10 MB decoded limit
_MAX_IMAGES_PER_ARTICLE = 50

def _validate_image_base64(v: Optional[str]) -> Optional[str]:
    """Reject base64 images that exceed 10 MB (decoded size ≈ len*3/4)."""
    if v is None:
        return v
    estimated_bytes = len(v) * 3 // 4
    if estimated_bytes > _MAX_IMAGE_BYTES:
        raise ValueError("Bild überschreitet die maximale Größe von 10 MB")
    return v

def _validate_images_list(v: List[str]) -> List[str]:
    """Reject image lists with too many entries or oversized individual images."""
    if len(v) > _MAX_IMAGES_PER_ARTICLE:
        raise ValueError(f"Maximal {_MAX_IMAGES_PER_ARTICLE} Bilder erlaubt")
    for img in v:
        _validate_image_base64(img)
    return v

class Category(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str = Field(..., max_length=255)
    description: Optional[str] = Field(None, max_length=2000)
    parent_id: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class Supplier(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    contact_person: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    website: Optional[str] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class SupplierCreate(BaseModel):
    name: str = Field(..., max_length=255)
    contact_person: Optional[str] = Field(None, max_length=255)
    email: Optional[EmailStr] = None
    phone: Optional[str] = Field(None, max_length=50)
    address: Optional[str] = Field(None, max_length=500)
    website: Optional[str] = Field(None, max_length=500)
    notes: Optional[str] = Field(None, max_length=5000)

class Team(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    members: List[str] = []  # List of user IDs
    created_at: datetime = Field(default_factory=datetime.utcnow)
    created_by: str

class TeamCreate(BaseModel):
    name: str = Field(..., max_length=255)
    description: Optional[str] = Field(None, max_length=2000)
    members: List[str] = []

class AuditLog(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    action: str  # CREATE, UPDATE, DELETE, LOGIN, LOGOUT
    entity_type: str  # articles, customers, events, etc.
    entity_id: Optional[str] = None
    entity_name: Optional[str] = None
    user_id: Optional[str] = None
    user_name: Optional[str] = None
    changes: Optional[dict] = None  # Before/after values
    ip_address: Optional[str] = None
    timestamp: datetime = Field(default_factory=datetime.utcnow)

class Article(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    category_id: Optional[str] = None  # Made optional
    supplier_id: Optional[str] = None
    serial_number: Optional[str] = None
    inventory_code: str
    base_unit: str = "Stück"
    current_stock: int = 0
    min_stock_level: int = 0
    price_per_unit: float = 0.0
    rental_price: Optional[float] = None  # Tagespreis
    # NEW: Gestaffelte Mietpreise
    rental_price_day: Optional[float] = None  # Tagespreis (1-3 Tage)
    rental_price_week: Optional[float] = None  # Wochenpreis (4-7 Tage)
    rental_price_month: Optional[float] = None  # Monatspreis (>7 Tage)
    # NEW: Equipment-spezifische Felder
    weight_kg: Optional[float] = None  # Gewicht in kg
    power_watt: Optional[int] = None  # Leistung in Watt
    power_type: Optional[str] = None  # "230V", "400V", "Akku", etc.
    operating_hours: Optional[float] = 0.0  # Betriebsstunden
    max_operating_hours: Optional[float] = None  # Max Stunden vor Wartung
    # Rental factors
    rental_factor_weekend: Optional[float] = 1.5  # Wochenend-Faktor
    rental_factor_week: Optional[float] = 3.0  # Wochen-Faktor
    # Consumable
    is_consumable: bool = False  # Ist Verbrauchsmaterial
    # Sub-Rental
    is_sub_rental: bool = False  # Ist Zumiet-Artikel
    sub_rental_supplier_id: Optional[str] = None  # Von wem gemietet
    sub_rental_cost: Optional[float] = None  # Mietkosten
    # Standard fields
    status: str = "OK"  # OK, defekt, gesperrt
    storage_location_id: Optional[str] = None  # Link to storage location
    last_maintenance: Optional[datetime] = None
    next_maintenance: Optional[datetime] = None
    maintenance_interval_days: Optional[int] = None
    image_base64: Optional[str] = None  # Legacy single image
    images: List[str] = []  # Multiple images as base64 strings
    qr_code: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class ArticleCreate(BaseModel):
    name: str = Field(..., max_length=255)
    description: Optional[str] = Field(None, max_length=5000)
    category_id: Optional[str] = None  # Made optional
    supplier_id: Optional[str] = None
    serial_number: Optional[str] = Field(None, max_length=100)
    inventory_code: Optional[str] = None
    base_unit: str = Field("Stück", max_length=50)
    current_stock: int = 0
    min_stock_level: int = 0
    price_per_unit: float = 0.0
    rental_price: Optional[float] = None
    # NEW: Gestaffelte Mietpreise
    rental_price_day: Optional[float] = None
    rental_price_week: Optional[float] = None
    rental_price_month: Optional[float] = None
    # NEW: Equipment-spezifische Felder
    weight_kg: Optional[float] = None
    power_watt: Optional[int] = None
    power_type: Optional[str] = Field(None, max_length=50)
    operating_hours: Optional[float] = 0.0
    max_operating_hours: Optional[float] = None
    rental_factor_weekend: Optional[float] = 1.5
    rental_factor_week: Optional[float] = 3.0
    is_consumable: bool = False
    is_sub_rental: bool = False
    sub_rental_supplier_id: Optional[str] = None
    sub_rental_cost: Optional[float] = None
    # Standard fields
    storage_location_id: Optional[str] = None  # Link to storage location
    maintenance_interval_days: Optional[int] = None
    image_base64: Optional[str] = None  # Legacy
    images: List[str] = []  # Multiple images

    @field_validator("image_base64", mode="before")
    @classmethod
    def validate_image_base64(cls, v):
        return _validate_image_base64(v)

    @field_validator("images", mode="before")
    @classmethod
    def validate_images(cls, v):
        return _validate_images_list(v)

class StorageZone(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    type: str  # Innenlager, Sperrlager, Transport
    description: Optional[str] = None
    image_base64: Optional[str] = None  # Zone photo
    images: List[str] = []  # Multiple images
    qr_code: Optional[str] = None
    grid_width: Optional[int] = None   # Anzahl Felder in X-Richtung (1 Feld = 1,5 m)
    grid_depth: Optional[int] = None   # Anzahl Felder in Z-Richtung (1 Feld = 1,5 m)
    created_at: datetime = Field(default_factory=datetime.utcnow)

class StorageLocation(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    zone_id: str
    name: str  # Regal-1, Fach-A, Case-123
    type: str  # Regal, Fach, Case, Container
    capacity: Optional[int] = None
    image_base64: Optional[str] = None  # Location photo
    images: List[str] = []  # Multiple images
    qr_code: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class InventoryMovement(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    article_id: str
    movement_type: str  # IN, OUT, TRANSFER
    quantity: int
    from_location_id: Optional[str] = None
    to_location_id: Optional[str] = None
    user_id: str
    reason: Optional[str] = None
    reference_number: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class MovementCreate(BaseModel):
    article_id: str
    movement_type: str = Field(..., max_length=20)
    quantity: int = Field(..., ge=1, le=100000)
    from_location_id: Optional[str] = None
    to_location_id: Optional[str] = None
    reason: Optional[str] = Field(None, max_length=1000)
    reference_number: Optional[str] = Field(None, max_length=100)

# Maintenance Models
class MaintenanceTask(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    article_id: str
    title: str
    description: Optional[str] = None
    task_type: str  # routine, repair, inspection, dguv_v3
    priority: str = "medium"  # low, medium, high, critical
    status: str = "pending"  # pending, in_progress, completed, cancelled
    assigned_to: Optional[str] = None  # user_id
    due_date: Optional[datetime] = None
    completed_at: Optional[datetime] = None
    estimated_duration: Optional[int] = None  # minutes
    actual_duration: Optional[int] = None  # minutes
    created_by: str  # user_id
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class MaintenanceTaskCreate(BaseModel):
    article_id: str
    title: str = Field(..., max_length=255)
    description: Optional[str] = Field(None, max_length=5000)
    task_type: str = Field("routine", max_length=50)
    priority: str = Field("medium", max_length=20)
    assigned_to: Optional[str] = None
    due_date: Optional[datetime] = None
    estimated_duration: Optional[int] = None

# DGUV V3 Prüfung Model
class DGUVV3Inspection(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    article_id: str
    inspection_date: datetime
    next_inspection_date: datetime
    inspector_name: str
    inspector_qualification: Optional[str] = None  # z.B. "Elektrofachkraft"
    inspection_type: str = "wiederkehrend"  # erstprüfung, wiederkehrend, prüfung_nach_reparatur
    result: str = "bestanden"  # bestanden, nicht_bestanden, bedingt_bestanden
    protocol_number: Optional[str] = None
    findings: Optional[List[str]] = None  # Liste der Mängel
    recommendations: Optional[str] = None
    measurement_values: Optional[Dict[str, Any]] = None  # z.B. {"schutzleiterwiderstand": 0.3, "iso_widerstand": 2.5}
    certificate_image: Optional[str] = None  # Base64 image of certificate
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    created_by: str

class DGUVV3InspectionCreate(BaseModel):
    article_id: str
    inspection_date: datetime
    next_inspection_date: datetime
    inspector_name: str
    inspector_qualification: Optional[str] = None
    inspection_type: str = "wiederkehrend"
    result: str = "bestanden"
    protocol_number: Optional[str] = None
    findings: Optional[List[str]] = None
    recommendations: Optional[str] = None
    measurement_values: Optional[Dict[str, Any]] = None
    certificate_image: Optional[str] = None
    notes: Optional[str] = None

# Reparatur-Ticket Model
class RepairTicket(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    ticket_number: str
    article_id: str
    title: str
    description: str
    defect_type: str = "other"  # electrical, mechanical, optical, software, other
    severity: str = "medium"  # low, medium, high, critical
    status: str = "open"  # open, in_progress, waiting_parts, repaired, closed, unrepairable
    reported_by: str
    assigned_to: Optional[str] = None
    defect_images: List[str] = []  # Base64 images of defects
    repair_images: List[str] = []  # Base64 images after repair
    repair_notes: Optional[str] = None
    parts_used: Optional[List[Dict[str, Any]]] = None
    repair_cost: Optional[float] = None
    repair_time_minutes: Optional[int] = None
    warranty_claim: bool = False
    external_repair: bool = False
    external_repair_vendor: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)
    closed_at: Optional[datetime] = None

class RepairTicketCreate(BaseModel):
    article_id: str
    title: str = Field(..., max_length=255)
    description: str = Field(..., max_length=10000)
    defect_type: str = Field("other", max_length=50)
    severity: str = Field("medium", max_length=20)
    defect_images: List[str] = []
    warranty_claim: bool = False

    @field_validator("defect_images", mode="before")
    @classmethod
    def validate_defect_images(cls, v):
        return _validate_images_list(v)

# Betriebsstunden-Update Model
class OperatingHoursUpdate(BaseModel):
    article_id: str
    hours_to_add: float
    notes: Optional[str] = None

class MaintenanceRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    task_id: str
    article_id: str
    performed_by: str  # user_id
    work_description: str
    parts_used: Optional[List[Dict[str, Any]]] = None  # [{"name": "Filter", "quantity": 1, "cost": 25.50}]
    cost: Optional[float] = None
    before_images: Optional[List[str]] = None  # base64 images
    after_images: Optional[List[str]] = None  # base64 images
    next_maintenance_date: Optional[datetime] = None
    status_after: str = "OK"  # OK, defekt, gesperrt
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class MaintenanceRecordCreate(BaseModel):
    task_id: str
    article_id: str
    work_description: str = Field(..., max_length=10000)
    parts_used: Optional[List[Dict[str, Any]]] = None
    cost: Optional[float] = None
    before_images: Optional[List[str]] = None
    after_images: Optional[List[str]] = None
    next_maintenance_date: Optional[datetime] = None
    status_after: str = Field("OK", max_length=20)
    notes: Optional[str] = Field(None, max_length=5000)

class MaintenanceChecklist(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    category_ids: Optional[List[str]] = None  # applicable categories
    items: List[Dict[str, Any]]  # [{"title": "Check cables", "required": true, "type": "checkbox"}]
    created_by: str
    is_template: bool = True
    created_at: datetime = Field(default_factory=datetime.utcnow)

class MaintenanceChecklistCreate(BaseModel):
    name: str = Field(..., max_length=255)
    description: Optional[str] = Field(None, max_length=2000)
    category_ids: Optional[List[str]] = None
    items: List[Dict[str, Any]]
    is_template: bool = True

class MaintenanceExecution(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    task_id: str
    checklist_id: str
    article_id: str
    performed_by: str
    checklist_results: List[Dict[str, Any]]  # completed checklist with results
    overall_status: str = "passed"  # passed, failed, partial
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class MaintenanceExecutionCreate(BaseModel):
    task_id: str
    checklist_id: str
    article_id: str
    checklist_results: List[Dict[str, Any]]
    overall_status: str = "passed"
    notes: Optional[str] = None

# ============================================================
# SETS & BUNDLES - Artikel-Pakete
# ============================================================

class BundleItem(BaseModel):
    article_id: str
    quantity: int = 1
    is_optional: bool = False  # Optional items can be excluded

class Bundle(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    bundle_code: str  # z.B. "PA-SET-001"
    category: str = "Standard"  # PA, Licht, Video, Rigging, Strom, Sonstige
    items: List[BundleItem] = []
    # Calculated fields (stored for quick access)
    total_items: int = 0
    total_weight_kg: float = 0.0
    total_power_watt: int = 0
    rental_price_day: float = 0.0
    rental_price_week: float = 0.0
    rental_price_month: float = 0.0
    # Discount when booking as bundle
    bundle_discount_percent: float = 0.0  # z.B. 10% Rabatt als Paket
    # Status
    is_active: bool = True
    image_base64: Optional[str] = None
    # Metadata
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)
    created_by: str

class BundleCreate(BaseModel):
    name: str = Field(..., max_length=255)
    description: Optional[str] = Field(None, max_length=5000)
    bundle_code: str = Field(..., max_length=100)
    category: str = Field("Standard", max_length=100)
    items: List[BundleItem] = []
    bundle_discount_percent: float = 0.0
    image_base64: Optional[str] = None

    @field_validator("image_base64", mode="before")
    @classmethod
    def validate_image_base64(cls, v):
        return _validate_image_base64(v)

class BundleUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    items: Optional[List[BundleItem]] = None
    bundle_discount_percent: Optional[float] = None
    is_active: Optional[bool] = None
    image_base64: Optional[str] = None

# Packing List Models for Lager-Logistik
class PackingListItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    event_id: str
    booking_id: Optional[str] = None
    article_id: str
    quantity: int
    # Check-Out Status
    checked_out: bool = False
    checked_out_by: Optional[str] = None
    checked_out_at: Optional[datetime] = None
    checkout_condition: Optional[str] = None  # OK, note
    checkout_notes: Optional[str] = None
    # Check-In Status  
    checked_in: bool = False
    checked_in_by: Optional[str] = None
    checked_in_at: Optional[datetime] = None
    checkin_condition: Optional[str] = None  # OK, DIRTY, DEFECT, MISSING
    checkin_notes: Optional[str] = None
    checkin_photos: List[str] = []  # Base64 photos of damage
    # Metadata
    created_at: datetime = Field(default_factory=datetime.utcnow)

class PackingListCheckout(BaseModel):
    item_ids: List[str]  # IDs of items to check out
    condition: str = "OK"  # OK or note
    notes: Optional[str] = None

class PackingListCheckin(BaseModel):
    item_id: str
    condition: str = Field(..., max_length=20)  # OK, DIRTY, DEFECT, MISSING
    notes: Optional[str] = Field(None, max_length=2000)
    photos: List[str] = []  # Base64 encoded photos

# Customer & Event Models
class ContactPerson(BaseModel):
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    role: Optional[str] = None

class Customer(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_number: str  # Auto-generated
    company_name: str
    contact_person: str
    phone: str
    email: EmailStr
    address_street: Optional[str] = None
    address_zip: Optional[str] = None
    address_city: Optional[str] = None
    address_country: Optional[str] = "Deutschland"
    payment_terms: Optional[str] = None  # e.g., "14 Tage netto"
    contract_info: Optional[str] = None
    notes: Optional[str] = None
    contact_persons: List[ContactPerson] = []
    is_active: bool = True
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class CustomerCreate(BaseModel):
    company_name: str = Field(..., max_length=255)
    contact_person: str = Field(..., max_length=255)
    phone: str = Field(..., max_length=50)
    email: EmailStr
    address_street: Optional[str] = Field(None, max_length=255)
    address_zip: Optional[str] = Field(None, max_length=20)
    address_city: Optional[str] = Field(None, max_length=100)
    address_country: Optional[str] = Field("Deutschland", max_length=100)
    payment_terms: Optional[str] = Field(None, max_length=255)
    contract_info: Optional[str] = Field(None, max_length=5000)
    notes: Optional[str] = Field(None, max_length=5000)
    contact_persons: List[ContactPerson] = []

class Event(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    event_number: str  # e.g., EVT-2025-001
    customer_id: str
    event_type: str  # Konzert, Messe, Hochzeit, etc.
    event_name: str
    description: Optional[str] = None
    location: str
    location_type: Optional[str] = None  # Indoor/Outdoor
    start_date: datetime
    end_date: datetime
    setup_date: Optional[datetime] = None
    teardown_date: Optional[datetime] = None
    status: str = "planned"  # planned, confirmed, in_progress, completed, cancelled
    total_value: Optional[float] = None
    notes: Optional[str] = None
    created_by: str  # user_id
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class EventCreate(BaseModel):
    customer_id: str
    event_type: str = Field(..., max_length=100)
    event_name: str = Field(..., max_length=255)
    description: Optional[str] = Field(None, max_length=5000)
    location: str = Field(..., max_length=500)
    location_type: Optional[str] = Field(None, max_length=50)
    start_date: datetime
    end_date: datetime
    setup_date: Optional[datetime] = None
    teardown_date: Optional[datetime] = None
    notes: Optional[str] = Field(None, max_length=10000)

class Booking(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    event_id: str
    article_id: str
    quantity: int
    booked_by: str  # user_id
    status: str = "booked"  # booked, returned, cancelled
    pickup_date: Optional[datetime] = None
    return_date: Optional[datetime] = None
    notes: Optional[str] = None
    serial_number_ids: List[str] = []  # F3: assigned serial numbers
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class BookingCreate(BaseModel):
    event_id: str
    article_id: str
    quantity: int
    pickup_date: Optional[datetime] = None
    return_date: Optional[datetime] = None
    notes: Optional[str] = Field(None, max_length=2000)
    serial_number_ids: List[str] = []  # F3: pre-assign serial numbers at booking time

    @field_validator("return_date", mode="after")
    @classmethod
    def return_date_after_pickup(cls, v, info):
        pickup = info.data.get("pickup_date") if hasattr(info, "data") else None
        if v and pickup and v <= pickup:
            raise ValueError("return_date muss nach pickup_date liegen")
        return v

    @field_validator("quantity", mode="before")
    @classmethod
    def quantity_positive(cls, v):
        if v <= 0:
            raise ValueError("quantity muss größer als 0 sein")
        return v

# Messaging Models
class Message(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sender_id: str
    recipient_id: str
    message_text: str
    is_read: bool = False
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class MessageCreate(BaseModel):
    recipient_id: str
    message_text: str = Field(..., max_length=10000)

class Conversation(BaseModel):
    user_id: str
    username: str
    last_message: Optional[str] = None
    last_message_time: Optional[datetime] = None
    unread_count: int = 0

# Settings Version History Entry
class SettingsVersionHistory(BaseModel):
    version: int
    changed_at: datetime = Field(default_factory=datetime.utcnow)
    changed_by: Optional[str] = None
    changes: dict = Field(default_factory=dict)

class AppSettings(BaseModel):
    # Version tracking
    settings_version: int = Field(1, ge=1)
    # Firmendaten
    company_name: str = Field("", max_length=255)
    company_address: str = Field("", max_length=500)
    company_phone: str = Field("", max_length=50)
    company_email: str = Field("", max_length=255)
    company_website: str = Field("", max_length=500)
    # App Branding (Admin konfigurierbar)
    app_display_name: str = Field("Inventar Pro", max_length=100)
    app_logo_icon: str = Field("cube", max_length=50)  # Ionicons name: cube, cube-outline, archive, business, etc.
    app_primary_color: str = Field("#007AFF", max_length=20)
    app_slogan: str = Field("Professionelle Lagerverwaltung", max_length=200)
    # Nummernkreise
    invoice_prefix: str = Field("INV", max_length=10)
    quote_prefix: str = Field("QUO", max_length=10)
    event_prefix: str = Field("EVT", max_length=10)
    customer_prefix: str = Field("CUST", max_length=10)
    invoice_next_number: int = Field(1, ge=1)
    quote_next_number: int = Field(1, ge=1)
    # Finanzen
    tax_rate: float = Field(19.0, ge=0.0, le=100.0)
    currency: str = Field("EUR", max_length=10)
    payment_terms: List[str] = Field(default_factory=lambda: ["Sofort fällig", "14 Tage netto", "30 Tage netto"])
    payment_methods: List[str] = Field(default_factory=lambda: ["Bar", "Überweisung", "PayPal"])
    # Sonstige
    timezone: str = Field("Europe/Berlin", max_length=50)
    date_format: str = Field("DD.MM.YYYY", max_length=20)
    # Briefpapier
    letterhead_logo_url: str = Field("", max_length=2000)
    letterhead_primary_color: str = Field("#FF9500", max_length=20)
    letterhead_footer_text: str = Field("", max_length=500)
    letterhead_slogan: str = Field("", max_length=200)
    letterhead_show_logo: bool = True
    # Online-Angebote
    online_quotes_enabled: bool = True
    online_quotes_expiry_days: int = Field(30, ge=1, le=365)
    # Digitale Unterschrift
    signature_require_quotes: bool = False
    signature_require_delivery: bool = False
    signature_footer_text: Optional[str] = Field(None, max_length=500)
    # Online-Rechnungen
    online_invoices_enabled: bool = False
    online_invoices_expiry_days: int = Field(30, ge=1, le=365)
    online_invoices_payment_text: Optional[str] = Field(None, max_length=500)
    # Version history (stored as separate collection)
    version_history: List[SettingsVersionHistory] = Field(default_factory=list)

# Invoice Models
class InvoiceItem(BaseModel):
    article_id: str
    article_name: str
    quantity: int
    unit_price: float
    total_price: float

class Invoice(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_number: str  # Auto-generated
    event_id: Optional[str] = None
    customer_id: str
    items: List[InvoiceItem]
    subtotal: float
    tax_rate: float = 19.0  # 19% MwSt
    tax_amount: float
    total_amount: float
    notes: Optional[str] = None
    status: str = "draft"  # draft, sent, paid, cancelled
    payment_status: str = "offen"  # offen, teilweise, bezahlt, überfällig
    issue_date: datetime = Field(default_factory=datetime.utcnow)
    due_date: Optional[datetime] = None
    paid_date: Optional[datetime] = None
    created_by: str
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)
    public_token: Optional[str] = None

class InvoiceCreate(BaseModel):
    event_id: Optional[str] = None
    customer_id: Optional[str] = None
    notes: Optional[str] = None
    due_days: int = Field(14, ge=0, le=365)  # max 1 year payment term

# Time Tracking Models
class TimeEntryCreate(BaseModel):
    crew_member_id: str
    event_id: Optional[str] = None
    date: str                       # "2026-03-18"
    start_time: str                 # "08:00"
    end_time: str                   # "17:00"
    break_minutes: int = Field(0, ge=0, le=480)  # max 8h break
    activity: Optional[str] = None
    notes: Optional[str] = None

class TimeEntry(TimeEntryCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    hours_worked: float = 0
    hourly_rate: float = 0
    total_pay: float = 0
    created_at: datetime = Field(default_factory=datetime.utcnow)

# Quote Models
class QuoteItem(BaseModel):
    article_id: Optional[str] = None
    article_name: str
    quantity: int = Field(1, ge=1, le=10000)
    unit_price: float = Field(0, ge=0)
    days: int = Field(1, ge=1, le=3650)  # max 10 years
    total: float = 0

class QuoteCreate(BaseModel):
    customer_id: Optional[str] = None
    customer_name: str
    event_name: str
    event_date: Optional[str] = None
    valid_until: Optional[str] = None
    items: List[QuoteItem] = []
    notes: Optional[str] = None
    discount_percent: float = Field(0, ge=0, le=100)  # max 100%

class Quote(QuoteCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    quote_number: str = ""
    status: str = "entwurf"
    total_net: float = 0
    created_at: datetime = Field(default_factory=datetime.utcnow)
    created_by: str = ""
    updated_at: datetime = Field(default_factory=datetime.utcnow)

# Bill of Materials Models
class BOMItem(BaseModel):
    article_id: str
    quantity: int
    is_optional: bool = False

class BillOfMaterials(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    category: Optional[str] = None  # Konzert, Messe, etc.
    items: List[BOMItem]
    package_price: Optional[float] = None  # Optional package discount price
    image_base64: Optional[str] = None
    is_active: bool = True
    created_by: str
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class BOMCreate(BaseModel):
    name: str
    description: Optional[str] = None
    category: Optional[str] = None
    items: List[BOMItem]
    package_price: Optional[float] = None
    image_base64: Optional[str] = None

class ProjectTemplateCreate(BaseModel):
    name: str
    description: Optional[str] = None
    event_type: Optional[str] = None
    location_type: Optional[str] = None
    notes_template: Optional[str] = None
    bom_id: Optional[str] = None

class ProjectTemplate(ProjectTemplateCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_by: str = ""
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class CustomFieldDefCreate(BaseModel):
    entity_type: str          # event | quote | customer | article
    field_label: str
    field_name: str
    field_type: str           # text | number | date | checkbox | select
    options: List[str] = []
    required: bool = False
    sort_order: int = 0

class CustomFieldDef(CustomFieldDefCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)

class EventInvitationCreate(BaseModel):
    event_id: str
    email: str
    name: str = ""

class EventInvitation(EventInvitationCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    status: str = "pending"   # pending | accepted | declined
    sent_at: datetime = Field(default_factory=datetime.utcnow)
    responded_at: Optional[datetime] = None

class WebhookCreate(BaseModel):
    url: str
    events: List[str] = []   # e.g. event_created, booking_confirmed, invoice_created
    active: bool = True
    secret: str = ""

class Webhook(WebhookCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class RentalCalculationRequest(BaseModel):
    article_ids: List[str]
    quantities: Dict[str, int] = {}

# Auth helper functions
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def parse_iso_date(value: str, field_name: str = "Datum") -> datetime:
    """Parse ISO-8601 date string and raise a clean 400 on invalid format."""
    try:
        return datetime.fromisoformat(value.replace('Z', '+00:00'))
    except (ValueError, AttributeError):
        raise HTTPException(
            status_code=400,
            detail=f"Ungültiges {field_name}-Format. Erwartet: ISO-8601 (z.B. 2024-01-31T00:00:00Z)"
        )

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire, "type": "access"})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def create_refresh_token(data: dict) -> str:
    """Create a long-lived refresh token"""
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)
    to_encode.update({"exp": expire, "type": "refresh"})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=["HS256"])  # M1: hardcoded, no algorithm confusion
        username: str = payload.get("sub")
        # No default — missing "type" field must also be rejected
        token_type: str = payload.get("type")

        # Only allow access tokens for regular API calls
        if token_type != "access":
            raise credentials_exception
            
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    user = await db.users.find_one({"username": username})
    if user is None:
        raise credentials_exception
    return User(**user)

def _sanitize_email_field(text: str) -> str:
    """Remove SMTP header injection characters (CR/LF) from user-controlled data."""
    return text.replace('\r', '').replace('\n', ' ').strip() if text else text

async def send_maintenance_notification(article_name: str, maintenance_date: datetime, is_overdue: bool = False):
    """Send maintenance notification email"""
    try:
        import aiosmtplib
        from email.message import EmailMessage

        if not SMTP_USERNAME or not SMTP_PASSWORD:
            logging.warning("SMTP credentials not configured, skipping email")
            return False

        safe_article_name = _sanitize_email_field(article_name)

        message = EmailMessage()
        message["From"] = SMTP_USERNAME
        message["To"] = NOTIFICATION_EMAIL

        if is_overdue:
            message["Subject"] = f"UEBERFAELLIG: Wartung fuer {safe_article_name}"
            body = f"""
Überfällige Wartung!

Artikel: {safe_article_name}
Fällig war: {maintenance_date.strftime('%d.%m.%Y')}
Status: ÜBERFÄLLIG

Bitte führen Sie die Wartung umgehend durch.

Mit freundlichen Grüßen
Ihr Inventar Management System
            """
        else:
            message["Subject"] = f"Wartungserinnerung: {safe_article_name}"
            body = f"""
Wartungserinnerung

Artikel: {safe_article_name}
Fällig am: {maintenance_date.strftime('%d.%m.%Y')}

Bitte planen Sie die Wartung rechtzeitig ein.

Mit freundlichen Grüßen
Ihr Inventar Management System
            """
        
        message.set_content(body)
        
        await aiosmtplib.send(
            message,
            hostname=SMTP_SERVER,
            port=SMTP_PORT,
            username=SMTP_USERNAME,
            password=SMTP_PASSWORD,
            start_tls=True
        )
        
        logging.info(f"Maintenance notification sent to {NOTIFICATION_EMAIL} for {article_name}")
        return True

    except aiosmtplib.SMTPException as e:
        logging.error(f"SMTP-Fehler beim Senden der Wartungsbenachrichtigung: {e}")
        return False
    except OSError as e:
        logging.error(f"Netzwerkfehler beim Senden der Wartungsbenachrichtigung: {e}")
        return False

async def send_user_registration_email(username: str, email: str, role: str):
    """Send email to admin when new user registers"""
    safe_username = _sanitize_email_field(username)
    safe_email = _sanitize_email_field(email)
    safe_role = _sanitize_email_field(role)
    email_body = f"""
Neue Benutzerregistrierung

Ein neuer Benutzer hat sich registriert und wartet auf Ihre Genehmigung:

Benutzername: {safe_username}
E-Mail: {safe_email}
Rolle: {safe_role}
Registriert am: {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')}

Bitte melden Sie sich im Admin-Panel an, um diesen Benutzer zu genehmigen oder abzulehnen.
    """
    # M2: Send actual email via SMTP (same pattern as send_maintenance_notification)
    if not SMTP_USERNAME or not SMTP_PASSWORD or not ADMIN_EMAIL:
        logging.warning("SMTP not configured — registration notification only logged.")
        logging.info(f"[REGISTRATION] {safe_username} ({safe_email}) role={safe_role}")
        return True
    try:
        import aiosmtplib
        from email.message import EmailMessage
        msg = EmailMessage()
        msg["From"] = SMTP_USERNAME
        msg["To"] = ADMIN_EMAIL
        msg["Subject"] = f"Neue Benutzerregistrierung: {safe_username}"
        msg.set_content(email_body)
        await aiosmtplib.send(
            msg,
            hostname=SMTP_SERVER,
            port=SMTP_PORT,
            username=SMTP_USERNAME,
            password=SMTP_PASSWORD,
            start_tls=True,
        )
        logging.info(f"Registration notification sent to {ADMIN_EMAIL} for {safe_username}")
        return True
    except Exception as e:
        logging.error(f"Failed to send registration email: {e}")
        return False

async def ensure_admin_user():
    """Ensure the fixed admin user exists"""
    admin_username = "Admin"
    admin_email = "admin@inventory.system"

    admin_password = os.environ.get("ADMIN_PASSWORD")
    _weak_defaults = {"changeme", "admin", "password", "123456", ""}
    if not admin_password or admin_password in _weak_defaults:
        raise RuntimeError(
            "⛔ CRITICAL: ADMIN_PASSWORD ist nicht gesetzt oder zu schwach.\n"
            "Bitte einen sicheren Wert in der .env setzen:\n"
            "  ADMIN_PASSWORD=<mindestens 12 Zeichen>\n"
            "Server wird nicht gestartet."
        )

    admin_user = await db.users.find_one({"username": admin_username})
    if not admin_user:
        hashed_password = get_password_hash(admin_password)
        admin_user_data = {
            "id": str(uuid.uuid4()),
            "username": admin_username,
            "email": admin_email,
            "role": "admin",
            "is_active": True,
            "is_approved": True,  # Admin is always approved
            "hashed_password": hashed_password,
            "created_at": datetime.now(timezone.utc)
        }
        await db.users.insert_one(admin_user_data)
        logging.info(f"Created fixed admin user: {admin_username}")
    else:
        # Ensure admin is always approved
        if not admin_user.get("is_approved", False):
            await db.users.update_one(
                {"username": admin_username},
                {"$set": {"is_approved": True}}
            )
            logging.info(f"Ensured admin user is approved: {admin_username}")

async def generate_customer_number():
    """Generate unique customer number"""
    # Get count of customers
    count = await db.customers.count_documents({})
    return f"CUST-{datetime.now(timezone.utc).year}-{str(count + 1).zfill(4)}"

async def generate_event_number():
    """Generate unique event number: EVT-YYYY-NNN"""
    year = datetime.now(timezone.utc).year
    # Count events in current year
    count = await db.events.count_documents({
        "event_number": {"$regex": f"^EVT-{year}"}
    })
    return f"EVT-{year}-{str(count + 1).zfill(3)}"

async def check_booking_conflict(article_id: str, start_date: datetime, end_date: datetime, exclude_booking_id: Optional[str] = None):
    """Check if article is already booked for the given date range"""
    query = {
        "article_id": article_id,
        "status": {"$in": ["booked"]},
        "$or": [
            # New booking starts during existing booking
            {"pickup_date": {"$lte": start_date}, "return_date": {"$gte": start_date}},
            # New booking ends during existing booking
            {"pickup_date": {"$lte": end_date}, "return_date": {"$gte": end_date}},
            # New booking completely contains existing booking
            {"pickup_date": {"$gte": start_date}, "return_date": {"$lte": end_date}},
        ]
    }
    
    if exclude_booking_id:
        query["id"] = {"$ne": exclude_booking_id}
    
    conflicts = await db.bookings.find(query).to_list(10)
    return conflicts

async def generate_invoice_number():
    """Generate unique invoice number: INV-YYYY-NNNN"""
    year = datetime.now(timezone.utc).year
    count = await db.invoices.count_documents({
        "invoice_number": {"$regex": f"^INV-{year}"}
    })
    return f"INV-{year}-{str(count + 1).zfill(4)}"

async def send_booking_notification_email(customer_email: str, event_name: str, action: str):
    """Send booking notification to customer"""
    safe_event_name = _sanitize_email_field(event_name)
    safe_action = _sanitize_email_field(action)
    safe_customer_email = _sanitize_email_field(customer_email)
    email_body = f"""
Buchungsbenachrichtigung

Aktion: {safe_action}
Veranstaltung: {safe_event_name}
E-Mail: {safe_customer_email}
Datum: {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')}
    """
    logging.info(f"Booking notification to {safe_customer_email}: {email_body}")
    return True

# Authentication routes
@api_router.post("/register")
@limiter.limit("3/minute")  # SECURITY: Limit registration to prevent spam accounts
async def register(request: Request, user_data: UserCreate):
    # SECURITY: Validate password strength
    if len(user_data.password) < 8:
        raise HTTPException(status_code=400, detail="Passwort muss mindestens 8 Zeichen lang sein")
    
    # Check if user exists
    existing_user = await db.users.find_one({"$or": [{"email": user_data.email}, {"username": user_data.username}]})
    if existing_user:
        raise HTTPException(status_code=400, detail="User already exists")
    
    # Hash password and create user
    hashed_password = get_password_hash(user_data.password)
    user_dict = user_data.dict()
    del user_dict['password']
    
    # New users are not approved by default and cannot login until approved
    user = User(**user_dict, is_approved=False)
    user_with_password = {**user.dict(), "hashed_password": hashed_password}
    
    await db.users.insert_one(user_with_password)
    
    # Log registration
    logging.info(f"New user registered: {user.username} from IP: {request.client.host}")
    
    # Send email notification to admin
    await send_user_registration_email(user.username, user.email, user.role)
    
    return {
        "message": "Registrierung erfolgreich. Ihr Konto wartet auf die Genehmigung des Administrators.",
        "username": user.username,
        "email": user.email,
        "is_approved": False
    }

@api_router.post("/login", response_model=Token)
@limiter.limit("10/minute")  # SECURITY: Rate limit login to 10 attempts per minute
async def login(request: Request, user_credentials: UserLogin):
    user = await db.users.find_one({"username": user_credentials.username})
    if not user or not verify_password(user_credentials.password, user["hashed_password"]):
        # SECURITY: Log failed login attempts
        logging.warning(f"Failed login attempt for user: {user_credentials.username} from IP: {request.client.host}")
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Check if user is approved
    if not user.get("is_approved", False):
        raise HTTPException(
            status_code=403, 
            detail="Ihr Konto wartet noch auf die Genehmigung des Administrators. Bitte versuchen Sie es später erneut."
        )
    
    # Create access token (short-lived)
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user["username"]}, expires_delta=access_token_expires
    )
    
    # Create refresh token (long-lived)
    refresh_token = create_refresh_token(data={"sub": user["username"]})
    
    # Store refresh token in database for validation
    await db.refresh_tokens.insert_one({
        "token": refresh_token,
        "user_id": user["id"],
        "created_at": datetime.now(timezone.utc),
        "expires_at": datetime.now(timezone.utc) + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS),
        "is_revoked": False
    })
    
    # Log successful login
    logging.info(f"Successful login for user: {user_credentials.username}")
    
    user_obj = User(**user)
    return Token(
        access_token=access_token, 
        refresh_token=refresh_token,
        token_type="bearer", 
        user=user_obj,
        expires_in=ACCESS_TOKEN_EXPIRE_MINUTES * 60  # in seconds
    )

@api_router.post("/refresh-token")
@limiter.limit("10/minute")
async def refresh_access_token(request: Request, token_data: RefreshTokenRequest):
    """V6: Refresh-token rotation — consume old token, issue new access + refresh pair."""
    try:
        payload = jwt.decode(token_data.refresh_token, SECRET_KEY, algorithms=["HS256"])  # M1: hardcoded
        username = payload.get("sub")
        token_type = payload.get("type")

        if token_type != "refresh" or not username:
            raise HTTPException(status_code=401, detail="Invalid refresh token")

        # Atomically delete the used token (one-time use).
        # If already consumed (reuse attack) delete_result.deleted_count == 0.
        delete_result = await db.refresh_tokens.delete_one({
            "token": token_data.refresh_token,
            "is_revoked": False
        })

        # Get user (needed both for reuse-attack cleanup and normal flow)
        user = await db.users.find_one({"username": username})

        if delete_result.deleted_count == 0:
            # Possible token-reuse attack: revoke ALL tokens for this user
            if user:
                await db.refresh_tokens.delete_many({"user_id": user["id"]})
            logging.warning("Refresh-token reuse detected for user: %s", username)
            raise HTTPException(status_code=401, detail="Refresh token already used or revoked")
        if not user:
            raise HTTPException(status_code=401, detail="User not found")

        # Issue new access token
        access_token = create_access_token(
            data={"sub": username},
            expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        )

        # Issue new refresh token (rotation)
        new_refresh_token = create_refresh_token(data={"sub": username})
        await db.refresh_tokens.insert_one({
            "token": new_refresh_token,
            "user_id": user["id"],
            "created_at": datetime.now(timezone.utc),
            "expires_at": datetime.now(timezone.utc) + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS),
            "is_revoked": False
        })

        return {
            "access_token": access_token,
            "refresh_token": new_refresh_token,
            "token_type": "bearer",
            "expires_in": ACCESS_TOKEN_EXPIRE_MINUTES * 60
        }

    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

@api_router.post("/logout")
async def logout(current_user: User = Depends(get_current_user)):
    """Revoke all refresh tokens for current user"""
    await db.refresh_tokens.update_many(
        {"user_id": current_user.id},
        {"$set": {"is_revoked": True}}
    )
    return {"message": "Erfolgreich abgemeldet"}

@api_router.get("/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

class ProfileUpdate(BaseModel):
    username: Optional[str] = Field(None, min_length=2, max_length=50)
    email: Optional[EmailStr] = None
    current_password: Optional[str] = Field(None, min_length=8, max_length=128)
    new_password: Optional[str] = Field(None, min_length=8, max_length=128)
    profile_image: Optional[str] = None  # Base64 encoded image

@api_router.put("/me", response_model=User)
async def update_profile(
    update_data: ProfileUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update current user's profile (username, email, password, profile image)."""
    update_fields = {}

    # Handle username update
    if update_data.username and update_data.username != current_user.username:
        existing = await db.users.find_one({"username": update_data.username, "id": {"$ne": current_user.id}})
        if existing:
            raise HTTPException(status_code=400, detail="Benutzername bereits vergeben")
        update_fields["username"] = update_data.username

    # Handle email update
    if update_data.email and update_data.email != current_user.email:
        existing = await db.users.find_one({"email": update_data.email, "id": {"$ne": current_user.id}})
        if existing:
            raise HTTPException(status_code=400, detail="E-Mail bereits registriert")
        update_fields["email"] = update_data.email

    # Handle password change
    if update_data.new_password:
        if not update_data.current_password:
            raise HTTPException(status_code=400, detail="Aktuelles Passwort erforderlich")
        if not verify_password(update_data.current_password, current_user.hashed_password):
            raise HTTPException(status_code=400, detail="Aktuelles Passwort ist falsch")
        update_fields["hashed_password"] = get_password_hash(update_data.new_password)

    # Handle profile image
    if update_data.profile_image is not None:
        # Validate base64 image (max ~5MB encoded = ~3.75MB decoded)
        if len(update_data.profile_image) > 5_000_000:
            raise HTTPException(status_code=400, detail="Bild zu groß (max. 5MB)")
        update_fields["profile_image"] = update_data.profile_image

    if update_fields:
        await db.users.update_one({"id": current_user.id}, {"$set": update_fields})
        updated_user = await db.users.find_one({"id": current_user.id})
        return User(**updated_user)

    return current_user


# ─── F2: Password Reset ───────────────────────────────────────────────────────

class ForgotPasswordRequest(BaseModel):
    email: EmailStr

class ResetPasswordRequest(BaseModel):
    token: str = Field(..., min_length=10, max_length=200)
    new_password: str = Field(..., min_length=8, max_length=128)

@api_router.post("/auth/forgot-password")
@limiter.limit("3/hour")
async def forgot_password(request: Request, body: ForgotPasswordRequest):
    """Send password-reset link via e-mail. Always returns 200 (prevents user enumeration)."""
    import secrets as _secrets
    user = await db.users.find_one({"email": body.email})
    if user:
        token = _secrets.token_urlsafe(32)
        expires_at = datetime.now(timezone.utc) + timedelta(hours=1)
        await db.password_reset_tokens.insert_one({
            "token": token,
            "user_id": user["id"],
            "expires_at": expires_at,
            "used": False,
        })
        reset_url = f"{os.getenv('FRONTEND_URL', 'http://localhost:8081')}/reset-password?token={token}"
        body_text = (
            f"Hallo {user['username']},\n\n"
            f"Sie haben ein Passwort-Reset angefordert. Bitte nutzen Sie folgenden Link "
            f"(gültig 1 Stunde):\n\n{reset_url}\n\n"
            f"Falls Sie das nicht angefordert haben, ignorieren Sie diese E-Mail.\n\n"
            f"Mit freundlichen Grüßen\nIhr InventarPro-Team"
        )
        if SMTP_USERNAME and SMTP_PASSWORD:
            try:
                import aiosmtplib
                from email.message import EmailMessage as _EM
                msg = _EM()
                msg["From"] = SMTP_USERNAME
                msg["To"] = body.email
                msg["Subject"] = "Passwort zurücksetzen — InventarPro"
                msg.set_content(body_text)
                await aiosmtplib.send(msg, hostname=SMTP_SERVER, port=SMTP_PORT,
                                      username=SMTP_USERNAME, password=SMTP_PASSWORD, start_tls=True)
            except Exception as e:
                logging.error("Passwort-Reset E-Mail konnte nicht gesendet werden: %s", e)
        else:
            logging.info("[RESET] Token für %s: %s", body.email, token)
    return {"message": "Falls ein Konto mit dieser E-Mail existiert, wurde ein Reset-Link gesendet."}

@api_router.post("/auth/reset-password")
async def reset_password(body: ResetPasswordRequest):
    """Consume a password-reset token and update the user's password."""
    record = await db.password_reset_tokens.find_one({"token": body.token, "used": False})
    if not record:
        raise HTTPException(status_code=400, detail="Token ungültig oder bereits verwendet")
    if record["expires_at"] < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Token abgelaufen. Bitte erneut anfordern.")
    hashed = get_password_hash(body.new_password)
    await db.users.update_one({"id": record["user_id"]}, {"$set": {"hashed_password": hashed}})
    await db.password_reset_tokens.update_one({"_id": record["_id"]}, {"$set": {"used": True}})
    # Invalidate all active sessions for security
    await db.refresh_tokens.delete_many({"user_id": record["user_id"]})
    return {"message": "Passwort erfolgreich geändert. Bitte melden Sie sich neu an."}

# ─── End F2 ───────────────────────────────────────────────────────────────────

# Categories
@api_router.post("/categories", response_model=Category)
async def create_category(category: Category, current_user: User = Depends(get_current_user)):
    await db.categories.insert_one(category.dict())
    return category

@api_router.get("/categories", response_model=List[Category])
async def get_categories(current_user: User = Depends(get_current_user)):
    categories = await db.categories.find().to_list(1000)
    return [Category(**cat) for cat in categories]

@api_router.get("/categories/{category_id}", response_model=Category)
async def get_category(category_id: str, current_user: User = Depends(get_current_user)):
    category = await db.categories.find_one({"id": category_id})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    return Category(**category)

@api_router.put("/categories/{category_id}", response_model=Category)
async def update_category(
    category_id: str,
    category_data: Category,
    current_user: User = Depends(get_current_user)
):
    category_dict = category_data.dict()
    category_dict.pop('id', None)  # Don't update ID
    
    result = await db.categories.update_one(
        {"id": category_id},
        {"$set": category_dict}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    
    updated_category = await db.categories.find_one({"id": category_id})
    return Category(**updated_category)

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, current_user: User = Depends(get_current_user)):
    # Check if category is used by articles
    articles_using_category = await db.articles.count_documents({"category_id": category_id})
    if articles_using_category > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Category is used by {articles_using_category} article(s). Please reassign or delete them first."
        )
    
    result = await db.categories.delete_one({"id": category_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    return {"message": "Category deleted successfully"}

# Suppliers
@api_router.post("/suppliers", response_model=Supplier)
async def create_supplier(supplier_data: SupplierCreate, current_user: User = Depends(get_current_user)):
    supplier = Supplier(**supplier_data.dict())
    await db.suppliers.insert_one(supplier.dict())
    return supplier

@api_router.get("/suppliers", response_model=List[Supplier])
async def get_suppliers(current_user: User = Depends(get_current_user)):
    suppliers = await db.suppliers.find().to_list(1000)
    return [Supplier(**sup) for sup in suppliers]

@api_router.get("/suppliers/{supplier_id}", response_model=Supplier)
async def get_supplier(supplier_id: str, current_user: User = Depends(get_current_user)):
    supplier = await db.suppliers.find_one({"id": supplier_id})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return Supplier(**supplier)

@api_router.put("/suppliers/{supplier_id}", response_model=Supplier)
async def update_supplier(
    supplier_id: str,
    supplier_data: SupplierCreate,
    current_user: User = Depends(get_current_user)
):
    supplier_dict = supplier_data.dict()
    
    result = await db.suppliers.update_one(
        {"id": supplier_id},
        {"$set": supplier_dict}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    updated_supplier = await db.suppliers.find_one({"id": supplier_id})
    return Supplier(**updated_supplier)

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str, current_user: User = Depends(get_current_user)):
    result = await db.suppliers.delete_one({"id": supplier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return {"message": "Supplier deleted successfully"}

# Teams
@api_router.post("/teams", response_model=Team)
async def create_team(team_data: TeamCreate, current_user: User = Depends(require_permission(Permission.MANAGE_USERS))):
    
    team = Team(**team_data.dict(), created_by=current_user.username)
    await db.teams.insert_one(team.dict())
    return team

@api_router.get("/teams", response_model=List[Team])
async def get_teams(current_user: User = Depends(get_current_user)):
    teams = await db.teams.find().to_list(1000)
    return [Team(**team) for team in teams]

@api_router.get("/teams/{team_id}", response_model=Team)
async def get_team(team_id: str, current_user: User = Depends(get_current_user)):
    team = await db.teams.find_one({"id": team_id})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    return Team(**team)

@api_router.put("/teams/{team_id}", response_model=Team)
async def update_team(
    team_id: str,
    team_data: TeamCreate,
    current_user: User = Depends(require_permission(Permission.MANAGE_USERS))
):
    
    team_dict = team_data.dict()
    result = await db.teams.update_one(
        {"id": team_id},
        {"$set": team_dict}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Team not found")
    
    updated_team = await db.teams.find_one({"id": team_id})
    return Team(**updated_team)

@api_router.delete("/teams/{team_id}")
async def delete_team(team_id: str, current_user: User = Depends(require_permission(Permission.MANAGE_USERS))):
    
    result = await db.teams.delete_one({"id": team_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Team not found")
    return {"message": "Team deleted successfully"}

@api_router.post("/teams/{team_id}/members/{user_id}")
async def add_team_member(team_id: str, user_id: str, current_user: User = Depends(require_permission(Permission.MANAGE_USERS))):
    
    # Check if user exists
    user = await db.users.find_one({"username": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Add user to team
    result = await db.teams.update_one(
        {"id": team_id},
        {"$addToSet": {"members": user_id}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Team not found")
    
    return {"message": "Member added successfully"}

@api_router.delete("/teams/{team_id}/members/{user_id}")
async def remove_team_member(team_id: str, user_id: str, current_user: User = Depends(require_permission(Permission.MANAGE_USERS))):
    
    result = await db.teams.update_one(
        {"id": team_id},
        {"$pull": {"members": user_id}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Team not found")
    
    return {"message": "Member removed successfully"}

# Audit Logs
_SENSITIVE_AUDIT_FIELDS = {'password', 'hashed_password', 'token', 'secret', 'api_key', 'refresh_token'}

def _sanitize_audit_changes(changes: dict) -> dict:
    """Redact sensitive field values from audit log change dicts."""
    if not changes:
        return changes
    return {
        k: "[REDACTED]" if any(s in k.lower() for s in _SENSITIVE_AUDIT_FIELDS) else v
        for k, v in changes.items()
    }

async def create_audit_log(
    action: str,
    entity_type: str,
    user: User,
    entity_id: str = None,
    entity_name: str = None,
    changes: dict = None
):
    """Helper function to create audit log entries"""
    log = AuditLog(
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        entity_name=entity_name,
        user_id=user.username,
        user_name=user.username,
        changes=_sanitize_audit_changes(changes),
    )
    await db.audit_logs.insert_one(log.dict())

@api_router.get("/audit-logs", response_model=List[AuditLog])
async def get_audit_logs(
    limit: int = 100,
    skip: int = 0,
    current_user: User = Depends(get_current_user)
):
    """Get audit logs (admin only sees all, users see their own)"""
    query = {} if current_user.role == "admin" else {"user_id": current_user.username}
    
    logs = await db.audit_logs.find(query).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)
    return [AuditLog(**log) for log in logs]

# Notifications
@api_router.post("/notifications/register")
async def register_notification_token(
    token_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Register push notification token for user"""
    await db.users.update_one(
        {"username": current_user.username},
        {"$set": {"push_token": token_data.get("token")}}
    )
    return {"message": "Token registered successfully"}

@api_router.get("/notifications/pending")
async def get_pending_notifications(current_user: User = Depends(get_current_user)):
    """Get pending maintenance reminders and unread messages"""
    # Get maintenance tasks due in next 7 days
    upcoming_tasks = await db.maintenance_tasks.find({
        "due_date": {
            "$gte": datetime.now(timezone.utc).isoformat(),
            "$lte": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()
        },
        "status": {"$ne": "completed"}
    }).to_list(100)
    
    # Count unread messages (simplified - in real app would track read status)
    message_count = await db.messages.count_documents({})
    
    return {
        "maintenance_reminders": len(upcoming_tasks),
        "unread_messages": message_count,
        "tasks": upcoming_tasks[:5]  # Return top 5
    }

@api_router.post("/sync/now")
async def sync_now(current_user: User = Depends(get_current_user)):
    """Trigger manual sync"""
    # Update last sync time
    await db.users.update_one(
        {"username": current_user.username},
        {"$set": {"last_sync_time": datetime.now(timezone.utc)}}
    )
    
    return {
        "message": "Sync completed",
        "timestamp": datetime.now(timezone.utc)
    }

# Articles
@api_router.post("/articles", response_model=Article)
async def create_article(article_data: ArticleCreate, current_user: User = Depends(require_permission(Permission.CREATE_ARTICLE))):
    # Auto-generate inventory_code if not provided
    if not article_data.inventory_code:
        short_id = str(uuid.uuid4())[:6].upper()
        article_data.inventory_code = f"ART-{short_id}"

    # Check for duplicate inventory_code
    existing_code = await db.articles.find_one({"inventory_code": article_data.inventory_code})
    if existing_code:
        raise HTTPException(status_code=409, detail=f"Inventarnummer '{article_data.inventory_code}' bereits vorhanden")

    # Use inventory_code directly as QR code (it already has the prefix)
    qr_code = article_data.inventory_code

    article = Article(**article_data.dict(), qr_code=qr_code)
    await db.articles.insert_one(article.dict())
    await manager.broadcast(json.dumps({"type": "article_created", "id": str(article.id)}))
    return article

@api_router.get("/articles", response_model=List[Article])
async def get_articles(
    category_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query: dict = {"deleted": {"$ne": True}}  # V7: exclude soft-deleted
    if category_id:
        query["category_id"] = category_id
    if search:
        safe_search = re.escape(search)
        query["$or"] = [
            {"name": {"$regex": safe_search, "$options": "i"}},
            {"description": {"$regex": safe_search, "$options": "i"}},
            {"inventory_code": {"$regex": safe_search, "$options": "i"}}
        ]

    articles = await db.articles.find(query).to_list(1000)
    return [Article(**art) for art in articles]

# Paginated articles endpoint for large datasets
@api_router.get("/articles/paginated/list")
async def get_articles_paginated(
    page: int = 1,
    page_size: int = 20,
    category_id: Optional[str] = None,
    search: Optional[str] = None,
    status: Optional[str] = None,
    sort_by: str = "name",
    sort_order: str = "asc",
    current_user: User = Depends(get_current_user)
):
    """Get paginated articles list for infinite scroll"""
    query: dict = {"deleted": {"$ne": True}}  # V7: exclude soft-deleted

    if category_id:
        query["category_id"] = category_id
    if status:
        query["status"] = status
    if search:
        safe_search = re.escape(search)
        query["$or"] = [
            {"name": {"$regex": safe_search, "$options": "i"}},
            {"inventory_code": {"$regex": safe_search, "$options": "i"}},
            {"description": {"$regex": safe_search, "$options": "i"}}
        ]

    # Calculate skip
    skip = (page - 1) * page_size
    
    # Get total count
    total = await db.articles.count_documents(query)
    
    # Sort direction
    sort_dir = 1 if sort_order == "asc" else -1
    
    # Get paginated data
    articles = await db.articles.find(query).sort(sort_by, sort_dir).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "items": [Article(**art).dict() for art in articles],
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
        "has_next": skip + page_size < total,
        "has_prev": page > 1
    }

@api_router.get("/articles/operating-hours-alerts")
async def get_operating_hours_alerts(
    threshold_percent: int = 80,
    current_user: User = Depends(get_current_user)
):
    """Get articles approaching or exceeding max operating hours"""
    articles = await db.articles.find({
        "max_operating_hours": {"$gt": 0}
    }).to_list(1000)
    
    alerts = []
    for article in articles:
        current = article.get("operating_hours", 0) or 0
        max_hours = article.get("max_operating_hours", 0) or 0
        
        if max_hours > 0:
            percentage = (current / max_hours) * 100
            if percentage >= threshold_percent:
                alerts.append({
                    "article_id": article["id"],
                    "article_name": article["name"],
                    "inventory_code": article.get("inventory_code", ""),
                    "current_hours": current,
                    "max_hours": max_hours,
                    "percentage": round(percentage, 1),
                    "hours_remaining": max(0, max_hours - current),
                    "status": "critical" if percentage >= 100 else ("warning" if percentage >= 90 else "attention")
                })
    
    # Sort by percentage (highest first)
    alerts.sort(key=lambda x: x["percentage"], reverse=True)
    
    return {
        "total_alerts": len(alerts),
        "critical_count": len([a for a in alerts if a["status"] == "critical"]),
        "warning_count": len([a for a in alerts if a["status"] == "warning"]),
        "alerts": alerts
    }

@api_router.get("/articles/consumable-alerts")
async def get_consumable_alerts(current_user: User = Depends(get_current_user)):
    """Returns consumable articles where current_stock <= min_stock_level."""
    articles = await db.articles.find({
        "is_consumable": True,
        "$expr": {"$lte": ["$current_stock", "$min_stock_level"]}
    }).to_list(1000)
    for a in articles:
        a.pop("_id", None)  # Remove MongoDB ObjectId
    return articles

class ArticleImportRow(BaseModel):
    name: str
    inventory_code: str
    description: Optional[str] = None
    category_id: Optional[str] = None
    base_unit: str = "Stück"
    current_stock: int = 0
    min_stock_level: int = 0
    price_per_unit: float = 0.0
    status: str = "OK"
    is_consumable: bool = False

class ArticleImportRequest(BaseModel):
    csv_content: str = ""   # raw CSV text (csv mode)
    file_content: str = ""  # F10: base64-encoded file bytes (excel mode)
    file_type: str = "csv"  # F10: "csv" | "excel"

# ===== FAHRZEUGE =====
class VehicleCreate(BaseModel):
    name: str
    license_plate: str = ""
    brand: str = ""
    model_name: str = ""
    year: Optional[int] = None
    tuev_date: Optional[str] = None
    fuel_type: str = "Diesel"
    status: str = "verfügbar"
    notes: Optional[str] = None

class Vehicle(VehicleCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)

# ===== AUFGABEN =====
class TaskCreate(BaseModel):
    title: str
    description: Optional[str] = None
    assigned_to_id: Optional[str] = None
    assigned_to_name: Optional[str] = None
    priority: str = "normal"
    due_date: Optional[str] = None
    event_id: Optional[str] = None
    event_name: Optional[str] = None
    status: str = "offen"

class Task(TaskCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)
    created_by: str = ""

# ===== SERIENNUMMERN =====
class SerialNumberCreate(BaseModel):
    article_id: str
    article_name: str = ""
    serial_number: str
    status: str = "verfügbar"
    condition: str = "gut"
    purchase_date: Optional[str] = None
    purchase_price: Optional[float] = None
    notes: Optional[str] = None

class SerialNumber(SerialNumberCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)

# ===== ABWESENHEITSANTRÄGE =====
class AbsenceRequestCreate(BaseModel):
    crew_member_id: str
    crew_member_name: str = ""
    start_date: str
    end_date: str
    type: str = "urlaub"
    reason: Optional[str] = None
    status: str = "ausstehend"

class AbsenceRequest(AbsenceRequestCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)

# ===== BESTANDSZÄHLUNGEN =====
class StockCountItem(BaseModel):
    article_id: str
    article_name: str
    expected_quantity: int = 0
    counted_quantity: int = 0

class StockCountCreate(BaseModel):
    name: str
    notes: Optional[str] = None
    items: List[StockCountItem] = []

class StockCount(StockCountCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    status: str = "offen"
    created_at: datetime = Field(default_factory=datetime.utcnow)
    created_by: str = ""

# ===== PRÜFUNGEN =====
class InspectionCreate(BaseModel):
    article_id: Optional[str] = None
    article_name: str = ""
    inspection_type: str = "Sicherheitsprüfung"
    due_date: str
    performed_date: Optional[str] = None
    performed_by: Optional[str] = None
    result: str = "ausstehend"
    next_due_date: Optional[str] = None
    cost: Optional[float] = None
    notes: Optional[str] = None

class Inspection(InspectionCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)

@api_router.post("/articles/import")
async def import_articles(
    request: ArticleImportRequest,
    current_user: User = Depends(get_current_user)
):
    """Import articles from CSV or Excel. Returns { imported: int, errors: List[str] }"""
    def _csv_safe(value: str) -> str:
        """H3: Prevent CSV/Excel formula injection by prefixing dangerous chars."""
        v = (value or "").strip()
        if v and v[0] in ('=', '+', '-', '@', '\t', '\r', '|', '%'):
            return "'" + v
        return v

    # F10: Build rows list from CSV or Excel
    rows: list[dict] = []
    if request.file_type == "excel":
        try:
            import base64
            import openpyxl
            raw_bytes = base64.b64decode(request.file_content)
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
            ws = wb.active
            header = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for xl_row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({header[j]: (str(v) if v is not None else "") for j, v in enumerate(xl_row)})
        except ImportError:
            raise HTTPException(status_code=501, detail="openpyxl nicht installiert — Excel-Import nicht verfügbar")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Excel-Datei konnte nicht gelesen werden: {e}")
    else:
        rows = list(csv.DictReader(io.StringIO(request.csv_content)))

    if len(rows) > MAX_IMPORT_ROWS:
        raise HTTPException(status_code=400, detail=f"Zu viele Zeilen — maximal {MAX_IMPORT_ROWS} erlaubt")

    imported = 0
    errors = []
    for i, row in enumerate(rows):
        try:
            # Validate required fields
            if not row.get("name") or not row.get("inventory_code"):
                errors.append(f"Zeile {i+2}: name und inventory_code sind Pflichtfelder")
                continue
            # Check duplicate inventory_code
            existing = await db.articles.find_one({"inventory_code": row["inventory_code"]})
            if existing:
                errors.append(f"Zeile {i+2}: Inventarnummer '{row['inventory_code']}' bereits vorhanden")
                continue
            # Validated numeric fields — reject negative or non-numeric values
            try:
                current_stock = int(row.get("current_stock", 0) or 0)
                if current_stock < 0:
                    errors.append(f"Zeile {i+2}: Bestand darf nicht negativ sein ({current_stock})")
                    continue
            except (ValueError, TypeError):
                errors.append(f"Zeile {i+2}: Bestand muss eine ganze Zahl sein")
                continue

            try:
                min_stock_level = int(row.get("min_stock_level", 0) or 0)
                if min_stock_level < 0:
                    errors.append(f"Zeile {i+2}: Mindestbestand darf nicht negativ sein")
                    continue
            except (ValueError, TypeError):
                errors.append(f"Zeile {i+2}: Mindestbestand muss eine ganze Zahl sein")
                continue

            try:
                price_per_unit = float(row.get("price_per_unit", 0) or 0)
                if price_per_unit < 0:
                    errors.append(f"Zeile {i+2}: Preis darf nicht negativ sein")
                    continue
            except (ValueError, TypeError):
                errors.append(f"Zeile {i+2}: Preis muss eine Zahl sein")
                continue

            valid_statuses = {"OK", "defekt", "gesperrt", "in_reparatur"}
            status = row.get("status", "OK").strip() or "OK"
            if status not in valid_statuses:
                errors.append(f"Zeile {i+2}: Ungültiger Status '{status}'. Erlaubt: {', '.join(valid_statuses)}")
                continue

            article_doc = {
                "id": str(uuid.uuid4()),
                "name": _csv_safe(row["name"]),
                "inventory_code": row["inventory_code"].strip(),  # code must stay unchanged
                "description": _csv_safe(row.get("description", "")) or None,
                "category_id": row.get("category_id", "").strip() or None,
                "base_unit": _csv_safe(row.get("base_unit", "Stück")) or "Stück",
                "current_stock": current_stock,
                "min_stock_level": min_stock_level,
                "price_per_unit": price_per_unit,
                "status": status,
                "is_consumable": row.get("is_consumable", "").strip().lower() in ("true", "1", "ja"),
                "images": [],
                "created_at": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc),
            }
            await db.articles.insert_one(article_doc)
            imported += 1
        except Exception as e:
            errors.append(f"Zeile {i+2}: {str(e)}")
    return {"imported": imported, "errors": errors}

@api_router.get("/articles/{article_id}", response_model=Article)
async def get_article(article_id: str, current_user: User = Depends(get_current_user)):
    article = await db.articles.find_one({"id": article_id})
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
    return Article(**article)

@api_router.put("/articles/{article_id}", response_model=Article)
async def update_article(
    article_id: str,
    article_data: ArticleCreate,
    current_user: User = Depends(require_permission(Permission.EDIT_ARTICLE))
):
    article_dict = article_data.dict()
    article_dict["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.articles.update_one(
        {"id": article_id}, 
        {"$set": article_dict}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Article not found")
    
    updated_article = await db.articles.find_one({"id": article_id})
    await manager.broadcast(json.dumps({"type": "article_updated", "id": article_id}))
    return Article(**updated_article)

@api_router.delete("/articles/{article_id}")
async def delete_article(article_id: str, current_user: User = Depends(require_permission(Permission.DELETE_ARTICLE))):
    # V7: Soft-delete — preserves history and allows recovery
    result = await db.articles.update_one(
        {"id": article_id, "deleted": {"$ne": True}},
        {"$set": {
            "deleted": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user.id
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Article not found")
    await manager.broadcast(json.dumps({"type": "article_deleted", "id": article_id}))
    return {"message": "Article deleted successfully"}

@api_router.post("/articles/{article_id}/images")
async def add_article_image(
    article_id: str,
    payload: dict,  # { "image": "base64string" }
    current_user: User = Depends(require_permission(Permission.EDIT_ARTICLE))
):
    """Append a base64 image to article.images list."""
    image_data = payload.get("image", "")
    if not image_data:
        raise HTTPException(status_code=400, detail="No image provided")
    result = await db.articles.update_one(
        {"id": article_id},
        {"$push": {"images": image_data}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Article not found")
    return {"message": "Image added"}

@api_router.delete("/articles/{article_id}/images/{image_index}")
async def delete_article_image(
    article_id: str,
    image_index: int,
    current_user: User = Depends(get_current_user)
):
    """Remove image at given index from article.images list."""
    article = await db.articles.find_one({"id": article_id})
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
    images = article.get("images", [])
    if image_index < 0 or image_index >= len(images):
        raise HTTPException(status_code=400, detail="Invalid image index")
    images.pop(image_index)
    await db.articles.update_one(
        {"id": article_id},
        {"$set": {"images": images}}
    )
    return {"message": "Image deleted"}

# F4 — Universal QR/barcode scan lookup and direct action

@api_router.get("/scan/{code}")
async def scan_lookup(code: str, current_user: User = Depends(get_current_user)):
    """Resolve a scanned code to an article, storage location, or booking."""
    code = code.strip()

    # 1. Try articles (qr_code or inventory_code)
    article = await db.articles.find_one({
        "$or": [{"qr_code": code}, {"inventory_code": code}],
        "deleted": {"$ne": True},
    })
    if article:
        article.pop("_id", None)
        return {"type": "article", "id": article["id"], "data": article}

    # 2. Try storage locations (qr_code)
    location = await db.storage_locations.find_one({"qr_code": code})
    if location:
        location.pop("_id", None)
        return {"type": "location", "id": location["id"], "data": location}

    # 3. Try booking by id
    booking = await db.bookings.find_one({"id": code})
    if booking:
        booking.pop("_id", None)
        return {"type": "booking", "id": booking["id"], "data": booking}

    raise HTTPException(status_code=404, detail="Code not found")


class ScanActionRequest(BaseModel):
    code: str = Field(..., max_length=500)
    action: str = Field(..., pattern="^(checkout|checkin|info)$")
    event_id: Optional[str] = None


@api_router.post("/scan/action")
async def scan_action(body: ScanActionRequest, current_user: User = Depends(get_current_user)):
    """Perform a direct checkout/checkin/info action from a scanned code."""
    code = body.code.strip()

    # Resolve code to a packing list item
    item = None
    article = await db.articles.find_one({
        "$or": [{"qr_code": code}, {"inventory_code": code}],
        "deleted": {"$ne": True},
    })
    if article and body.event_id:
        item = await db.packing_list_items.find_one({
            "article_id": article["id"],
            "event_id": body.event_id,
        })

    if body.action == "info":
        if article:
            article.pop("_id", None)
            return {"success": True, "message": "Artikel gefunden", "item": article}
        raise HTTPException(status_code=404, detail="Article not found for code")

    if not item:
        raise HTTPException(status_code=404, detail="No packing list item found for this code and event")

    now = datetime.now(timezone.utc)
    if body.action == "checkout":
        if item.get("checked_out"):
            return {"success": False, "message": "Bereits ausgecheckt", "item": item}
        await db.packing_list_items.update_one(
            {"id": item["id"]},
            {"$set": {"checked_out": True, "checked_out_by": current_user.username, "checked_out_at": now}}
        )
        return {"success": True, "message": "Ausgecheckt", "item": {**item, "checked_out": True}}

    if body.action == "checkin":
        if item.get("checked_in"):
            return {"success": False, "message": "Bereits eingecheckt", "item": item}
        await db.packing_list_items.update_one(
            {"id": item["id"]},
            {"$set": {"checked_in": True, "checked_in_by": current_user.username, "checked_in_at": now}}
        )
        return {"success": True, "message": "Eingecheckt", "item": {**item, "checked_in": True}}

    raise HTTPException(status_code=400, detail="Unknown action")


# Storage Zones
@api_router.post("/storage-zones", response_model=StorageZone)
async def create_storage_zone(zone: StorageZone, current_user: User = Depends(get_current_user)):
    zone.qr_code = f"ZONE-{zone.name.upper()}"
    await db.storage_zones.insert_one(zone.dict())
    return zone

@api_router.get("/storage-zones", response_model=List[StorageZone])
async def get_storage_zones(current_user: User = Depends(get_current_user)):
    zones = await db.storage_zones.find().to_list(1000)
    return [StorageZone(**zone) for zone in zones]

@api_router.put("/storage-zones/{zone_id}", response_model=StorageZone)
async def update_storage_zone(zone_id: str, zone_update: dict, current_user: User = Depends(get_current_user)):
    zone_update.pop("id", None)
    zone_update["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.storage_zones.update_one(
        {"id": zone_id},
        {"$set": zone_update}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Lagerzone nicht gefunden")
    zone = await db.storage_zones.find_one({"id": zone_id})
    return StorageZone(**zone)

@api_router.delete("/storage-zones/{zone_id}")
async def delete_storage_zone(zone_id: str, current_user: User = Depends(get_current_user)):
    # Check if zone has locations
    locations_count = await db.storage_locations.count_documents({"zone_id": zone_id})
    if locations_count > 0:
        raise HTTPException(status_code=400, detail=f"Zone hat noch {locations_count} Lagerorte. Bitte zuerst löschen.")
    
    result = await db.storage_zones.delete_one({"id": zone_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lagerzone nicht gefunden")
    return {"message": "Lagerzone gelöscht"}

# Storage Locations
@api_router.post("/storage-locations", response_model=StorageLocation)
async def create_storage_location(location: StorageLocation, current_user: User = Depends(get_current_user)):
    location.qr_code = f"LOC-{location.zone_id}-{location.name}"
    await db.storage_locations.insert_one(location.dict())
    return location

@api_router.get("/storage-locations", response_model=List[StorageLocation])
async def get_storage_locations(zone_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if zone_id:
        query["zone_id"] = zone_id
    
    locations = await db.storage_locations.find(query).to_list(1000)
    return [StorageLocation(**loc) for loc in locations]

@api_router.put("/storage-locations/{location_id}", response_model=StorageLocation)
async def update_storage_location(
    location_id: str,
    location_update: StorageLocation,
    current_user: User = Depends(get_current_user)
):
    """Update storage location"""
    existing_location = await db.storage_locations.find_one({"id": location_id})
    if not existing_location:
        raise HTTPException(status_code=404, detail="Storage location not found")
    
    # Update fields
    location_update.qr_code = f"LOC-{location_update.zone_id}-{location_update.name}"
    await db.storage_locations.update_one(
        {"id": location_id},
        {"$set": location_update.dict()}
    )
    
    return location_update

@api_router.delete("/storage-locations/{location_id}")
async def delete_storage_location(
    location_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete storage location"""
    result = await db.storage_locations.delete_one({"id": location_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Storage location not found")
    
    return {"message": "Storage location deleted successfully"}


# Inventory Movements
@api_router.post("/movements", response_model=InventoryMovement)
async def create_movement(
    movement_data: MovementCreate, 
    current_user: User = Depends(get_current_user)
):
    movement = InventoryMovement(**movement_data.dict(), user_id=current_user.id)
    
    # Verify article exists first
    article = await db.articles.find_one({"id": movement.article_id})
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")

    # Atomic stock update — prevents race conditions with concurrent requests
    if movement.movement_type == "IN":
        await db.articles.update_one(
            {"id": movement.article_id},
            {"$inc": {"current_stock": movement.quantity}, "$set": {"updated_at": datetime.now(timezone.utc)}}
        )
    elif movement.movement_type == "OUT":
        result = await db.articles.update_one(
            {"id": movement.article_id, "current_stock": {"$gte": movement.quantity}},
            {"$inc": {"current_stock": -movement.quantity}, "$set": {"updated_at": datetime.now(timezone.utc)}}
        )
        if result.modified_count == 0:
            raise HTTPException(status_code=400, detail="Insufficient stock")
    # TRANSFER: no stock change needed
    
    await db.movements.insert_one(movement.dict())

    # Log to sync_history for cross-platform sync
    await db.sync_history.insert_one({
        "entity_type": "movement",
        "entity_id": movement.id,
        "action": "create",
        "user_id": current_user.id,
        "username": current_user.username,
        "timestamp": datetime.now(timezone.utc),
        "data": movement.dict()
    })

    return movement

@api_router.get("/movements", response_model=List[InventoryMovement])
async def get_movements(
    article_id: Optional[str] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if article_id:
        query["article_id"] = article_id
    
    movements = await db.movements.find(query).sort("created_at", -1).limit(limit).to_list(limit)
    return [InventoryMovement(**mov) for mov in movements]

# Maintenance Tasks
@api_router.post("/maintenance/tasks", response_model=MaintenanceTask)
async def create_maintenance_task(
    task_data: MaintenanceTaskCreate,
    current_user: User = Depends(get_current_user)
):
    task = MaintenanceTask(**task_data.dict(), created_by=current_user.id)
    await db.maintenance_tasks.insert_one(task.dict())

    # Send notification if due date is within 7 days
    if task.due_date and task.due_date <= datetime.now(timezone.utc) + timedelta(days=7):
        article = await db.articles.find_one({"id": task.article_id})
        if article:
            await send_maintenance_notification(article["name"], task.due_date)

    await manager.broadcast(json.dumps({"type": "maintenance_created", "id": str(task.id)}))
    return task

@api_router.get("/maintenance/tasks", response_model=List[MaintenanceTask])
async def get_maintenance_tasks(
    article_id: Optional[str] = None,
    status: Optional[str] = None,
    assigned_to: Optional[str] = None,
    overdue_only: bool = False,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if article_id:
        query["article_id"] = article_id
    if status:
        query["status"] = status
    if assigned_to:
        query["assigned_to"] = assigned_to
    if overdue_only:
        query["due_date"] = {"$lt": datetime.now(timezone.utc)}
        query["status"] = {"$ne": "completed"}
    
    tasks = await db.maintenance_tasks.find(query).sort("due_date", 1).to_list(1000)
    return [MaintenanceTask(**task) for task in tasks]

@api_router.get("/maintenance/tasks/{task_id}", response_model=MaintenanceTask)
async def get_maintenance_task(task_id: str, current_user: User = Depends(get_current_user)):
    task = await db.maintenance_tasks.find_one({"id": task_id})
    if not task:
        raise HTTPException(status_code=404, detail="Maintenance task not found")
    return MaintenanceTask(**task)

@api_router.put("/maintenance/tasks/{task_id}", response_model=MaintenanceTask)
async def update_maintenance_task(
    task_id: str,
    task_data: MaintenanceTaskCreate,
    current_user: User = Depends(get_current_user)
):
    task_dict = task_data.dict()
    task_dict["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.maintenance_tasks.update_one(
        {"id": task_id},
        {"$set": task_dict}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Maintenance task not found")
    
    updated_task = await db.maintenance_tasks.find_one({"id": task_id})
    await manager.broadcast(json.dumps({"type": "maintenance_updated", "id": task_id}))
    return MaintenanceTask(**updated_task)

@api_router.put("/maintenance/tasks/{task_id}/complete")
async def complete_maintenance_task(
    task_id: str,
    current_user: User = Depends(get_current_user)
):
    result = await db.maintenance_tasks.update_one(
        {"id": task_id},
        {
            "$set": {
                "status": "completed",
                "completed_at": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc)
            }
        }
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Maintenance task not found")
    
    return {"message": "Task completed successfully"}

# Maintenance Records
@api_router.post("/maintenance/records", response_model=MaintenanceRecord)
async def create_maintenance_record(
    record_data: MaintenanceRecordCreate,
    current_user: User = Depends(get_current_user)
):
    record = MaintenanceRecord(**record_data.dict(), performed_by=current_user.id)
    await db.maintenance_records.insert_one(record.dict())
    
    # Update article status and next maintenance date
    update_data = {
        "status": record.status_after,
        "last_maintenance": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    if record.next_maintenance_date:
        update_data["next_maintenance"] = record.next_maintenance_date
    
    await db.articles.update_one(
        {"id": record.article_id},
        {"$set": update_data}
    )
    
    # Complete the associated task
    await db.maintenance_tasks.update_one(
        {"id": record.task_id},
        {
            "$set": {
                "status": "completed",
                "completed_at": datetime.now(timezone.utc),
                "actual_duration": None,  # Would need to calculate from start time
                "updated_at": datetime.now(timezone.utc)
            }
        }
    )
    
    return record

@api_router.get("/maintenance/records", response_model=List[MaintenanceRecord])
async def get_maintenance_records(
    article_id: Optional[str] = None,
    task_id: Optional[str] = None,
    performed_by: Optional[str] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if article_id:
        query["article_id"] = article_id
    if task_id:
        query["task_id"] = task_id
    if performed_by:
        query["performed_by"] = performed_by
    
    records = await db.maintenance_records.find(query).sort("created_at", -1).limit(limit).to_list(limit)
    return [MaintenanceRecord(**record) for record in records]

# Maintenance Checklists
@api_router.post("/maintenance/checklists", response_model=MaintenanceChecklist)
async def create_maintenance_checklist(
    checklist_data: MaintenanceChecklistCreate,
    current_user: User = Depends(get_current_user)
):
    checklist = MaintenanceChecklist(**checklist_data.dict(), created_by=current_user.id)
    await db.maintenance_checklists.insert_one(checklist.dict())
    return checklist

@api_router.get("/maintenance/checklists", response_model=List[MaintenanceChecklist])
async def get_maintenance_checklists(
    category_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {"is_template": True}
    if category_id:
        query["category_ids"] = {"$in": [category_id]}
    
    checklists = await db.maintenance_checklists.find(query).to_list(1000)
    return [MaintenanceChecklist(**checklist) for checklist in checklists]

# ============================================================
# WARTUNG ERWEITERT - DGUV V3, Betriebsstunden, Reparatur-Tickets
# ============================================================

# === DGUV V3 Prüfungen ===

async def generate_dguv_protocol_number():
    """Generate unique DGUV V3 protocol number"""
    year = datetime.now().year
    count = await db.dguv_inspections.count_documents({
        "protocol_number": {"$regex": f"^DGUV-{year}"}
    })
    return f"DGUV-{year}-{str(count + 1).zfill(5)}"

@api_router.post("/dguv-v3/inspections")
async def create_dguv_inspection(
    inspection_data: DGUVV3InspectionCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a new DGUV V3 inspection record"""
    try:
        article = await db.articles.find_one({"id": inspection_data.article_id})
        if not article:
            raise HTTPException(status_code=404, detail="Artikel nicht gefunden")
        
        protocol_number = inspection_data.protocol_number or await generate_dguv_protocol_number()
        
        inspection = DGUVV3Inspection(
            **inspection_data.dict(),
            protocol_number=protocol_number,
            created_by=current_user.id
        )
        
        await db.dguv_inspections.insert_one(inspection.dict())
        
        # Update article with next inspection date
        await db.articles.update_one(
            {"id": inspection_data.article_id},
            {"$set": {
                "next_maintenance": inspection_data.next_inspection_date,
                "last_maintenance": inspection_data.inspection_date,
                "status": "OK" if inspection_data.result == "bestanden" else "gesperrt",
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        
        # Create maintenance task for next inspection
        next_task = MaintenanceTask(
            article_id=inspection_data.article_id,
            title=f"DGUV V3 Prüfung - {article['name']}",
            description="Wiederkehrende DGUV V3 Prüfung fällig",
            task_type="dguv_v3",
            priority="high",
            due_date=inspection_data.next_inspection_date,
            created_by=current_user.id
        )
        await db.maintenance_tasks.insert_one(next_task.dict())
        
        return inspection
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating DGUV inspection: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")

@api_router.get("/dguv-v3/inspections")
async def get_dguv_inspections(
    article_id: Optional[str] = None,
    result: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all DGUV V3 inspections"""
    query = {}
    if article_id:
        query["article_id"] = article_id
    if result:
        query["result"] = result
    
    inspections = await db.dguv_inspections.find(query).sort("inspection_date", -1).to_list(500)
    return inspections

@api_router.get("/dguv-v3/due")
async def get_dguv_due_inspections(
    days: int = 30,
    current_user: User = Depends(get_current_user)
):
    """Get articles with DGUV V3 inspections due within X days"""
    due_date = datetime.now(timezone.utc) + timedelta(days=days)
    
    # Find tasks of type dguv_v3 due soon
    tasks = await db.maintenance_tasks.find({
        "task_type": "dguv_v3",
        "status": {"$ne": "completed"},
        "due_date": {"$lte": due_date}
    }).sort("due_date", 1).to_list(100)
    
    results = []
    for task in tasks:
        article = await db.articles.find_one({"id": task.get("article_id")})
        last_inspection = await db.dguv_inspections.find_one(
            {"article_id": task.get("article_id")},
            sort=[("inspection_date", -1)]
        )
        
        is_overdue = task.get("due_date") and task["due_date"] < datetime.now(timezone.utc)
        
        results.append({
            "task_id": task.get("id"),
            "article_id": task.get("article_id"),
            "article_name": article.get("name") if article else "Unbekannt",
            "inventory_code": article.get("inventory_code") if article else "",
            "due_date": task.get("due_date"),
            "is_overdue": is_overdue,
            "days_until_due": (task.get("due_date") - datetime.now(timezone.utc)).days if task.get("due_date") else None,
            "last_inspection": last_inspection.get("inspection_date") if last_inspection else None,
            "last_result": last_inspection.get("result") if last_inspection else None
        })
    
    return {
        "due_count": len(results),
        "overdue_count": len([r for r in results if r.get("is_overdue")]),
        "inspections": results
    }

# === Betriebsstunden-Tracking ===

@api_router.post("/articles/{article_id}/operating-hours")
async def update_operating_hours(
    article_id: str,
    hours_update: OperatingHoursUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update operating hours for an article and check for maintenance threshold"""
    try:
        article = await db.articles.find_one({"id": article_id})
        if not article:
            raise HTTPException(status_code=404, detail="Artikel nicht gefunden")
        
        current_hours = article.get("operating_hours", 0) or 0
        new_hours = current_hours + hours_update.hours_to_add
        max_hours = article.get("max_operating_hours")
        
        update_data = {
            "operating_hours": new_hours,
            "updated_at": datetime.now(timezone.utc)
        }
        
        # Check if maintenance is needed
        maintenance_needed = False
        if max_hours and new_hours >= max_hours:
            maintenance_needed = True
            # Create automatic maintenance task
            task = MaintenanceTask(
                article_id=article_id,
                title=f"Wartung nach {int(new_hours)} Betriebsstunden",
                description=f"Artikel hat {int(new_hours)} von {int(max_hours)} max. Betriebsstunden erreicht. Wartung erforderlich!",
                task_type="routine",
                priority="high",
                due_date=datetime.now(timezone.utc) + timedelta(days=7),
                created_by=current_user.id
            )
            await db.maintenance_tasks.insert_one(task.dict())
        
        await db.articles.update_one({"id": article_id}, {"$set": update_data})
        
        # Log the update
        log_entry = {
            "id": str(uuid.uuid4()),
            "article_id": article_id,
            "previous_hours": current_hours,
            "added_hours": hours_update.hours_to_add,
            "new_hours": new_hours,
            "notes": hours_update.notes,
            "user_id": current_user.id,
            "timestamp": datetime.now(timezone.utc)
        }
        await db.operating_hours_log.insert_one(log_entry)
        
        return {
            "article_id": article_id,
            "previous_hours": current_hours,
            "added_hours": hours_update.hours_to_add,
            "new_hours": new_hours,
            "max_hours": max_hours,
            "maintenance_needed": maintenance_needed,
            "percentage_used": round((new_hours / max_hours) * 100, 1) if max_hours else None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating operating hours: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")

# === Reparatur-Tickets mit Fotos ===

async def generate_ticket_number():
    """Generate unique repair ticket number"""
    year = datetime.now().year
    count = await db.repair_tickets.count_documents({
        "ticket_number": {"$regex": f"^REP-{year}"}
    })
    return f"REP-{year}-{str(count + 1).zfill(5)}"

@api_router.post("/repair-tickets")
async def create_repair_ticket(
    ticket_data: RepairTicketCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a new repair ticket with defect photos"""
    try:
        article = await db.articles.find_one({"id": ticket_data.article_id})
        if not article:
            raise HTTPException(status_code=404, detail="Artikel nicht gefunden")
        
        ticket_number = await generate_ticket_number()
        
        ticket = RepairTicket(
            **ticket_data.dict(),
            ticket_number=ticket_number,
            reported_by=current_user.id
        )
        
        await db.repair_tickets.insert_one(ticket.dict())
        
        # Update article status to defective
        if ticket_data.severity in ["high", "critical"]:
            await db.articles.update_one(
                {"id": ticket_data.article_id},
                {"$set": {"status": "defekt", "updated_at": datetime.now(timezone.utc)}}
            )
        
        # Create a maintenance task
        task = MaintenanceTask(
            article_id=ticket_data.article_id,
            title=f"Reparatur: {ticket_data.title}",
            description=ticket_data.description,
            task_type="repair",
            priority=ticket_data.severity,
            created_by=current_user.id
        )
        await db.maintenance_tasks.insert_one(task.dict())
        
        return ticket
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating repair ticket: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")

@api_router.get("/repair-tickets")
async def get_repair_tickets(
    status: Optional[str] = None,
    severity: Optional[str] = None,
    article_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all repair tickets"""
    query = {}
    if status:
        query["status"] = status
    if severity:
        query["severity"] = severity
    if article_id:
        query["article_id"] = article_id
    
    tickets = await db.repair_tickets.find(query).sort("created_at", -1).to_list(500)
    for t in tickets:
        t["id"] = str(t.pop("_id", t.get("id", "")))
    return tickets

@api_router.get("/repair-tickets/{ticket_id}")
async def get_repair_ticket(
    ticket_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get a single repair ticket"""
    ticket = await db.repair_tickets.find_one({"id": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket nicht gefunden")
    
    # Get article info
    article = await db.articles.find_one({"id": ticket.get("article_id")})
    ticket["article"] = article
    
    return ticket

@api_router.put("/repair-tickets/{ticket_id}")
async def update_repair_ticket(
    ticket_id: str,
    status: Optional[str] = None,
    assigned_to: Optional[str] = None,
    repair_notes: Optional[str] = None,
    repair_images: Optional[List[str]] = None,
    parts_used: Optional[List[Dict[str, Any]]] = None,
    repair_cost: Optional[float] = None,
    repair_time_minutes: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Update a repair ticket"""
    ticket = await db.repair_tickets.find_one({"id": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket nicht gefunden")
    
    update_data = {"updated_at": datetime.now(timezone.utc)}
    
    if status:
        update_data["status"] = status
        if status in ["repaired", "closed"]:
            update_data["closed_at"] = datetime.now(timezone.utc)
            # Update article status back to OK
            await db.articles.update_one(
                {"id": ticket.get("article_id")},
                {"$set": {"status": "OK", "updated_at": datetime.now(timezone.utc)}}
            )
    if assigned_to:
        update_data["assigned_to"] = assigned_to
    if repair_notes:
        update_data["repair_notes"] = repair_notes
    if repair_images:
        update_data["repair_images"] = repair_images
    if parts_used:
        update_data["parts_used"] = parts_used
    if repair_cost is not None:
        update_data["repair_cost"] = repair_cost
    if repair_time_minutes is not None:
        update_data["repair_time_minutes"] = repair_time_minutes
    
    await db.repair_tickets.update_one({"id": ticket_id}, {"$set": update_data})
    
    return {"message": "Ticket aktualisiert", "ticket_id": ticket_id}

@api_router.post("/repair-tickets/{ticket_id}/images")
async def add_repair_images(
    ticket_id: str,
    image_type: str,  # "defect" or "repair"
    images: List[str],  # Base64 images
    current_user: User = Depends(get_current_user)
):
    """Add images to a repair ticket"""
    ticket = await db.repair_tickets.find_one({"id": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket nicht gefunden")
    
    field = "defect_images" if image_type == "defect" else "repair_images"
    current_images = ticket.get(field, [])
    updated_images = current_images + images
    
    await db.repair_tickets.update_one(
        {"id": ticket_id},
        {"$set": {field: updated_images, "updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"message": f"{len(images)} Bilder hinzugefügt", "total_images": len(updated_images)}

# Combined maintenance dashboard
@api_router.get("/maintenance/dashboard")
async def get_maintenance_dashboard(
    current_user: User = Depends(get_current_user)
):
    """Get comprehensive maintenance dashboard data"""
    now = datetime.now(timezone.utc)
    
    # Pending tasks by type
    tasks = await db.maintenance_tasks.find({"status": {"$ne": "completed"}}).to_list(1000)
    
    routine_tasks = [t for t in tasks if t.get("task_type") == "routine"]
    repair_tasks = [t for t in tasks if t.get("task_type") == "repair"]
    dguv_tasks = [t for t in tasks if t.get("task_type") == "dguv_v3"]
    
    # Overdue counts
    overdue_count = len([t for t in tasks if t.get("due_date") and t["due_date"] < now])
    
    # Open repair tickets
    open_tickets = await db.repair_tickets.count_documents({"status": {"$in": ["open", "in_progress"]}})
    critical_tickets = await db.repair_tickets.count_documents({"status": {"$in": ["open", "in_progress"]}, "severity": "critical"})
    
    # DGUV due soon (30 days)
    dguv_due = await db.maintenance_tasks.count_documents({
        "task_type": "dguv_v3",
        "status": {"$ne": "completed"},
        "due_date": {"$lte": now + timedelta(days=30)}
    })
    
    # Operating hours alerts
    articles_with_hours = await db.articles.find({"max_operating_hours": {"$gt": 0}}).to_list(1000)
    hours_critical = len([a for a in articles_with_hours 
                         if a.get("operating_hours", 0) >= (a.get("max_operating_hours", 1) * 0.9)])
    
    return {
        "summary": {
            "total_pending_tasks": len(tasks),
            "overdue_tasks": overdue_count,
            "open_repair_tickets": open_tickets,
            "critical_tickets": critical_tickets,
            "dguv_due_30_days": dguv_due,
            "operating_hours_critical": hours_critical
        },
        "tasks_by_type": {
            "routine": len(routine_tasks),
            "repair": len(repair_tasks),
            "dguv_v3": len(dguv_tasks)
        },
        "upcoming_tasks": sorted(
            [t for t in tasks if t.get("due_date")],
            key=lambda x: x.get("due_date")
        )[:10]
    }

# Maintenance Executions
@api_router.post("/maintenance/executions", response_model=MaintenanceExecution)
async def create_maintenance_execution(
    execution_data: MaintenanceExecutionCreate,
    current_user: User = Depends(get_current_user)
):
    execution = MaintenanceExecution(**execution_data.dict(), performed_by=current_user.id)
    await db.maintenance_executions.insert_one(execution.dict())
    return execution

# Dashboard stats (updated with maintenance)
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: User = Depends(get_current_user)):
    import time
    global _dashboard_cache, _dashboard_cache_ts
    # F9: Return cached result if still fresh (30 s TTL)
    if _dashboard_cache and (time.monotonic() - _dashboard_cache_ts) < DASHBOARD_CACHE_TTL:
        return _dashboard_cache

    total_articles = await db.articles.count_documents({})
    low_stock_articles = await db.articles.count_documents({
        "$expr": {
            "$and": [
                {"$gt": ["$min_stock_level", 0]},
                {"$lt": ["$current_stock", "$min_stock_level"]}
            ]
        }
    })
    
    # Get articles needing maintenance
    maintenance_due = await db.maintenance_tasks.count_documents({
        "due_date": {"$lte": datetime.now(timezone.utc) + timedelta(days=7)},
        "status": {"$ne": "completed"}
    })
    
    # Overdue maintenance
    overdue_maintenance = await db.maintenance_tasks.count_documents({
        "due_date": {"$lt": datetime.now(timezone.utc)},
        "status": {"$ne": "completed"}
    })
    
    total_movements_today = await db.movements.count_documents({
        "created_at": {"$gte": datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)}
    })
    
    # Articles by status
    defective_articles = await db.articles.count_documents({"status": "defekt"})
    blocked_articles = await db.articles.count_documents({"status": "gesperrt"})
    
    # New statistics
    total_customers = await db.customers.count_documents({"is_active": True})
    total_events = await db.events.count_documents({})
    active_events = await db.events.count_documents({
        "status": {"$in": ["confirmed", "ongoing"]}
    })
    
    # Events this month
    now = datetime.now(timezone.utc)
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    events_this_month = await db.events.count_documents({
        "start_date": {"$gte": start_of_month.isoformat()}
    })
    
    # Total inventory value — computed server-side via aggregation, no full collection load
    _value_pipeline = [
        {"$project": {"value": {"$multiply": [
            {"$ifNull": ["$price_per_unit", 0]},
            {"$ifNull": ["$current_stock", 0]}
        ]}}},
        {"$group": {"_id": None, "total": {"$sum": "$value"}}}
    ]
    _value_result = await db.articles.aggregate(_value_pipeline).to_list(1)
    total_inventory_value = _value_result[0]["total"] if _value_result else 0
    
    # Movements statistics (last 7 days)
    seven_days_ago = datetime.now(timezone.utc) - timedelta(days=7)
    movements_last_7_days = await db.movements.count_documents({
        "created_at": {"$gte": seven_days_ago}
    })
    
    result = {
        "total_articles": total_articles,
        "low_stock_articles": low_stock_articles,
        "maintenance_due": maintenance_due,
        "overdue_maintenance": overdue_maintenance,
        "movements_today": total_movements_today,
        "defective_articles": defective_articles,
        "blocked_articles": blocked_articles,
        "total_customers": total_customers,
        "total_events": total_events,
        "active_events": active_events,
        "events_this_month": events_this_month,
        "total_inventory_value": round(total_inventory_value, 2),
        "movements_last_7_days": movements_last_7_days,
    }
    # F9: Store result in cache
    _dashboard_cache = result
    _dashboard_cache_ts = time.monotonic()
    return result

# Check maintenance alerts
@api_router.get("/maintenance/alerts")
async def get_maintenance_alerts(current_user: User = Depends(get_current_user)):
    # Get overdue tasks
    overdue_tasks = await db.maintenance_tasks.find({
        "due_date": {"$lt": datetime.now(timezone.utc)},
        "status": {"$ne": "completed"}
    }).to_list(100)
    
    # Get upcoming tasks (next 7 days)
    upcoming_tasks = await db.maintenance_tasks.find({
        "due_date": {
            "$gte": datetime.now(timezone.utc),
            "$lte": datetime.now(timezone.utc) + timedelta(days=7)
        },
        "status": {"$ne": "completed"}
    }).to_list(100)
    
    return {
        "overdue": [MaintenanceTask(**task) for task in overdue_tasks],
        "upcoming": [MaintenanceTask(**task) for task in upcoming_tasks]
    }

# Basic route for testing
@api_router.get("/")
async def root():
    return {"message": "Inventory Management System API", "version": "1.0.0"}

# Messaging System
@api_router.post("/messages", response_model=Message)
@limiter.limit("10/minute")  # K1: Prevent message flooding / email spam
async def send_message(request: Request, message_data: MessageCreate, current_user: User = Depends(get_current_user)):
    # Validate recipient exists
    recipient = await db.users.find_one({"id": message_data.recipient_id})
    if not recipient:
        raise HTTPException(status_code=404, detail="Recipient not found")
    
    message = Message(
        sender_id=current_user.id,
        recipient_id=message_data.recipient_id,
        message_text=message_data.message_text
    )
    
    await db.messages.insert_one(message.dict())
    return message

@api_router.get("/messages/conversations")
async def get_conversations(current_user: User = Depends(get_current_user)):
    """Get list of conversations with other users"""
    # Find all users who have exchanged messages with current user
    conversations = []
    
    # Get all unique users current user has communicated with
    sent_to = await db.messages.find({"sender_id": current_user.id}).distinct("recipient_id")
    received_from = await db.messages.find({"recipient_id": current_user.id}).distinct("sender_id")
    
    unique_user_ids = list(set(sent_to + received_from))
    
    for user_id in unique_user_ids:
        user = await db.users.find_one({"id": user_id})
        if not user:
            continue
        
        # Get last message
        last_message = await db.messages.find_one(
            {
                "$or": [
                    {"sender_id": current_user.id, "recipient_id": user_id},
                    {"sender_id": user_id, "recipient_id": current_user.id}
                ]
            },
            sort=[("created_at", -1)]
        )
        
        # Count unread messages from this user
        unread_count = await db.messages.count_documents({
            "sender_id": user_id,
            "recipient_id": current_user.id,
            "is_read": False
        })
        
        conversations.append({
            "user_id": user_id,
            "username": user["username"],
            "last_message": last_message.get("message_text") if last_message else None,
            "last_message_time": last_message.get("created_at") if last_message else None,
            "unread_count": unread_count
        })
    
    # Sort by last message time
    conversations.sort(key=lambda x: x["last_message_time"] or datetime.min, reverse=True)
    
    return conversations

@api_router.get("/messages/{other_user_id}")
async def get_messages_with_user(
    other_user_id: str,
    limit: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Get all messages between current user and another user"""
    messages = await db.messages.find({
        "$or": [
            {"sender_id": current_user.id, "recipient_id": other_user_id},
            {"sender_id": other_user_id, "recipient_id": current_user.id}
        ]
    }).sort("created_at", -1).limit(limit).to_list(limit)
    
    # Mark messages from other user as read
    await db.messages.update_many(
        {
            "sender_id": other_user_id,
            "recipient_id": current_user.id,
            "is_read": False
        },
        {"$set": {"is_read": True, "updated_at": datetime.now(timezone.utc)}}
    )
    
    # Reverse to show oldest first
    messages.reverse()
    
    return [Message(**msg) for msg in messages]

@api_router.get("/messages/unread/count")
async def get_unread_count(current_user: User = Depends(get_current_user)):
    """Get total unread message count"""
    count = await db.messages.count_documents({
        "recipient_id": current_user.id,
        "is_read": False
    })
    return {"unread_count": count}

@api_router.get("/users/all")
async def get_all_users(current_user: User = Depends(get_current_user)):
    """Get all users for starting new conversations"""
    users = await db.users.find(
        {"id": {"$ne": current_user.id}, "is_approved": True, "is_active": True}
    ).to_list(1000)
    
    return [{
        "id": user["id"],
        "username": user["username"],
        "role": user["role"]
    } for user in users]

# Customer Management
@api_router.post("/customers", response_model=Customer)
async def create_customer(customer_data: CustomerCreate, current_user: User = Depends(get_current_user)):
    customer_number = await generate_customer_number()
    customer = Customer(**customer_data.dict(), customer_number=customer_number)
    await db.customers.insert_one(customer.dict())
    return customer

@api_router.get("/customers", response_model=List[Customer])
async def get_customers(
    search: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {"is_active": True}
    if search:
        safe_search = re.escape(search)
        query["$or"] = [
            {"company_name": {"$regex": safe_search, "$options": "i"}},
            {"contact_person": {"$regex": safe_search, "$options": "i"}},
            {"customer_number": {"$regex": safe_search, "$options": "i"}},
            {"email": {"$regex": safe_search, "$options": "i"}}
        ]
    
    customers = await db.customers.find(query).sort("created_at", -1).to_list(1000)
    return [Customer(**customer) for customer in customers]

@api_router.get("/customers/{customer_id}", response_model=Customer)
async def get_customer(customer_id: str, current_user: User = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return Customer(**customer)

@api_router.put("/customers/{customer_id}", response_model=Customer)
async def update_customer(
    customer_id: str,
    customer_data: CustomerCreate,
    current_user: User = Depends(get_current_user)
):
    customer_dict = customer_data.dict()
    customer_dict["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.customers.update_one(
        {"id": customer_id},
        {"$set": customer_dict}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    updated_customer = await db.customers.find_one({"id": customer_id})
    return Customer(**updated_customer)

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, current_user: User = Depends(get_current_user)):
    # V7: Soft-delete with audit trail
    result = await db.customers.update_one(
        {"id": customer_id, "is_active": {"$ne": False}},
        {"$set": {
            "is_active": False,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user.id,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"message": "Customer deleted successfully"}

# Event Management  
@api_router.post("/events", response_model=Event)
async def create_event(event_data: EventCreate, current_user: User = Depends(require_permission(Permission.CREATE_EVENT))):
    # Validate customer exists
    customer = await db.customers.find_one({"id": event_data.customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Validate dates
    if event_data.end_date <= event_data.start_date:
        raise HTTPException(status_code=400, detail="End date must be after start date")
    
    event_number = await generate_event_number()
    event = Event(**event_data.dict(), event_number=event_number, created_by=current_user.id)
    await db.events.insert_one(event.dict())
    await manager.broadcast(json.dumps({"type": "event_created", "id": str(event.id)}))
    return event

@api_router.get("/events", response_model=List[Event])
async def get_events(
    customer_id: Optional[str] = None,
    status: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query: dict = {"deleted": {"$ne": True}}  # V7: exclude soft-deleted
    if customer_id:
        query["customer_id"] = customer_id
    if status:
        query["status"] = status
    if from_date:
        query["start_date"] = {"$gte": parse_iso_date(from_date, "from_date")}
    if to_date:
        if "start_date" in query:
            query["start_date"]["$lte"] = parse_iso_date(to_date, "to_date")
        else:
            query["start_date"] = {"$lte": parse_iso_date(to_date, "to_date")}

    events = await db.events.find(query).sort("start_date", -1).to_list(1000)
    return [Event(**event) for event in events]

@api_router.get("/events/{event_id}", response_model=Event)
async def get_event(event_id: str, current_user: User = Depends(get_current_user)):
    event = await db.events.find_one({"id": event_id})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    return Event(**event)

@api_router.put("/events/{event_id}", response_model=Event)
async def update_event(
    event_id: str,
    event_data: EventCreate,
    current_user: User = Depends(require_permission(Permission.EDIT_EVENT))
):
    event_dict = event_data.dict()
    event_dict["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.events.update_one(
        {"id": event_id},
        {"$set": event_dict}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    
    updated_event = await db.events.find_one({"id": event_id})
    await manager.broadcast(json.dumps({"type": "event_updated", "id": event_id}))
    return Event(**updated_event)

@api_router.delete("/events/{event_id}")
async def delete_event(event_id: str, current_user: User = Depends(require_permission(Permission.DELETE_EVENT))):
    # V7: Soft-delete — preserves booking history
    result = await db.events.update_one(
        {"id": event_id, "deleted": {"$ne": True}},
        {"$set": {
            "deleted": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user.id
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    return {"message": "Event deleted successfully"}

# F1 — Generate invoice directly from an event's booked items
@api_router.post("/events/{event_id}/generate-invoice")
async def generate_invoice_from_event(
    event_id: str,
    due_days: int = 30,
    notes: str = "",
    current_user: User = Depends(require_permission(Permission.CREATE_INVOICE))
):
    event = await db.events.find_one({"id": event_id, "deleted": {"$ne": True}})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    bookings = await db.bookings.find({"event_id": event_id, "status": "booked"}).to_list(1000)
    if not bookings:
        raise HTTPException(status_code=400, detail="No booked items found for this event")

    # Bulk-fetch articles to avoid N+1
    article_ids = list({b["article_id"] for b in bookings if b.get("article_id")})
    articles_list = await db.articles.find({"id": {"$in": article_ids}}).to_list(len(article_ids) + 1)
    articles_map = {a["id"]: a for a in articles_list}

    # F6: use tiered pricing based on event duration
    try:
        ev_start = datetime.fromisoformat(event["start_date"].replace("Z", "+00:00"))
        ev_end = datetime.fromisoformat(event["end_date"].replace("Z", "+00:00"))
        event_days = max(1, (ev_end - ev_start).days + 1)
    except Exception:
        event_days = 1

    items = []
    subtotal = 0.0
    for booking in bookings:
        article = articles_map.get(booking.get("article_id", ""))
        if not article:
            continue
        quantity = int(booking.get("quantity", 1))
        # F6: tiered price per period (not multiplied by days again — period price is the unit)
        unit_price = _calculate_rental_price(article, event_days) if event_days > 0 else float(
            article.get("rental_price", article.get("price", 0.0))
        )
        total_price = unit_price * quantity
        items.append({
            "article_id": article["id"],
            "article_name": article["name"],
            "quantity": quantity,
            "unit_price": unit_price,
            "total_price": total_price,
        })
        subtotal += total_price

    tax_amount = round(subtotal * 0.19, 2)
    total_amount = round(subtotal + tax_amount, 2)
    subtotal = round(subtotal, 2)

    invoice_number = await generate_invoice_number()
    invoice = Invoice(
        invoice_number=invoice_number,
        event_id=event_id,
        customer_id=event.get("customer_id", ""),
        items=items,
        subtotal=subtotal,
        tax_amount=tax_amount,
        total_amount=total_amount,
        notes=notes,
        due_date=datetime.now(timezone.utc) + timedelta(days=due_days),
        created_by=current_user.id,
    )
    await db.invoices.insert_one(invoice.dict())
    return {"invoice_id": invoice.id, "invoice_number": invoice_number, "total_amount": total_amount}


# F7 — Clone an event (with optional booking copy)
@api_router.post("/events/{event_id}/clone")
async def clone_event(
    event_id: str,
    new_start_date: str = "",
    new_end_date: str = "",
    clone_bookings: bool = True,
    current_user: User = Depends(require_permission(Permission.CREATE_EVENT))
):
    source = await db.events.find_one({"id": event_id, "deleted": {"$ne": True}})
    if not source:
        raise HTTPException(status_code=404, detail="Event not found")

    # Calculate date offset for booking shift
    def _parse_date(val) -> datetime:
        if isinstance(val, datetime):
            return val
        return datetime.fromisoformat(str(val).replace("Z", "+00:00"))

    try:
        old_start = _parse_date(source["start_date"])
        if new_start_date:
            new_start = _parse_date(new_start_date)
        else:
            new_start = old_start + timedelta(days=7)  # default: one week later
        date_offset = new_start - old_start

        if new_end_date:
            new_end = _parse_date(new_end_date)
        else:
            new_end = _parse_date(source["end_date"]) + date_offset
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid date format — use ISO 8601")

    event_number = await generate_event_number()
    new_event_id = str(uuid.uuid4())
    clone = {k: v for k, v in source.items() if k not in ("_id", "id", "event_number", "status",
                                                             "deleted", "deleted_at", "deleted_by")}
    clone.update({
        "id": new_event_id,
        "event_number": event_number,
        "status": "anfrage",
        "start_date": new_start.isoformat(),
        "end_date": new_end.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat() + "Z",
        "created_by": current_user.id,
    })
    await db.events.insert_one(clone)

    cloned_bookings = 0
    if clone_bookings:
        bookings = await db.bookings.find({"event_id": event_id}).to_list(1000)
        for b in bookings:
            new_b = {k: v for k, v in b.items() if k not in ("_id", "id")}
            new_b["id"] = str(uuid.uuid4())
            new_b["event_id"] = new_event_id
            # Shift pickup/return dates by the same offset
            for date_field in ("pickup_date", "return_date"):
                if new_b.get(date_field):
                    try:
                        d = datetime.fromisoformat(new_b[date_field].replace("Z", "+00:00"))
                        new_b[date_field] = (d + date_offset).isoformat()
                    except Exception:
                        pass
            await db.bookings.insert_one(new_b)
            cloned_bookings += 1

    return {
        "event_id": new_event_id,
        "event_number": event_number,
        "cloned_bookings": cloned_bookings,
    }


# Booking Management
@api_router.post("/bookings", response_model=Booking)
async def create_booking(booking_data: BookingCreate, current_user: User = Depends(get_current_user)):
    # Validate event exists
    event = await db.events.find_one({"id": booking_data.event_id})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    # Validate article exists
    article = await db.articles.find_one({"id": booking_data.article_id})
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")

    # Early stock check — before insert to give clear error message
    article_stock = article.get("current_stock", 0)
    if article_stock < booking_data.quantity:
        raise HTTPException(
            status_code=400,
            detail=f"Nicht genug Bestand. Verfügbar: {article_stock}, angefragt: {booking_data.quantity}"
        )

    # Determine dates
    pickup_date = booking_data.pickup_date or event["start_date"]
    return_date = booking_data.return_date or event["end_date"]

    # K2: Two-phase conflict check to close the check→insert race window.
    # Phase 1: optimistic pre-check (fast path for non-contested slots)
    conflicts = await check_booking_conflict(booking_data.article_id, pickup_date, return_date)
    if conflicts:
        raise HTTPException(status_code=409, detail="Article is already booked for this time period")

    # Phase 2: insert immediately, then post-validate while our booking is in DB
    booking_dict = booking_data.dict()
    booking_dict.update({
        "booked_by": current_user.id,
        "pickup_date": pickup_date,
        "return_date": return_date
    })
    booking = Booking(**booking_dict)
    await db.bookings.insert_one(booking.dict())

    # Re-check conflicts now that our record is persisted (catches concurrent inserts)
    post_conflicts = await db.bookings.count_documents({
        "article_id": booking_data.article_id,
        "id": {"$ne": booking.id},
        "status": {"$in": ["booked", "confirmed", "active"]},
        "$or": [
            {"pickup_date": {"$lt": return_date}, "return_date": {"$gt": pickup_date}},
        ]
    })
    if post_conflicts > 0:
        # Another booking won the race — roll back ours
        await db.bookings.delete_one({"id": booking.id})
        raise HTTPException(status_code=409, detail="Article is already booked for this time period")

    # Atomares Abziehen des Bestands – verhindert Race Condition bei parallelen Buchungen
    stock_result = await db.articles.update_one(
        {"id": booking_data.article_id, "current_stock": {"$gte": booking_data.quantity}},
        {"$inc": {"current_stock": -booking_data.quantity}, "$set": {"updated_at": datetime.now(timezone.utc)}}
    )
    if stock_result.modified_count == 0:
        # Buchung rückgängig machen – kein Bestand mehr verfügbar
        await db.bookings.delete_one({"id": booking.id})
        # Get current stock for error message
        current_article = await db.articles.find_one({"id": booking_data.article_id})
        raise HTTPException(
            status_code=400,
            detail=f"Nicht genug Bestand (Race Condition). Verfügbarer Bestand: {current_article.get('current_stock', 0)}"
        )
    
    # Send notification to customer
    customer = await db.customers.find_one({"id": event["customer_id"]})
    if customer:
        await send_booking_notification_email(
            customer["email"],
            event["event_name"],
            "Neue Buchung"
        )
    
    return booking

@api_router.get("/bookings", response_model=List[Booking])
async def get_bookings(
    event_id: Optional[str] = None,
    article_id: Optional[str] = None,
    status: Optional[str] = None,
    page: int = Query(default=1, ge=1, description="Seitennummer (ab 1)"),
    page_size: int = Query(default=100, ge=1, le=500, description="Einträge pro Seite (max. 500)"),
    current_user: User = Depends(get_current_user)
):
    query = {}
    if event_id:
        query["event_id"] = event_id
    if article_id:
        query["article_id"] = article_id
    if status:
        query["status"] = status

    skip = (page - 1) * page_size
    bookings = await db.bookings.find(query).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    return [Booking(**booking) for booking in bookings]

@api_router.get("/bookings/{booking_id}", response_model=Booking)
async def get_booking(booking_id: str, current_user: User = Depends(get_current_user)):
    booking = await db.bookings.find_one({"id": booking_id})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    return Booking(**booking)

@api_router.put("/bookings/{booking_id}/return")
async def return_booking(booking_id: str, current_user: User = Depends(get_current_user)):
    booking = await db.bookings.find_one({"id": booking_id})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")

    # Atomarer Status-Check + Update – verhindert doppeltes Gutschreiben bei parallelen Requests
    result = await db.bookings.update_one(
        {"id": booking_id, "status": "booked"},
        {"$set": {"status": "returned", "updated_at": datetime.now(timezone.utc)}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=400, detail="Booking is not active")

    # Atomares Gutschreiben des Bestands
    await db.articles.update_one(
        {"id": booking["article_id"]},
        {"$inc": {"current_stock": booking["quantity"]}, "$set": {"updated_at": datetime.now(timezone.utc)}}
    )
    
    # Send notification to customer
    event = await db.events.find_one({"id": booking["event_id"]})
    if event:
        customer = await db.customers.find_one({"id": event["customer_id"]})
        if customer:
            await send_booking_notification_email(
                customer["email"],
                event["event_name"],
                "Rückgabe bestätigt"
            )
    
    return {"message": "Booking returned successfully"}

@api_router.delete("/bookings/{booking_id}")
async def cancel_booking(booking_id: str, current_user: User = Depends(get_current_user)):
    booking = await db.bookings.find_one({"id": booking_id})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if booking["status"] == "returned":
        raise HTTPException(status_code=400, detail="Cannot cancel returned booking")
    
    # Update booking status
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {"status": "cancelled", "updated_at": datetime.now(timezone.utc)}}
    )
    
    # Return stock atomically if booking was active — prevents race conditions
    if booking["status"] == "booked":
        await db.articles.update_one(
            {"id": booking["article_id"]},
            {"$inc": {"current_stock": booking["quantity"]}, "$set": {"updated_at": datetime.now(timezone.utc)}}
        )
    
    # Send notification to customer
    event = await db.events.find_one({"id": booking["event_id"]})
    if event:
        customer = await db.customers.find_one({"id": event["customer_id"]})
        if customer:
            await send_booking_notification_email(
                customer["email"],
                event["event_name"],
                "Buchung storniert"
            )
    
    return {"message": "Booking cancelled successfully"}


# F3 — Assign serial numbers to a booking
class AssignSerialsRequest(BaseModel):
    serial_number_ids: List[str] = Field(..., min_length=1, max_length=200)

@api_router.post("/bookings/{booking_id}/assign-serials")
async def assign_serials_to_booking(
    booking_id: str,
    body: AssignSerialsRequest,
    current_user: User = Depends(get_current_user)
):
    booking = await db.bookings.find_one({"id": booking_id})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")

    # Validate all serials belong to the correct article and are available
    serials = await db.serial_numbers.find({"id": {"$in": body.serial_number_ids}}).to_list(200)
    if len(serials) != len(body.serial_number_ids):
        raise HTTPException(status_code=400, detail="One or more serial numbers not found")

    wrong_article = [s["id"] for s in serials if s.get("article_id") != booking["article_id"]]
    if wrong_article:
        raise HTTPException(status_code=400, detail=f"Serial numbers belong to wrong article: {wrong_article}")

    unavailable = [s["id"] for s in serials if s.get("status") != "verfügbar"]
    if unavailable:
        raise HTTPException(status_code=400, detail=f"Serial numbers not available: {unavailable}")

    # Mark serials as lent
    await db.serial_numbers.update_many(
        {"id": {"$in": body.serial_number_ids}},
        {"$set": {"status": "verliehen", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    # Store on booking
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {"serial_number_ids": body.serial_number_ids, "updated_at": datetime.now(timezone.utc)}}
    )
    return {"assigned": len(body.serial_number_ids)}


# Time Tracking Endpoints
@api_router.post("/time-entries")
async def create_time_entry(entry_data: TimeEntryCreate, current_user: User = Depends(get_current_user)):
    # Calculate hours worked
    try:
        start_h, start_m = map(int, entry_data.start_time.split(":"))
        end_h, end_m = map(int, entry_data.end_time.split(":"))
        # L2 — Reject if start is not before end
        if start_h * 60 + start_m >= end_h * 60 + end_m:
            raise HTTPException(status_code=400, detail="start_time muss vor end_time liegen")
        total_minutes = (end_h * 60 + end_m) - (start_h * 60 + start_m) - entry_data.break_minutes
        hours_worked = max(0, total_minutes / 60)
    except HTTPException:
        raise
    except (ValueError, AttributeError):
        hours_worked = 0

    # Get crew member's hourly rate
    crew_member = await db.crew.find_one({"id": entry_data.crew_member_id})
    hourly_rate = crew_member.get("hourly_rate", 0) if crew_member else 0

    entry = TimeEntry(
        **entry_data.dict(),
        hours_worked=round(hours_worked, 2),
        hourly_rate=hourly_rate,
        total_pay=round(hours_worked * hourly_rate, 2),
    )
    await db.time_entries.insert_one(entry.dict())
    return entry.dict()

@api_router.get("/time-entries")
async def get_time_entries(
    crew_member_id: Optional[str] = None,
    event_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query: dict = {}
    if crew_member_id:
        query["crew_member_id"] = crew_member_id
    if event_id:
        query["event_id"] = event_id
    if date_from or date_to:
        query["date"] = {}
        if date_from:
            query["date"]["$gte"] = date_from
        if date_to:
            query["date"]["$lte"] = date_to
    entries = await db.time_entries.find(query).sort("date", -1).to_list(5000)
    for e in entries:
        e.pop("_id", None)
    return entries

@api_router.put("/time-entries/{entry_id}")
async def update_time_entry(entry_id: str, update_data: TimeEntryCreate, current_user: User = Depends(get_current_user)):
    try:
        start_h, start_m = map(int, update_data.start_time.split(":"))
        end_h, end_m = map(int, update_data.end_time.split(":"))
        # L2 — Reject if start is not before end
        if start_h * 60 + start_m >= end_h * 60 + end_m:
            raise HTTPException(status_code=400, detail="start_time muss vor end_time liegen")
        total_minutes = (end_h * 60 + end_m) - (start_h * 60 + start_m) - update_data.break_minutes
        hours_worked = max(0, total_minutes / 60)
    except HTTPException:
        raise
    except (ValueError, AttributeError):
        hours_worked = 0
    crew_member = await db.crew.find_one({"id": update_data.crew_member_id})
    hourly_rate = crew_member.get("hourly_rate", 0) if crew_member else 0
    update_dict = update_data.dict()
    update_dict.update({
        "hours_worked": round(hours_worked, 2),
        "hourly_rate": hourly_rate,
        "total_pay": round(hours_worked * hourly_rate, 2),
    })
    result = await db.time_entries.update_one({"id": entry_id}, {"$set": update_dict})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Zeiteintrag nicht gefunden")
    return {"message": "Aktualisiert"}

@api_router.delete("/time-entries/{entry_id}")
async def delete_time_entry(entry_id: str, current_user: User = Depends(get_current_user)):
    result = await db.time_entries.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Zeiteintrag nicht gefunden")
    return {"message": "Gelöscht"}

@api_router.get("/time-entries/summary")
async def get_time_entries_summary(
    month: Optional[str] = None,  # format: "2026-03"
    current_user: User = Depends(get_current_user)
):
    """Payroll summary per crew member for a given month"""
    query: dict = {}
    if month:
        query["date"] = {"$gte": f"{month}-01", "$lte": f"{month}-31"}

    entries = await db.time_entries.find(query).to_list(10000)
    crew_list = await db.crew.find().to_list(500)
    crew_map = {c["id"]: c for c in crew_list}

    from collections import defaultdict
    summary: dict = defaultdict(lambda: {"crew_id": "", "crew_name": "", "hourly_rate": 0, "total_hours": 0, "total_pay": 0, "entry_count": 0})

    for entry in entries:
        cid = entry.get("crew_member_id", "")
        cm = crew_map.get(cid, {})
        s = summary[cid]
        s["crew_id"] = cid
        s["crew_name"] = cm.get("name", cid)
        s["hourly_rate"] = cm.get("hourly_rate", 0)
        s["total_hours"] = round(s["total_hours"] + entry.get("hours_worked", 0), 2)
        s["total_pay"] = round(s["total_pay"] + entry.get("total_pay", 0), 2)
        s["entry_count"] += 1

    return list(summary.values())

# Quote Management
@api_router.post("/quotes")
async def create_quote(quote_data: QuoteCreate, current_user: User = Depends(get_current_user)):
    count = await db.quotes.count_documents({})
    quote_number = f"QUO-{datetime.now(timezone.utc).year}-{count + 1:04d}"
    total_net = sum(item.unit_price * item.quantity * item.days for item in quote_data.items)
    if quote_data.discount_percent > 0:
        total_net = total_net * (1 - quote_data.discount_percent / 100)
    quote = Quote(
        **quote_data.dict(),
        quote_number=quote_number,
        total_net=round(total_net, 2),
        created_by=getattr(current_user, "id", str(current_user)),
    )
    await db.quotes.insert_one(quote.dict())
    return quote.dict()

@api_router.get("/quotes")
async def get_quotes(
    status: Optional[str] = None,
    customer_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if customer_id:
        query["customer_id"] = customer_id
    quotes = await db.quotes.find(query).sort("created_at", -1).to_list(1000)
    for q in quotes:
        q.pop("_id", None)
    return quotes

@api_router.get("/quotes/{quote_id}")
async def get_quote(quote_id: str, current_user: User = Depends(get_current_user)):
    quote = await db.quotes.find_one({"id": quote_id})
    if not quote:
        raise HTTPException(status_code=404, detail="Angebot nicht gefunden")
    quote.pop("_id", None)
    return quote

@api_router.put("/quotes/{quote_id}")
async def update_quote(quote_id: str, update_data: QuoteCreate, current_user: User = Depends(get_current_user)):
    total_net = sum(item.unit_price * item.quantity * item.days for item in update_data.items)
    if update_data.discount_percent > 0:
        total_net = total_net * (1 - update_data.discount_percent / 100)
    update_dict = update_data.dict()
    update_dict["total_net"] = round(total_net, 2)
    update_dict["updated_at"] = datetime.now(timezone.utc)
    result = await db.quotes.update_one({"id": quote_id}, {"$set": update_dict})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Angebot nicht gefunden")
    return {"message": "Angebot aktualisiert"}

@api_router.put("/quotes/{quote_id}/status")
async def update_quote_status(quote_id: str, update: dict, current_user: User = Depends(get_current_user)):
    valid = ["entwurf", "gesendet", "akzeptiert", "abgelehnt"]
    new_status = update.get("status")
    if new_status not in valid:
        raise HTTPException(status_code=400, detail="Ungültiger Status")
    result = await db.quotes.update_one(
        {"id": quote_id},
        {"$set": {"status": new_status, "updated_at": datetime.now(timezone.utc)}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Angebot nicht gefunden")
    return {"message": "Status aktualisiert"}

@api_router.put("/quotes/{quote_id}/sign")
async def sign_quote(quote_id: str, body: dict, current_user: User = Depends(get_current_user)):
    """Save customer signature and accept quote"""
    result = await db.quotes.update_one(
        {"id": quote_id},
        {"$set": {
            "signature_customer": body.get("signature_customer", ""),
            "signature_date": datetime.now(timezone.utc),
            "signed_by": body.get("signed_by", ""),
            "status": "akzeptiert",
            "updated_at": datetime.now(timezone.utc),
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Angebot nicht gefunden")
    return {"message": "Angebot akzeptiert und unterschrieben"}

@api_router.post("/quotes/{quote_id}/share")
async def share_quote(quote_id: str, current_user: User = Depends(get_current_user)):
    """Generate public share token for a quote"""
    token = str(uuid.uuid4())
    result = await db.quotes.update_one(
        {"id": quote_id},
        {"$set": {"public_token": token, "public_token_created_at": datetime.now(timezone.utc)}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Angebot nicht gefunden")
    return {"public_token": token}

@api_router.delete("/quotes/{quote_id}/share")
async def revoke_quote_share(quote_id: str, current_user: User = Depends(get_current_user)):
    """Revoke public share token"""
    await db.quotes.update_one({"id": quote_id}, {"$set": {"public_token": None, "public_token_created_at": None}})
    return {"message": "Link widerrufen"}

@api_router.delete("/quotes/{quote_id}")
async def delete_quote(quote_id: str, current_user: User = Depends(get_current_user)):
    result = await db.quotes.delete_one({"id": quote_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Angebot nicht gefunden")
    return {"message": "Angebot gelöscht"}

# Invoice Management
@api_router.post("/invoices", response_model=Invoice)
async def create_invoice(invoice_data: InvoiceCreate, current_user: User = Depends(get_current_user)):
    # Determine customer_id
    customer_id = invoice_data.customer_id

    # Get event if provided
    if invoice_data.event_id:
        event = await db.events.find_one({"id": invoice_data.event_id})
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")

        # Use customer_id from event if not provided in request
        if not customer_id:
            customer_id = event["customer_id"]

        # Get bookings for this event
        bookings = await db.bookings.find({"event_id": invoice_data.event_id, "status": "booked"}).to_list(1000)
        if not bookings:
            raise HTTPException(status_code=400, detail="No bookings found for this event")

        # Calculate invoice items
        items = []
        subtotal = 0.0

        for booking in bookings:
            article = await db.articles.find_one({"id": booking["article_id"]})
            if article:
                unit_price = article.get("price", 0.0)
                quantity = booking["quantity"]
                total_price = unit_price * quantity

                items.append({
                    "article_id": article["id"],
                    "article_name": article["name"],
                    "quantity": quantity,
                    "unit_price": unit_price,
                    "total_price": total_price
                })
                subtotal += total_price
    else:
        # Free invoice without event
        if not customer_id:
            raise HTTPException(status_code=400, detail="customer_id is required for free invoices (without event)")
        items = []
        subtotal = 0.0

    # Validate customer exists
    customer = await db.customers.find_one({"id": customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    # Calculate tax
    tax_amount = subtotal * 0.19  # 19% MwSt
    total_amount = subtotal + tax_amount

    # Generate invoice number
    invoice_number = await generate_invoice_number()

    # Create invoice
    invoice = Invoice(
        invoice_number=invoice_number,
        event_id=invoice_data.event_id,
        customer_id=customer_id,
        items=items,
        subtotal=subtotal,
        tax_amount=tax_amount,
        total_amount=total_amount,
        notes=invoice_data.notes,
        due_date=datetime.now(timezone.utc) + timedelta(days=invoice_data.due_days),
        created_by=current_user.id
    )

    await db.invoices.insert_one(invoice.dict())
    return invoice

@api_router.get("/invoices", response_model=List[Invoice])
async def get_invoices(
    customer_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if customer_id:
        query["customer_id"] = customer_id
    if status:
        query["status"] = status
    
    invoices = await db.invoices.find(query).sort("created_at", -1).to_list(1000)
    return [Invoice(**invoice) for invoice in invoices]

@api_router.get("/invoices/{invoice_id}", response_model=Invoice)
async def get_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    invoice = await db.invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return Invoice(**invoice)

@api_router.put("/invoices/{invoice_id}/status")
async def update_invoice_status(
    invoice_id: str,
    status: str,
    current_user: User = Depends(get_current_user)
):
    valid_statuses = ["draft", "sent", "paid", "cancelled"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")
    
    update_data = {"status": status, "updated_at": datetime.now(timezone.utc)}
    if status == "paid":
        update_data["paid_date"] = datetime.now(timezone.utc)
    
    result = await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    return {"message": f"Invoice status updated to {status}"}

@api_router.put("/invoices/{invoice_id}/payment-status")
async def update_invoice_payment_status(invoice_id: str, update: dict, current_user: User = Depends(get_current_user)):
    valid_statuses = ["offen", "teilweise", "bezahlt", "überfällig"]
    new_status = update.get("payment_status")
    if new_status not in valid_statuses:
        raise HTTPException(status_code=400, detail="Ungültiger Zahlungsstatus")
    result = await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {"payment_status": new_status, "updated_at": datetime.now(timezone.utc)}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Rechnung nicht gefunden")
    return {"message": "Zahlungsstatus aktualisiert"}

@api_router.get("/conflicts")
async def get_booking_conflicts(current_user: User = Depends(get_current_user)):
    """Detect articles that are double-booked for overlapping time periods"""
    # Load all active bookings (not cancelled, not returned)
    bookings = await db.bookings.find({"status": {"$nin": ["cancelled", "returned"]}}).to_list(5000)

    # Load events to get date ranges
    events_list = await db.events.find().to_list(5000)
    events_map = {e["id"]: e for e in events_list}

    # Load articles for names
    articles_list = await db.articles.find().to_list(5000)
    articles_map = {a["id"]: a for a in articles_list}

    conflicts = []
    seen_pairs = set()

    # Group bookings by article_id
    from collections import defaultdict
    by_article = defaultdict(list)
    for b in bookings:
        by_article[b.get("article_id", "")].append(b)

    for article_id, article_bookings in by_article.items():
        if len(article_bookings) < 2:
            continue
        article_name = articles_map.get(article_id, {}).get("name", article_id)

        for i in range(len(article_bookings)):
            for j in range(i + 1, len(article_bookings)):
                b1 = article_bookings[i]
                b2 = article_bookings[j]

                if b1.get("event_id") == b2.get("event_id"):
                    continue

                # Get date ranges: use pickup/return_date if set, else use event dates
                e1 = events_map.get(b1.get("event_id", ""), {})
                e2 = events_map.get(b2.get("event_id", ""), {})

                start1 = b1.get("pickup_date") or e1.get("start_date")
                end1 = b1.get("return_date") or e1.get("end_date")
                start2 = b2.get("pickup_date") or e2.get("start_date")
                end2 = b2.get("return_date") or e2.get("end_date")

                if not all([start1, end1, start2, end2]):
                    continue

                # Convert to datetime if string
                from datetime import datetime as dt
                def to_dt(v):
                    if isinstance(v, str):
                        try:
                            return dt.fromisoformat(v.replace("Z", "+00:00"))
                        except (ValueError, AttributeError):
                            return None
                    return v

                s1, e1d, s2, e2d = to_dt(start1), to_dt(end1), to_dt(start2), to_dt(end2)
                if not all([s1, e1d, s2, e2d]):
                    continue

                # Check overlap: s1 < e2 and s2 < e1
                if s1 < e2d and s2 < e1d:
                    pair_key = tuple(sorted([b1.get("id",""), b2.get("id","")]))
                    if pair_key not in seen_pairs:
                        seen_pairs.add(pair_key)
                        conflicts.append({
                            "article_name": article_name,
                            "article_id": article_id,
                            "event1_name": e1.get("event_name", "Unbekannt"),
                            "event1_id": b1.get("event_id", ""),
                            "start1": str(s1.date()),
                            "end1": str(e1d.date()),
                            "event2_name": e2.get("event_name", "Unbekannt"),
                            "event2_id": b2.get("event_id", ""),
                            "start2": str(s2.date()),
                            "end2": str(e2d.date()),
                        })

    return conflicts

@api_router.get("/conflicts/crew")
async def get_crew_conflicts(current_user: User = Depends(get_current_user)):
    """Detect crew members assigned to multiple events at overlapping times"""
    assignments = await db.crew_assignments.find().to_list(5000)
    events_list = await db.events.find().to_list(5000)
    events_map = {e["id"]: e for e in events_list}
    crew_list = await db.crew.find().to_list(500)
    crew_map = {c["id"]: c for c in crew_list}

    # Expand: one record per (crew_member_id, event_id, start, end)
    crew_event_entries = []
    for a in assignments:
        e = events_map.get(a.get("event_id", ""), {})
        start = a.get("start_time") or e.get("start_date")
        end = a.get("end_time") or e.get("end_date")
        for crew_id in a.get("crew_ids", []):
            crew_event_entries.append({
                "crew_id": crew_id,
                "event_id": a.get("event_id", ""),
                "event_name": e.get("event_name", "Unbekannt"),
                "start": start,
                "end": end,
            })

    from collections import defaultdict
    from datetime import datetime as dt
    def to_dt(v):
        if isinstance(v, str):
            try:
                return dt.fromisoformat(v.replace("Z", "+00:00"))
            except:
                return None
        return v

    conflicts = []
    seen_pairs = set()
    by_crew = defaultdict(list)
    for entry in crew_event_entries:
        by_crew[entry["crew_id"]].append(entry)

    for crew_id, entries in by_crew.items():
        if len(entries) < 2:
            continue
        crew_name = crew_map.get(crew_id, {}).get("name", crew_id)
        for i in range(len(entries)):
            for j in range(i + 1, len(entries)):
                e1, e2 = entries[i], entries[j]
                if e1["event_id"] == e2["event_id"]:
                    continue
                s1, end1 = to_dt(e1["start"]), to_dt(e1["end"])
                s2, end2 = to_dt(e2["start"]), to_dt(e2["end"])
                if not all([s1, end1, s2, end2]):
                    continue
                if s1 < end2 and s2 < end1:
                    pair_key = tuple(sorted([e1["event_id"], e2["event_id"], crew_id]))
                    if pair_key not in seen_pairs:
                        seen_pairs.add(pair_key)
                        conflicts.append({
                            "crew_name": crew_name,
                            "crew_id": crew_id,
                            "event1_name": e1["event_name"],
                            "event1_id": e1["event_id"],
                            "start1": str(s1.date()),
                            "end1": str(end1.date()),
                            "event2_name": e2["event_name"],
                            "event2_id": e2["event_id"],
                            "start2": str(s2.date()),
                            "end2": str(end2.date()),
                        })

    return conflicts

@api_router.get("/events/{event_id}/bookings-summary")
async def get_event_bookings_summary(event_id: str, current_user: User = Depends(get_current_user)):
    """Get summary of all bookings for an event with article details"""
    bookings = await db.bookings.find({"event_id": event_id}).to_list(1000)

    # F5: Bulk-fetch all articles in one $in query instead of N find_one calls
    article_ids = list({b["article_id"] for b in bookings if b.get("article_id")})
    articles_list = await db.articles.find({"id": {"$in": article_ids}}).to_list(len(article_ids) + 1)
    articles_map = {a["id"]: a for a in articles_list}

    summary = []
    total_value = 0.0

    for booking in bookings:
        article = articles_map.get(booking["article_id"])
        if article:
            item_total = article.get("price", 0.0) * booking["quantity"]
            summary.append({
                "booking_id": booking["id"],
                "article_id": article["id"],
                "article_name": article["name"],
                "article_code": article.get("inventory_code", ""),
                "quantity": booking["quantity"],
                "unit_price": article.get("price", 0.0),
                "total_price": item_total,
                "status": booking["status"],
                "pickup_date": booking.get("pickup_date"),
                "return_date": booking.get("return_date")
            })
            if booking["status"] == "booked":
                total_value += item_total
    
    return {
        "event_id": event_id,
        "bookings": summary,
        "total_items": len(summary),
        "total_value": total_value
    }

# Admin-only endpoints for user approval
@api_router.get("/admin/pending-users")
async def get_pending_users(current_user: User = Depends(require_permission(Permission.MANAGE_USERS))):
    
    pending_users = await db.users.find({"is_approved": False}).to_list(100)
    return [{
        "id": user["id"],
        "username": user["username"],
        "email": user["email"],
        "role": user["role"],
        "created_at": user["created_at"]
    } for user in pending_users]

@api_router.put("/admin/approve-user/{user_id}")
async def approve_user(user_id: str, current_user: User = Depends(require_permission(Permission.MANAGE_USERS))):
    
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"is_approved": True}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")

    updated_user = await db.users.find_one({"id": user_id})
    # M3: Audit trail for admin approval actions
    await create_audit_log("APPROVE_USER", "user", current_user,
                           entity_id=user_id, entity_name=updated_user.get("username"),
                           changes={"approved_by": current_user.username})
    return {
        "message": "User approved successfully",
        "username": updated_user["username"],
        "email": updated_user["email"]
    }

@api_router.delete("/admin/reject-user/{user_id}")
async def reject_user(user_id: str, current_user: User = Depends(require_permission(Permission.MANAGE_USERS))):
    target = await db.users.find_one({"id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")

    # M3: Audit trail for admin rejection actions
    await create_audit_log("REJECT_USER", "user", current_user,
                           entity_id=user_id, entity_name=target.get("username"),
                           changes={"rejected_by": current_user.username})
    return {"message": "User rejected and deleted successfully"}

# Backup & Restore Endpoints (Admin only)
@api_router.post("/admin/backup/create")
async def create_backup(current_user: User = Depends(require_permission(Permission.BACKUP_DATABASE))):
    
    success = await create_database_backup()
    
    if success:
        # Get backup file stats
        backup_size = BACKUP_FILE.stat().st_size / (1024 * 1024)  # MB
        return {
            "message": "Backup created successfully",
            "file": str(BACKUP_FILE),
            "size_mb": round(backup_size, 2),
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
    else:
        raise HTTPException(status_code=500, detail="Backup failed")

@api_router.post("/admin/backup/restore")
async def restore_backup(current_user: User = Depends(require_permission(Permission.BACKUP_DATABASE))):
    
    success, message = await restore_database_from_backup()
    
    if success:
        return {
            "message": "Database restored successfully",
            "details": message,
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
    else:
        raise HTTPException(status_code=500, detail=f"Restore failed: {message}")

@api_router.get("/admin/backup/download")
async def download_backup(current_user: User = Depends(require_permission(Permission.BACKUP_DATABASE))):
    """Download backup as ZIP file"""
    import tempfile
    import zipfile
    from pathlib import Path
    from fastapi.responses import FileResponse

    try:
        # Create a cross-platform temp file (auto-cleaned after response in background)
        tmp = tempfile.NamedTemporaryFile(suffix='.zip', delete=False)
        zip_path = Path(tmp.name)
        tmp.close()

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            backup_dir = Path("/app/backups")
            if backup_dir.exists():
                for file in backup_dir.rglob('*'):
                    if file.is_file():
                        zipf.write(file, file.relative_to(backup_dir.parent))

        # H2: Reject oversized backups to prevent memory exhaustion on large deployments
        zip_size_mb = zip_path.stat().st_size / (1024 * 1024)
        if zip_size_mb > MAX_BACKUP_SIZE_MB:
            zip_path.unlink(missing_ok=True)
            raise HTTPException(
                status_code=413,
                detail=f"Backup ({zip_size_mb:.0f} MB) überschreitet das Limit von {MAX_BACKUP_SIZE_MB} MB"
            )

        def _cleanup(path: Path):
            try:
                path.unlink(missing_ok=True)
            except Exception:
                pass

        response = FileResponse(
            path=str(zip_path),
            media_type='application/zip',
            filename=f'inventar_backup_{datetime.now(timezone.utc).strftime("%Y%m%d")}.zip',
            background=None,
        )
        # Schedule temp-file deletion via Starlette background
        from starlette.background import BackgroundTask
        response.background = BackgroundTask(_cleanup, zip_path)
        return response
    except Exception as e:
        logging.error(f"Backup download failed: {e}")
        raise HTTPException(status_code=500, detail="Download fehlgeschlagen. Details wurden protokolliert.")

@api_router.get("/admin/backup/info")
async def get_backup_info(current_user: User = Depends(require_permission(Permission.BACKUP_DATABASE))):
    
    if not BACKUP_FILE.exists():
        return {
            "exists": False,
            "message": "No backup file found"
        }
    
    try:
        # Get backup file info
        backup_size = BACKUP_FILE.stat().st_size / (1024 * 1024)  # MB
        backup_time = datetime.fromtimestamp(BACKUP_FILE.stat().st_mtime)
        
        # Read backup metadata
        with open(BACKUP_FILE, 'r') as f:
            backup_data = json.load(f)
        
        collections_info = {
            name: len(docs) 
            for name, docs in backup_data["collections"].items()
        }
        
        return {
            "exists": True,
            "file": str(BACKUP_FILE),
            "size_mb": round(backup_size, 2),
            "created_at": backup_data.get("timestamp"),
            "last_modified": backup_time.isoformat(),
            "collections": collections_info,
            "total_documents": sum(collections_info.values())
        }
    except Exception as e:
        return {
            "exists": True,
            "error": str(e)
        }

# Alias for backup list (for compatibility)
@api_router.get("/backup/list")
async def get_backup_list(current_user: User = Depends(get_current_user)):
    """Alias für /admin/backup/info - Backup-Liste abrufen"""
    return await get_backup_info(current_user)

# Cross-platform sync endpoints (basic version for now)
@api_router.get("/sync/status")
async def get_sync_status(current_user: User = Depends(get_current_user)):
    """Get synchronization status"""
    return {
        "status": "active",
        "server_time": datetime.now(timezone.utc).isoformat(),
        "user_id": current_user.id,
        "username": current_user.username,
        "platform_support": {
            "android": True,
            "ios": True, 
            "web": True,
            "windows": True
        }
    }

@api_router.get("/sync/changes")
async def get_recent_changes(
    since: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get recent changes for synchronization"""
    try:
        if since:
            since_date = parse_iso_date(since, "since")
        else:
            since_date = datetime.now(timezone.utc) - timedelta(hours=24)
        
        # Get recent articles changes
        articles_changes = await db.articles.find({
            "updated_at": {"$gte": since_date}
        }).sort("updated_at", -1).limit(100).to_list(100)
        
        # Get recent movements
        movements_changes = await db.movements.find({
            "created_at": {"$gte": since_date}
        }).sort("created_at", -1).limit(100).to_list(100)
        
        # Get recent maintenance updates
        maintenance_changes = await db.maintenance_tasks.find({
            "updated_at": {"$gte": since_date}
        }).sort("updated_at", -1).limit(100).to_list(100)
        
        return {
            "server_time": datetime.now(timezone.utc).isoformat(),
            "changes": {
                "articles": [Article(**article).dict() for article in articles_changes],
                "movements": [InventoryMovement(**movement).dict() for movement in movements_changes], 
                "maintenance_tasks": [MaintenanceTask(**task).dict() for task in maintenance_changes]
            },
            "total_changes": len(articles_changes) + len(movements_changes) + len(maintenance_changes)
        }
        
    except Exception as e:
        logger.error(f"Error getting sync changes: {e}")
        raise HTTPException(status_code=500, detail="Failed to get sync changes")

# Bill of Materials (BOM) Management
@api_router.post("/bom", response_model=BillOfMaterials)
async def create_bom(bom_data: BOMCreate, current_user: User = Depends(get_current_user)):
    """Create a new Bill of Materials (Equipment Package)"""
    # Validate all articles exist
    for item in bom_data.items:
        article = await db.articles.find_one({"id": item.article_id})
        if not article:
            raise HTTPException(status_code=404, detail=f"Article {item.article_id} not found")
    
    bom = BillOfMaterials(**bom_data.dict(), created_by=current_user.id)
    await db.bom.insert_one(bom.dict())
    return bom

@api_router.get("/bom", response_model=List[BillOfMaterials])
async def get_boms(
    category: Optional[str] = None,
    is_active: bool = True,
    current_user: User = Depends(get_current_user)
):
    """Get all Bills of Materials"""
    query = {"is_active": is_active}
    if category:
        query["category"] = category
    
    boms = await db.bom.find(query).sort("created_at", -1).to_list(1000)
    return [BillOfMaterials(**bom) for bom in boms]

@api_router.get("/bom/{bom_id}", response_model=BillOfMaterials)
async def get_bom(bom_id: str, current_user: User = Depends(get_current_user)):
    """Get BOM details"""
    bom = await db.bom.find_one({"id": bom_id})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")
    return BillOfMaterials(**bom)

@api_router.get("/bom/{bom_id}/details")
async def get_bom_with_articles(bom_id: str, current_user: User = Depends(get_current_user)):
    """Get BOM with full article details"""
    bom = await db.bom.find_one({"id": bom_id})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")
    
    # Get full article details for each item
    items_with_details = []
    total_regular_price = 0.0
    
    for item in bom["items"]:
        article = await db.articles.find_one({"id": item["article_id"]})
        if article:
            rental_price = article.get("rental_price") or 0.0
            item_price = rental_price * item["quantity"]
            total_regular_price += item_price
            
            items_with_details.append({
                "article_id": article["id"],
                "article_name": article["name"],
                "article_code": article.get("inventory_code", ""),
                "quantity": item["quantity"],
                "is_optional": item.get("is_optional", False),
                "unit_price": rental_price,
                "total_price": item_price,
                "current_stock": article.get("current_stock", 0),
                "image_base64": article.get("image_base64")
            })
    
    package_price = bom.get("package_price", total_regular_price)
    discount = total_regular_price - package_price if package_price else 0
    discount_percent = (discount / total_regular_price * 100) if total_regular_price > 0 else 0
    
    return {
        "bom": BillOfMaterials(**bom),
        "items": items_with_details,
        "total_regular_price": total_regular_price,
        "package_price": package_price,
        "discount_amount": discount,
        "discount_percent": round(discount_percent, 1)
    }

@api_router.put("/bom/{bom_id}", response_model=BillOfMaterials)
async def update_bom(
    bom_id: str,
    bom_data: BOMCreate,
    current_user: User = Depends(get_current_user)
):
    """Update BOM"""
    bom_dict = bom_data.dict()
    bom_dict["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.bom.update_one(
        {"id": bom_id},
        {"$set": bom_dict}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="BOM not found")
    
    updated_bom = await db.bom.find_one({"id": bom_id})
    return BillOfMaterials(**updated_bom)

@api_router.delete("/bom/{bom_id}")
async def delete_bom(bom_id: str, current_user: User = Depends(get_current_user)):
    """Soft delete BOM"""
    result = await db.bom.update_one(
        {"id": bom_id},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="BOM not found")
    return {"message": "BOM deleted successfully"}

@api_router.post("/bom/{bom_id}/book-to-event")
async def book_bom_to_event(
    bom_id: str,
    event_id: str,
    current_user: User = Depends(get_current_user)
):
    """Book entire BOM package to an event"""
    # Get BOM
    bom = await db.bom.find_one({"id": bom_id})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")
    
    # Get Event
    event = await db.events.find_one({"id": event_id})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    # Check stock and conflicts for all items
    conflicts = []
    stock_issues = []
    
    for item in bom["items"]:
        article = await db.articles.find_one({"id": item["article_id"]})
        if not article:
            continue
        
        # Check stock
        if article["current_stock"] < item["quantity"]:
            stock_issues.append({
                "article_name": article["name"],
                "requested": item["quantity"],
                "available": article["current_stock"]
            })
        
        # Check conflicts
        item_conflicts = await check_booking_conflict(
            item["article_id"],
            event["start_date"],
            event["end_date"]
        )
        if item_conflicts:
            conflicts.append({
                "article_name": article["name"],
                "conflicts": len(item_conflicts)
            })
    
    if stock_issues:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "Insufficient stock",
                "issues": stock_issues
            }
        )
    
    if conflicts:
        raise HTTPException(
            status_code=409,
            detail={
                "error": "Booking conflicts detected",
                "conflicts": conflicts
            }
        )
    
    # Create bookings for all items
    bookings_created = []
    
    for item in bom["items"]:
        if item.get("is_optional", False):
            continue  # Skip optional items
        
        booking = Booking(
            event_id=event_id,
            article_id=item["article_id"],
            quantity=item["quantity"],
            booked_by=current_user.id,
            pickup_date=event["start_date"],
            return_date=event["end_date"],
            notes=f"Gebucht als Teil von Paket: {bom['name']}"
        )
        
        await db.bookings.insert_one(booking.dict())

        # Atomic stock deduction — prevents race condition if booked concurrently
        stock_result = await db.articles.update_one(
            {"id": item["article_id"], "current_stock": {"$gte": item["quantity"]}},
            {"$inc": {"current_stock": -item["quantity"]}, "$set": {"updated_at": datetime.now(timezone.utc)}}
        )
        if stock_result.modified_count == 0:
            # Rollback the just-created booking and abort
            await db.bookings.delete_one({"id": booking.id})
            raise HTTPException(
                status_code=400,
                detail=f"Bestand für Artikel {item['article_id']} nicht mehr ausreichend"
            )
        
        bookings_created.append({
            "booking_id": booking.id,
            "article_id": item["article_id"],
            "quantity": item["quantity"]
        })
    
    return {
        "message": f"Package '{bom['name']}' successfully booked",
        "bookings_created": len(bookings_created),
        "bookings": bookings_created
    }

# Project Templates
@api_router.get("/project-templates")
async def get_project_templates(current_user: User = Depends(get_current_user)):
    templates = await db.project_templates.find().sort("name", 1).to_list(500)
    return [ProjectTemplate(**t) for t in templates]

@api_router.post("/project-templates", response_model=ProjectTemplate)
async def create_project_template(data: ProjectTemplateCreate, current_user: User = Depends(get_current_user)):
    template = ProjectTemplate(**data.dict(), created_by=current_user.id)
    await db.project_templates.insert_one(template.dict())
    return template

@api_router.put("/project-templates/{template_id}", response_model=ProjectTemplate)
async def update_project_template(template_id: str, data: ProjectTemplateCreate, current_user: User = Depends(get_current_user)):
    update = {**data.dict(), "updated_at": datetime.now(timezone.utc)}
    result = await db.project_templates.update_one({"id": template_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vorlage nicht gefunden")
    updated = await db.project_templates.find_one({"id": template_id})
    return ProjectTemplate(**updated)

@api_router.delete("/project-templates/{template_id}")
async def delete_project_template(template_id: str, current_user: User = Depends(get_current_user)):
    result = await db.project_templates.delete_one({"id": template_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Vorlage nicht gefunden")
    return {"message": "Vorlage gelöscht"}

# Custom Fields CRUD
@api_router.get("/custom-fields")
async def get_custom_fields(entity_type: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {"entity_type": entity_type} if entity_type else {}
    fields = await db.custom_fields.find(query).sort("sort_order", 1).to_list(200)
    return [CustomFieldDef(**f) for f in fields]

@api_router.post("/custom-fields", response_model=CustomFieldDef)
async def create_custom_field(data: CustomFieldDefCreate, current_user: User = Depends(get_current_user)):
    field = CustomFieldDef(**data.dict())
    await db.custom_fields.insert_one(field.dict())
    return field

@api_router.put("/custom-fields/{field_id}", response_model=CustomFieldDef)
async def update_custom_field(field_id: str, data: CustomFieldDefCreate, current_user: User = Depends(get_current_user)):
    result = await db.custom_fields.update_one({"id": field_id}, {"$set": data.dict()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Feld nicht gefunden")
    updated = await db.custom_fields.find_one({"id": field_id})
    return CustomFieldDef(**updated)

@api_router.delete("/custom-fields/{field_id}")
async def delete_custom_field(field_id: str, current_user: User = Depends(get_current_user)):
    result = await db.custom_fields.delete_one({"id": field_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Feld nicht gefunden")
    return {"message": "Feld gelöscht"}

# Event Invitations
@api_router.get("/events/{event_id}/invitations")
async def get_event_invitations(event_id: str, current_user: User = Depends(get_current_user)):
    invitations = await db.event_invitations.find({"event_id": event_id}).sort("sent_at", -1).to_list(500)
    return [EventInvitation(**i) for i in invitations]

@api_router.post("/events/{event_id}/invitations", response_model=EventInvitation)
async def create_event_invitation(event_id: str, data: EventInvitationCreate, current_user: User = Depends(get_current_user)):
    inv = EventInvitation(**{**data.dict(), "event_id": event_id})
    await db.event_invitations.insert_one(inv.dict())
    return inv

@api_router.delete("/invitations/{invitation_id}")
async def delete_invitation(invitation_id: str, current_user: User = Depends(get_current_user)):
    result = await db.event_invitations.delete_one({"id": invitation_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Einladung nicht gefunden")
    return {"message": "Einladung gelöscht"}

@api_router.get("/invitations")
async def get_all_invitations(current_user: User = Depends(get_current_user)):
    invitations = await db.event_invitations.find().sort("sent_at", -1).to_list(1000)
    result = []
    for inv in invitations:
        event = await db.events.find_one({"id": inv.get("event_id")})
        item = EventInvitation(**inv).dict()
        item["event_name"] = event.get("title", event.get("name", "")) if event else ""
        result.append(item)
    return result

# Invoice Share + Online
@api_router.post("/invoices/{invoice_id}/share")
async def share_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    token = str(uuid.uuid4())
    result = await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {"public_token": token}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Rechnung nicht gefunden")
    return {"public_token": token}

# Webhooks CRUD
@api_router.get("/webhooks")
async def get_webhooks(current_user: User = Depends(get_current_user)):
    hooks = await db.webhooks.find().sort("created_at", -1).to_list(100)
    return [Webhook(**h) for h in hooks]

@api_router.post("/webhooks", response_model=Webhook)
async def create_webhook(data: WebhookCreate, current_user: User = Depends(get_current_user)):
    hook = Webhook(**data.dict())
    await db.webhooks.insert_one(hook.dict())
    return hook

@api_router.put("/webhooks/{hook_id}", response_model=Webhook)
async def update_webhook(hook_id: str, data: WebhookCreate, current_user: User = Depends(get_current_user)):
    update = {**data.dict(), "updated_at": datetime.now(timezone.utc)}
    result = await db.webhooks.update_one({"id": hook_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Webhook nicht gefunden")
    updated = await db.webhooks.find_one({"id": hook_id})
    return Webhook(**updated)

@api_router.delete("/webhooks/{hook_id}")
async def delete_webhook(hook_id: str, current_user: User = Depends(get_current_user)):
    result = await db.webhooks.delete_one({"id": hook_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Webhook nicht gefunden")
    return {"message": "Webhook gelöscht"}

# Reports & Export APIs
@api_router.get("/reports/inventory")
async def get_inventory_report(current_user: User = Depends(get_current_user)):
    """Get inventory report as JSON"""
    articles = await db.articles.find().to_list(10000)
    categories = await db.categories.find().to_list(1000)
    
    category_map = {cat["id"]: cat["name"] for cat in categories}
    
    report_data = []
    for article in articles:
        report_data.append({
            "id": article.get("id"),
            "name": article.get("name"),
            "inventory_code": article.get("inventory_code"),
            "category": category_map.get(article.get("category_id", ""), "N/A"),
            "current_stock": article.get("current_stock", 0),
            "min_stock_level": article.get("min_stock_level", 0),
            "price_per_unit": article.get("price_per_unit", 0),
            "rental_price": article.get("rental_price", 0),
            "status": article.get("status", "active")
        })
    
    return {
        "total_items": len(report_data),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "items": report_data
    }

@api_router.get("/reports/customers")
async def get_customers_report(current_user: User = Depends(get_current_user)):
    """Get customers report as JSON"""
    customers = await db.customers.find({"is_active": True}).to_list(10000)
    
    report_data = []
    for customer in customers:
        report_data.append({
            "id": customer.get("id"),
            "customer_number": customer.get("customer_number"),
            "company_name": customer.get("company_name"),
            "contact_person": customer.get("contact_person"),
            "email": customer.get("email"),
            "phone": customer.get("phone"),
            "city": customer.get("address_city", "")
        })
    
    return {
        "total_customers": len(report_data),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "customers": report_data
    }

# =====================================================
# NEUE FEATURES: Berechnungs-Endpoints
# =====================================================

@api_router.post("/calculate/power")
async def calculate_power_requirements(
    article_ids: List[str],
    current_user: User = Depends(get_current_user)
):
    """Berechnet den Gesamtstrombedarf für eine Liste von Artikeln"""
    articles = await db.articles.find({"id": {"$in": article_ids}}).to_list(1000)
    
    total_watt = 0
    power_230v = 0
    power_400v = 0
    power_details = []
    
    for article in articles:
        watt = article.get("power_watt", 0) or 0
        power_type = article.get("power_type", "230V")
        quantity = article.get("current_stock", 1)
        
        article_power = watt * quantity
        total_watt += article_power
        
        if power_type == "400V":
            power_400v += article_power
        else:
            power_230v += article_power
        
        if watt > 0:
            power_details.append({
                "name": article.get("name"),
                "power_watt": watt,
                "power_type": power_type,
                "quantity": quantity,
                "total_watt": article_power
            })
    
    # Berechne benötigte Anschlüsse
    ampere_230v = power_230v / 230 if power_230v > 0 else 0
    ampere_400v = power_400v / 400 if power_400v > 0 else 0
    
    # Standard-Anschlüsse
    schuko_16a = 3680  # 16A * 230V
    cee_16a = 11000    # 16A * 400V * sqrt(3)
    cee_32a = 22000    # 32A * 400V * sqrt(3)
    cee_63a = 43000    # 63A * 400V * sqrt(3)
    
    warnings = []
    if power_230v > schuko_16a:
        warnings.append(f"⚠️ 230V Leistung ({power_230v}W) übersteigt Schuko 16A ({schuko_16a}W)")
    
    return {
        "total_watt": total_watt,
        "power_230v": power_230v,
        "power_400v": power_400v,
        "ampere_230v": round(ampere_230v, 1),
        "ampere_400v": round(ampere_400v, 1),
        "details": power_details,
        "warnings": warnings,
        "recommendations": {
            "schuko_16a_needed": max(1, int(power_230v / schuko_16a) + 1) if power_230v > 0 else 0,
            "cee_32a_needed": max(1, int(power_400v / cee_32a) + 1) if power_400v > 0 else 0
        }
    }

@api_router.post("/calculate/weight")
async def calculate_weight(
    article_ids: List[str],
    current_user: User = Depends(get_current_user)
):
    """Berechnet das Gesamtgewicht für eine Liste von Artikeln"""
    articles = await db.articles.find({"id": {"$in": article_ids}}).to_list(1000)
    
    total_weight = 0
    weight_details = []
    
    for article in articles:
        weight = article.get("weight_kg", 0) or 0
        quantity = article.get("current_stock", 1)
        article_weight = weight * quantity
        total_weight += article_weight
        
        if weight > 0:
            weight_details.append({
                "name": article.get("name"),
                "weight_kg": weight,
                "quantity": quantity,
                "total_kg": article_weight
            })
    
    # LKW Zuladungs-Grenzen
    warnings = []
    if total_weight > 1000:
        warnings.append(f"⚠️ Gewicht ({total_weight}kg) erfordert Transporter >1t Zuladung")
    if total_weight > 2500:
        warnings.append(f"⚠️ Gewicht ({total_weight}kg) erfordert LKW oder mehrere Fahrten")
    
    return {
        "total_weight_kg": round(total_weight, 2),
        "details": weight_details,
        "warnings": warnings,
        "recommendations": {
            "vehicle_type": "PKW" if total_weight < 500 else "Transporter" if total_weight < 1500 else "LKW"
        }
    }

@api_router.post("/calculate/rental-price")
async def calculate_rental_price(
    request: RentalCalculationRequest,
    days: int = 1,
    is_weekend: bool = False,
    current_user: User = Depends(get_current_user)
):
    article_ids = request.article_ids
    quantities = request.quantities
    """Berechnet den Mietpreis mit Tagesfaktoren"""
    articles = await db.articles.find({"id": {"$in": article_ids}}).to_list(1000)

    total_base_price = 0
    total_final_price = 0
    price_details = []
    breakdown = []
    rates_daily_total = 0
    rates_weekend_total = 0
    rates_week_total = 0

    for article in articles:
        rental_price = article.get("rental_price", 0) or 0
        article_id = article.get("id", "")
        quantity = quantities.get(article_id, 1)

        weekend_factor = article.get("rental_factor_weekend", 1.5)
        week_factor = article.get("rental_factor_week", 3.0)

        base = rental_price * quantity
        total_base_price += base

        if days <= 1:
            factor = 1.0
        elif days <= 3:
            factor = weekend_factor if is_weekend else 1.5
        elif days <= 7:
            factor = week_factor
        else:
            factor = week_factor + ((days - 7) * 0.3)

        final = base * factor
        total_final_price += final

        daily_rate = rental_price
        weekend_rate = round(rental_price * weekend_factor, 2)
        week_rate = round(rental_price * week_factor * 7, 2)

        rates_daily_total += daily_rate * quantity
        rates_weekend_total += weekend_rate * quantity
        rates_week_total += week_rate * quantity

        if rental_price > 0:
            price_details.append({
                "name": article.get("name"),
                "daily_price": rental_price,
                "quantity": quantity,
                "days": days,
                "factor": factor,
                "final_price": round(final, 2)
            })
            breakdown.append({
                "article_id": article_id,
                "name": article.get("name"),
                "quantity": quantity,
                "daily_rate": daily_rate,
                "weekend_rate": weekend_rate,
                "week_rate": week_rate,
                "subtotal": round(final, 2)
            })

    return {
        "days": days,
        "is_weekend": is_weekend,
        "base_price": round(total_base_price, 2),
        "factor_applied": round(total_final_price / total_base_price, 2) if total_base_price > 0 else 1,
        "final_price": round(total_final_price, 2),
        "details": price_details,
        "breakdown": breakdown,
        "rates_summary": {
            "daily_total": round(rates_daily_total, 2),
            "weekend_total": round(rates_weekend_total, 2),
            "week_total": round(rates_week_total, 2),
        }
    }

@api_router.put("/articles/{article_id}/operating-hours")
async def update_operating_hours(
    article_id: str,
    hours_to_add: float,
    current_user: User = Depends(get_current_user)
):
    """Aktualisiert die Betriebsstunden eines Artikels"""
    article = await db.articles.find_one({"id": article_id})
    if not article:
        raise HTTPException(status_code=404, detail="Artikel nicht gefunden")
    
    current_hours = article.get("operating_hours", 0) or 0
    new_hours = current_hours + hours_to_add
    max_hours = article.get("max_operating_hours")
    
    update_data = {
        "operating_hours": new_hours,
        "updated_at": datetime.now(timezone.utc)
    }
    
    warnings = []
    if max_hours and new_hours >= max_hours:
        warnings.append(f"⚠️ Maximale Betriebsstunden ({max_hours}h) erreicht - Wartung erforderlich!")
        update_data["status"] = "wartung_erforderlich"
    
    await db.articles.update_one(
        {"id": article_id},
        {"$set": update_data}
    )
    
    return {
        "article_id": article_id,
        "previous_hours": current_hours,
        "hours_added": hours_to_add,
        "new_total_hours": new_hours,
        "max_hours": max_hours,
        "warnings": warnings
    }

# =====================================================
# FEATURE 6: Packliste sortiert nach Lagerort
# =====================================================

@api_router.get("/events/{event_id}/packing-list")
async def get_packing_list(
    event_id: str,
    sort_by: str = "location",
    current_user: User = Depends(get_current_user)
):
    """Generiert eine Packliste sortiert nach Lagerort"""
    # Get all bookings for this event
    bookings = await db.bookings.find({"event_id": event_id}).to_list(1000)
    
    if not bookings:
        return {"items": [], "message": "Keine Buchungen für dieses Event"}
    
    # Get all article IDs
    article_ids = [b.get("article_id") for b in bookings]
    
    # Get articles with location info
    articles = await db.articles.find({"id": {"$in": article_ids}}).to_list(1000)
    
    # Get storage locations
    location_ids = [a.get("storage_location_id") for a in articles if a.get("storage_location_id")]
    locations = await db.storage_locations.find({"id": {"$in": location_ids}}).to_list(1000)
    location_map = {loc["id"]: loc for loc in locations}
    
    # Get zones
    zone_ids = [loc.get("zone_id") for loc in locations if loc.get("zone_id")]
    zones = await db.storage_zones.find({"id": {"$in": zone_ids}}).to_list(1000)
    zone_map = {z["id"]: z for z in zones}
    
    # Build packing list
    packing_items = []
    for booking in bookings:
        article = next((a for a in articles if a.get("id") == booking.get("article_id")), None)
        if not article:
            continue
        
        location_id = article.get("storage_location_id")
        location = location_map.get(location_id, {}) if location_id else {}
        zone_id = location.get("zone_id")
        zone = zone_map.get(zone_id, {}) if zone_id else {}
        
        packing_items.append({
            "article_id": article.get("id"),
            "name": article.get("name"),
            "inventory_code": article.get("inventory_code"),
            "quantity": booking.get("quantity", 1),
            "zone_name": zone.get("name", "Unbekannt"),
            "zone_order": zone.get("name", "ZZZ"),  # For sorting
            "location_name": location.get("name", "Kein Lagerort"),
            "location_order": location.get("name", "ZZZ"),
            "weight_kg": article.get("weight_kg", 0) or 0,
            "checked": False
        })
    
    # Sort by zone, then location
    if sort_by == "location":
        packing_items.sort(key=lambda x: (x["zone_order"], x["location_order"]))
    elif sort_by == "name":
        packing_items.sort(key=lambda x: x["name"])
    
    # Calculate totals
    total_weight = sum(item["weight_kg"] * item["quantity"] for item in packing_items)
    
    return {
        "event_id": event_id,
        "total_items": len(packing_items),
        "total_weight_kg": round(total_weight, 2),
        "items": packing_items,
        "sorted_by": sort_by
    }

# ============================================
# LAGER-LOGISTIK: Packing List Check-Out/Check-In
# ============================================

@api_router.get("/events/{event_id}/packing-list-items")
async def get_packing_list_items(event_id: str, current_user: User = Depends(get_current_user)):
    """Get all packing list items with their check-out/check-in status"""
    items = await db.packing_list_items.find({"event_id": event_id}).sort(
        [("storage_location", 1), ("zone_name", 1)]
    ).to_list(1000)
    
    # If no items exist yet, create them from bookings
    if not items:
        bookings = await db.bookings.find({"event_id": event_id}).to_list(1000)
        articles = await db.articles.find().to_list(10000)
        locations = await db.storage_locations.find().to_list(10000)
        zones = await db.storage_zones.find().to_list(1000)
        
        location_map = {loc["id"]: loc for loc in locations}
        zone_map = {zone["id"]: zone for zone in zones}
        
        for booking in bookings:
            article = next((a for a in articles if a.get("id") == booking.get("article_id")), None)
            if not article:
                continue
            
            location_id = article.get("storage_location_id")
            location = location_map.get(location_id, {}) if location_id else {}
            zone_id = location.get("zone_id")
            zone = zone_map.get(zone_id, {}) if zone_id else {}
            
            item = PackingListItem(
                event_id=event_id,
                booking_id=booking.get("id"),
                article_id=article.get("id"),
                quantity=booking.get("quantity", 1)
            )
            
            await db.packing_list_items.insert_one(item.model_dump())

        items = await db.packing_list_items.find({"event_id": event_id}).sort(
            [("storage_location", 1), ("zone_name", 1)]
        ).to_list(1000)
    
    # Enrich with article and location info
    articles = await db.articles.find().to_list(10000)
    locations = await db.storage_locations.find().to_list(10000)
    zones = await db.storage_zones.find().to_list(1000)
    
    # Remove MongoDB ObjectId from all documents to prevent JSON serialization errors
    for a in articles:
        a.pop("_id", None)
    for loc in locations:
        loc.pop("_id", None)
    for z in zones:
        z.pop("_id", None)
    
    article_map = {a["id"]: a for a in articles}
    location_map = {loc["id"]: loc for loc in locations}
    zone_map = {zone["id"]: zone for zone in zones}
    
    enriched_items = []
    for item in items:
        item.pop("_id", None)  # Remove MongoDB ObjectId — not JSON-serializable
        article = article_map.get(item.get("article_id"), {})
        location_id = article.get("storage_location_id")
        location = location_map.get(location_id, {}) if location_id else {}
        zone_id = location.get("zone_id")
        zone = zone_map.get(zone_id, {}) if zone_id else {}
        
        enriched_items.append({
            **item,
            "article_name": article.get("name", "Unbekannt"),
            "inventory_code": article.get("inventory_code", ""),
            "zone_name": zone.get("name", "Kein Lagerort"),
            "zone_order": zone.get("name", "ZZZ"),
            "location_name": location.get("name", ""),
            "weight_kg": article.get("weight_kg", 0) or 0,
            "image_base64": article.get("image_base64"),
        })
    
    # Sort by zone, then location
    enriched_items.sort(key=lambda x: (x["zone_order"], x["location_name"]))
    
    # Calculate stats
    total_items = len(enriched_items)
    checked_out = sum(1 for i in enriched_items if i.get("checked_out"))
    checked_in = sum(1 for i in enriched_items if i.get("checked_in"))
    missing = sum(1 for i in enriched_items if i.get("checkin_condition") == "MISSING")
    defect = sum(1 for i in enriched_items if i.get("checkin_condition") == "DEFECT")
    dirty = sum(1 for i in enriched_items if i.get("checkin_condition") == "DIRTY")
    
    return {
        "event_id": event_id,
        "items": enriched_items,
        "stats": {
            "total": total_items,
            "checked_out": checked_out,
            "checked_in": checked_in,
            "missing": missing,
            "defect": defect,
            "dirty": dirty,
            "pending_checkout": total_items - checked_out,
            "pending_checkin": checked_out - checked_in
        }
    }

@api_router.post("/packing-list/checkout")
async def checkout_items(checkout: PackingListCheckout, current_user: User = Depends(get_current_user)):
    """Check out multiple items from the packing list"""
    now = datetime.now(timezone.utc)
    
    updated_count = 0
    for item_id in checkout.item_ids:
        result = await db.packing_list_items.update_one(
            {"id": item_id},
            {"$set": {
                "checked_out": True,
                "checked_out_by": current_user.username,
                "checked_out_at": now,
                "checkout_condition": checkout.condition,
                "checkout_notes": checkout.notes
            }}
        )
        if result.modified_count > 0:
            updated_count += 1
    
    return {"success": True, "updated": updated_count}

@api_router.post("/packing-list/checkout-all/{event_id}")
async def checkout_all_items(event_id: str, current_user: User = Depends(get_current_user)):
    """Check out all items for an event"""
    now = datetime.now(timezone.utc)
    
    result = await db.packing_list_items.update_many(
        {"event_id": event_id, "checked_out": False},
        {"$set": {
            "checked_out": True,
            "checked_out_by": current_user.username,
            "checked_out_at": now,
            "checkout_condition": "OK"
        }}
    )
    
    return {"success": True, "updated": result.modified_count}

@api_router.post("/packing-list/checkin")
async def checkin_item(checkin: PackingListCheckin, current_user: User = Depends(get_current_user)):
    """Check in a single item with condition assessment"""
    now = datetime.now(timezone.utc)
    
    # Get the item first
    item = await db.packing_list_items.find_one({"id": checkin.item_id})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # Update the item
    update_data = {
        "checked_in": True,
        "checked_in_by": current_user.username,
        "checked_in_at": now,
        "checkin_condition": checkin.condition,
        "checkin_notes": checkin.notes,
        "checkin_photos": checkin.photos
    }
    
    await db.packing_list_items.update_one(
        {"id": checkin.item_id},
        {"$set": update_data}
    )
    
    # F3: Release serial numbers assigned to the booking back to "verfügbar"
    booking_id = item.get("booking_id")
    if booking_id:
        booking = await db.bookings.find_one({"id": booking_id})
        if booking and booking.get("serial_number_ids"):
            await db.serial_numbers.update_many(
                {"id": {"$in": booking["serial_number_ids"]}},
                {"$set": {"status": "verfügbar", "updated_at": datetime.now(timezone.utc).isoformat()}}
            )

    # If defect, create a maintenance task
    if checkin.condition == "DEFECT":
        article = await db.articles.find_one({"id": item.get("article_id")})
        maintenance_task = MaintenanceTask(
            article_id=item.get("article_id"),
            title=f"Defekt gemeldet: {article.get('name', 'Unbekannt') if article else 'Unbekannt'}",
            description=f"Bei Rückgabe als defekt markiert. Notiz: {checkin.notes or 'Keine'}",
            task_type="Reparatur",
            priority="high",
            status="pending",
            created_by=current_user.username
        )
        await db.maintenance_tasks.insert_one(maintenance_task.model_dump())

    return {"success": True, "condition": checkin.condition}

@api_router.post("/packing-list/checkin-all/{event_id}")
async def checkin_all_items(event_id: str, condition: str = "OK", current_user: User = Depends(get_current_user)):
    """Check in all checked-out items as OK"""
    now = datetime.now(timezone.utc)
    
    result = await db.packing_list_items.update_many(
        {"event_id": event_id, "checked_out": True, "checked_in": False},
        {"$set": {
            "checked_in": True,
            "checked_in_by": current_user.username,
            "checked_in_at": now,
            "checkin_condition": condition
        }}
    )
    
    return {"success": True, "updated": result.modified_count}

@api_router.get("/packing-list/missing/{event_id}")
async def get_missing_items(event_id: str, current_user: User = Depends(get_current_user)):
    """Get all items that are checked out but not checked in (potentially missing)"""
    items = await db.packing_list_items.find({
        "event_id": event_id,
        "checked_out": True,
        "checked_in": False
    }).to_list(1000)
    
    # Also get items marked as MISSING
    missing_items = await db.packing_list_items.find({
        "event_id": event_id,
        "checkin_condition": "MISSING"
    }).to_list(1000)
    
    # Combine and deduplicate
    all_missing_ids = set(i["id"] for i in items) | set(i["id"] for i in missing_items)
    
    # Enrich with article info
    articles = await db.articles.find().to_list(10000)
    article_map = {a["id"]: a for a in articles}
    
    result = []
    for item in items + missing_items:
        if item["id"] in all_missing_ids:
            all_missing_ids.discard(item["id"])  # Avoid duplicates
            article = article_map.get(item.get("article_id"), {})
            result.append({
                **item,
                "article_name": article.get("name", "Unbekannt"),
                "inventory_code": article.get("inventory_code", ""),
            })
    
    return {
        "event_id": event_id,
        "missing_count": len(result),
        "items": result
    }

@api_router.delete("/packing-list/reset/{event_id}")
async def reset_packing_list(event_id: str, current_user: User = Depends(require_permission(Permission.ADMIN_ACCESS))):
    """Reset packing list for an event (admin only)"""
    
    result = await db.packing_list_items.delete_many({"event_id": event_id})
    return {"success": True, "deleted": result.deleted_count}

@api_router.post("/events/{event_id}/packing-list/sign")
async def sign_packing_list(event_id: str, body: dict, current_user: User = Depends(get_current_user)):
    """Save sign-off signature for a packing list"""
    await db.events.update_one(
        {"id": event_id},
        {"$set": {
            "packing_list_signature": body.get("signature", ""),
            "packing_list_signed_by": body.get("signed_by", ""),
            "packing_list_signed_at": datetime.now(timezone.utc),
        }}
    )
    return {"message": "Packliste abgezeichnet"}

@api_router.get("/reports/inventory-csv")
async def export_inventory_csv(current_user: User = Depends(get_current_user)):
    """Export inventory as CSV"""
    articles = await db.articles.find({"deleted": {"$ne": True}}).to_list(MAX_EXPORT_ROWS)
    
    csv_lines = ["Name,Inventory Code,Category,Stock,Min Stock,Price,Rental Price"]
    
    for article in articles:
        category = await db.categories.find_one({"id": article.get("category_id", "")})
        category_name = category["name"] if category else "N/A"
        
        line = f'"{article["name"]}","{article["inventory_code"]}","{category_name}",{article.get("current_stock", 0)},{article.get("min_stock_level", 0)},{article.get("price_per_unit", 0)},{article.get("rental_price", 0) or 0}'
        csv_lines.append(line)
    
    csv_data = "\n".join(csv_lines)
    return Response(content=csv_data, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=inventory.csv"})

@api_router.get("/reports/customers-csv")
async def export_customers_csv(current_user: User = Depends(get_current_user)):
    """Export customers as CSV"""
    customers = await db.customers.find({"is_active": True}).to_list(MAX_EXPORT_ROWS)
    
    csv_lines = ["Customer Number,Company,Contact,Email,Phone,City"]
    
    for customer in customers:
        line = f'"{customer.get("customer_number", "")}","{customer.get("company_name", "")}","{customer.get("contact_person", "")}","{customer.get("email", "")}","{customer.get("phone", "")}","{customer.get("address_city", "")}"'
        csv_lines.append(line)
    
    csv_data = "\n".join(csv_lines)
    return Response(content=csv_data, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=customers.csv"})

@api_router.get("/reports/monthly")
async def generate_monthly_report(
    period: str = "current",
    current_user: User = Depends(get_current_user)
):
    """Generate monthly report"""
    from datetime import datetime
    now = datetime.now(timezone.utc)
    if period == "current":
        start = datetime(now.year, now.month, 1)
        if now.month == 12:
            end = datetime(now.year + 1, 1, 1)
        else:
            end = datetime(now.year, now.month + 1, 1)
    else:  # last month
        if now.month == 1:
            start = datetime(now.year - 1, 12, 1)
            end = datetime(now.year, 1, 1)
        else:
            start = datetime(now.year, now.month - 1, 1)
            end = datetime(now.year, now.month, 1)
    
    events = await db.events.find({
        "start_date": {"$gte": start, "$lt": end}
    }).to_list(1000)
    
    bookings = await db.bookings.find({
        "created_at": {"$gte": start, "$lt": end}
    }).to_list(1000)
    
    # Calculate revenue (simplified)
    total_revenue = 0.0
    for booking in bookings:
        article = await db.articles.find_one({"id": booking["article_id"]})
        if article and article.get("rental_price"):
            total_revenue += article["rental_price"] * booking["quantity"]
    
    return {
        "period": period,
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "total_events": len(events),
        "total_bookings": len(bookings),
        "total_revenue": round(total_revenue, 2)
    }


@api_router.get("/reports/inventory-pdf")
async def export_inventory_pdf(current_user: User = Depends(get_current_user)):
    """Export inventory as PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    
    articles = await db.articles.find({"deleted": {"$ne": True}}).to_list(MAX_EXPORT_ROWS)

    # Bulk-fetch all categories in a single query — avoids N+1 (one query per article)
    category_ids = list({a.get("category_id") for a in articles if a.get("category_id")})
    categories_raw = await db.categories.find({"id": {"$in": category_ids}}).to_list(len(category_ids) or 1)
    category_map = {c["id"]: c["name"] for c in categories_raw}

    # Create PDF in memory
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#007AFF'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    # Title
    title = Paragraph("Inventarliste", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))

    # Date
    date_text = Paragraph(f"Erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal'])
    elements.append(date_text)
    elements.append(Spacer(1, 1*cm))

    # Table data
    data = [['Name', 'Inventarcode', 'Kategorie', 'Bestand', 'Min. Bestand', 'Preis (€)', 'Mietpreis (€)']]

    for article in articles:
        category_name = category_map.get(article.get("category_id", ""), "N/A")
        
        row = [
            article["name"][:30],
            article["inventory_code"],
            category_name[:20],
            str(article.get("current_stock", 0)),
            str(article.get("min_stock_level", 0)),
            f"{article.get('price_per_unit', 0):.2f}",
            f"{article.get('rental_price', 0) or 0:.2f}"
        ]
        data.append(row)
    
    # Create table
    table = Table(data, colWidths=[6*cm, 3*cm, 4*cm, 2*cm, 2.5*cm, 2*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007AFF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 1*cm))
    
    # Footer
    footer = Paragraph(f"Gesamt: {len(articles)} Artikel", styles['Normal'])
    elements.append(footer)
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF data
    pdf_data = buffer.getvalue()
    buffer.close()
    
    return Response(
        content=pdf_data,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=inventar.pdf"}
    )

@api_router.get("/reports/customers-pdf")
async def export_customers_pdf(current_user: User = Depends(get_current_user)):
    """Export customers as PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    
    customers = await db.customers.find({"is_active": True}).to_list(MAX_EXPORT_ROWS)

    # Create PDF in memory
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#34C759'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Title
    title = Paragraph("Kundenliste", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))
    
    # Date
    date_text = Paragraph(f"Erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal'])
    elements.append(date_text)
    elements.append(Spacer(1, 1*cm))
    
    # Table data
    data = [['Kundennr.', 'Firma', 'Kontakt', 'E-Mail', 'Telefon', 'Stadt']]
    
    for customer in customers:
        row = [
            customer.get("customer_number", "")[:15],
            customer.get("company_name", "")[:30],
            customer.get("contact_person", "")[:25],
            customer.get("email", "")[:30],
            customer.get("phone", "")[:20],
            customer.get("address_city", "")[:20]
        ]
        data.append(row)
    
    # Create table
    table = Table(data, colWidths=[3*cm, 5*cm, 4*cm, 5*cm, 3.5*cm, 3.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34C759')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 1*cm))
    
    # Footer
    footer = Paragraph(f"Gesamt: {len(customers)} Kunden", styles['Normal'])
    elements.append(footer)
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF data
    pdf_data = buffer.getvalue()
    buffer.close()
    
    return Response(
        content=pdf_data,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=kunden.pdf"}
    )


# ============================================================
# DOKUMENTE & EXPORT - Packlisten-PDF, Lademeter, Excel
# ============================================================

@api_router.get("/events/{event_id}/packing-list/pdf")
async def export_packing_list_pdf(
    event_id: str,
    current_user: User = Depends(get_current_user)
):
    """Export packing list as PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from io import BytesIO
    
    try:
        # Get event
        event = await db.events.find_one({"id": event_id})
        if not event:
            raise HTTPException(status_code=404, detail="Event nicht gefunden")
        
        # Get customer
        customer = await db.customers.find_one({"id": event.get("customer_id")})
        
        # Get bookings
        bookings = await db.bookings.find({"event_id": event_id, "status": "booked"}).to_list(500)
        
        # Build items list with article details
        items = []
        total_weight = 0.0
        total_power = 0
        
        for booking in bookings:
            article = await db.articles.find_one({"id": booking.get("article_id")})
            if article:
                location = await db.storage_locations.find_one({"id": article.get("storage_location_id")})
                weight = (article.get("weight_kg") or 0) * booking.get("quantity", 1)
                power = (article.get("power_watt") or 0) * booking.get("quantity", 1)
                total_weight += weight
                total_power += power
                
                items.append({
                    "name": article.get("name"),
                    "code": article.get("inventory_code"),
                    "qty": booking.get("quantity", 1),
                    "location": location.get("name") if location else "-",
                    "weight": f"{weight:.1f} kg",
                    "status": "☐"  # Checkbox for packing
                })
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=20)
        elements.append(Paragraph("📦 Packliste", title_style))
        
        # Event Info
        info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, spaceAfter=5)
        elements.append(Paragraph(f"<b>Event:</b> {event.get('event_name')} ({event.get('event_number')})", info_style))
        elements.append(Paragraph(f"<b>Kunde:</b> {customer.get('company_name') if customer else 'N/A'}", info_style))
        
        # Format dates safely
        start_date = event.get('start_date', '')
        end_date = event.get('end_date', '')
        if hasattr(start_date, 'strftime'):
            start_date = start_date.strftime('%d.%m.%Y')
        elif isinstance(start_date, str) and len(start_date) >= 10:
            start_date = start_date[:10]
        if hasattr(end_date, 'strftime'):
            end_date = end_date.strftime('%d.%m.%Y')
        elif isinstance(end_date, str) and len(end_date) >= 10:
            end_date = end_date[:10]
        
        elements.append(Paragraph(f"<b>Datum:</b> {start_date} - {end_date}", info_style))
        elements.append(Paragraph(f"<b>Ort:</b> {event.get('location')}", info_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # Summary
        elements.append(Paragraph(f"<b>Gesamt:</b> {len(items)} Positionen | {total_weight:.1f} kg | {total_power} W", info_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # Items Table
        table_data = [["✓", "Artikel", "Code", "Menge", "Lagerort", "Gewicht"]]
        for item in items:
            table_data.append([item["status"], item["name"][:30], item["code"], str(item["qty"]), item["location"], item["weight"]])
        
        table = Table(table_data, colWidths=[1*cm, 7*cm, 3*cm, 1.5*cm, 3*cm, 2*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007AFF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 0), (3, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 1*cm))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
        elements.append(Paragraph(f"Erstellt am: {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')} | Inventar System", footer_style))
        
        doc.build(elements)
        pdf_data = buffer.getvalue()
        
        filename = f"packliste_{event.get('event_number', 'event')}.pdf"
        return Response(
            content=pdf_data,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error generating packing list PDF: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


# ===========================================
# DELIVERY CONFIRMATION WITH DIGITAL SIGNATURE
# ===========================================

class DeliveryConfirmation(BaseModel):
    signature: str  # Base64 SVG signature
    signed_by: str
    signed_at: str
    confirmed_items: List[str]  # List of article IDs

@api_router.post("/events/{event_id}/delivery-confirmation")
async def confirm_delivery(
    event_id: str,
    confirmation: DeliveryConfirmation,
    current_user: User = Depends(get_current_user)
):
    """Save delivery confirmation with digital signature"""
    try:
        event = await db.events.find_one({"id": event_id})
        if not event:
            raise HTTPException(status_code=404, detail="Event nicht gefunden")
        
        # Create delivery record
        delivery_record = {
            "id": str(uuid.uuid4()),
            "event_id": event_id,
            "signature": confirmation.signature,
            "signed_by": confirmation.signed_by,
            "signed_at": confirmation.signed_at,
            "confirmed_items": confirmation.confirmed_items,
            "confirmed_by_user": current_user.id,
            "created_at": datetime.now(timezone.utc)
        }
        
        await db.delivery_confirmations.insert_one(delivery_record)
        
        # Update event status
        await db.events.update_one(
            {"id": event_id},
            {"$set": {"delivery_status": "confirmed", "delivery_confirmed_at": datetime.now(timezone.utc)}}
        )
        
        # Update booking status for confirmed items
        await db.bookings.update_many(
            {"event_id": event_id, "article_id": {"$in": confirmation.confirmed_items}},
            {"$set": {"status": "delivered", "delivered_at": datetime.now(timezone.utc)}}
        )
        
        return {"status": "success", "delivery_id": delivery_record["id"]}
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error confirming delivery: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


@api_router.get("/events/{event_id}/delivery-status")
async def get_delivery_status(
    event_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get delivery confirmation status for an event"""
    try:
        confirmation = await db.delivery_confirmations.find_one({"event_id": event_id})
        if confirmation:
            confirmation.pop("_id", None)
            return {"status": "confirmed", "confirmation": confirmation}
        return {"status": "pending", "confirmation": None}
    except Exception as e:
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


# ===========================================
# BOOKING CONFLICT & OVERBOOKING WARNINGS
# ===========================================

@api_router.get("/articles/{article_id}/availability-check")
async def check_article_availability(
    article_id: str,
    start_date: str,
    end_date: str,
    quantity: int = 1,
    current_user: User = Depends(get_current_user)
):
    """Check if article is available and warn about conflicts"""
    try:
        article = await db.articles.find_one({"id": article_id})
        if not article:
            raise HTTPException(status_code=404, detail="Artikel nicht gefunden")
        
        # Parse dates — raises 400 on invalid format
        start_dt = parse_iso_date(start_date, "start_date")
        end_dt = parse_iso_date(end_date, "end_date")
        
        # Find conflicting bookings
        conflicts = await db.bookings.find({
            "article_id": article_id,
            "status": {"$in": ["booked", "confirmed", "delivered"]},
            "$or": [
                {"start_date": {"$lte": end_dt}, "end_date": {"$gte": start_dt}},
            ]
        }).to_list(100)
        
        # Calculate booked quantity in period
        booked_quantity = sum(b.get("quantity", 1) for b in conflicts)
        available_quantity = article.get("current_stock", 0) - booked_quantity
        
        # Get conflict details
        conflict_details = []
        for conflict in conflicts:
            event = await db.events.find_one({"id": conflict.get("event_id")})
            conflict_details.append({
                "event_id": conflict.get("event_id"),
                "event_name": event.get("event_name") if event else "Unbekannt",
                "quantity": conflict.get("quantity", 1),
                "start_date": conflict.get("start_date"),
                "end_date": conflict.get("end_date")
            })
        
        # Check for shortage
        is_available = available_quantity >= quantity
        shortage = max(0, quantity - available_quantity)
        
        return {
            "article_id": article_id,
            "article_name": article.get("name"),
            "requested_quantity": quantity,
            "current_stock": article.get("current_stock", 0),
            "booked_quantity": booked_quantity,
            "available_quantity": max(0, available_quantity),
            "is_available": is_available,
            "shortage": shortage,
            "conflicts": conflict_details,
            "suggestion": f"Zumieten: {shortage} Stück" if shortage > 0 else None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error checking availability: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


@api_router.get("/scanner/article-warning/{inventory_code}")
async def get_scanner_warning(
    inventory_code: str,
    current_user: User = Depends(get_current_user)
):
    """Get warnings when scanning an article (conflicts with other events)"""
    try:
        article = await db.articles.find_one({"inventory_code": inventory_code})
        if not article:
            return {"warning": None, "article": None}
        
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        tomorrow = today + timedelta(days=1)
        
        # Check if article is booked for today
        active_booking = await db.bookings.find_one({
            "article_id": article["id"],
            "status": {"$in": ["booked", "confirmed"]},
            "start_date": {"$lte": tomorrow},
            "end_date": {"$gte": today}
        })
        
        if active_booking:
            event = await db.events.find_one({"id": active_booking.get("event_id")})
            return {
                "warning": {
                    "type": "conflict",
                    "message": f"⚠️ Dieser Artikel ist heute für '{event.get('event_name') if event else 'ein Event'}' gebucht!",
                    "event_id": active_booking.get("event_id"),
                    "event_name": event.get("event_name") if event else "Unbekannt",
                    "quantity_booked": active_booking.get("quantity", 1)
                },
                "article": {
                    "id": article["id"],
                    "name": article.get("name"),
                    "inventory_code": article.get("inventory_code"),
                    "current_stock": article.get("current_stock", 0),
                    "status": article.get("status")
                }
            }
        
        return {
            "warning": None,
            "article": {
                "id": article["id"],
                "name": article.get("name"),
                "inventory_code": article.get("inventory_code"),
                "current_stock": article.get("current_stock", 0),
                "status": article.get("status")
            }
        }
        
    except Exception as e:
        logging.error(f"Error getting scanner warning: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


@api_router.get("/overbooking-alerts")
async def get_overbooking_alerts(
    days_ahead: int = 30,
    current_user: User = Depends(get_current_user)
):
    """Check for material shortages (overbooking) across all events in the next X days"""
    try:
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = today + timedelta(days=days_ahead)
        
        # Get all active bookings in the period
        # Buchungen verwenden pickup_date/return_date (nicht start_date/end_date)
        bookings = await db.bookings.find({
            "status": {"$in": ["booked", "confirmed"]},
            "pickup_date": {"$lte": end_date},
            "return_date": {"$gte": today}
        }).to_list(5000)

        # Group bookings by date and article
        from collections import defaultdict
        date_article_demand = defaultdict(lambda: defaultdict(int))

        for booking in bookings:
            start = booking.get("pickup_date")
            end = booking.get("return_date")
            article_id = booking.get("article_id")
            qty = booking.get("quantity", 1)
            
            try:
                if isinstance(start, str):
                    start = datetime.fromisoformat(start.replace('Z', '+00:00'))
                if isinstance(end, str):
                    end = datetime.fromisoformat(end.replace('Z', '+00:00'))
            except (ValueError, AttributeError):
                logging.error(f"Buchung {booking.get('id')} hat nicht-parsebares Datum – wird übersprungen.")
                continue

            # Ungültige Buchungen überspringen – verhindert Endlosschleife
            if not start or not end or start > end:
                logging.warning(f"Buchung {booking.get('id')} hat ungültige Datumsangaben (start={start}, end={end}) – wird übersprungen.")
                continue

            # Mark demand for each day of the booking
            current = start
            while current <= end:
                date_key = current.strftime("%Y-%m-%d")
                date_article_demand[date_key][article_id] += qty
                current += timedelta(days=1)
        
        # F5: Pre-fetch all unique articles and events in bulk (eliminates N+1)
        all_article_ids = list({b.get("article_id") for b in bookings if b.get("article_id")})
        all_event_ids = list({b.get("event_id") for b in bookings if b.get("event_id")})
        _art_list = await db.articles.find({"id": {"$in": all_article_ids}}).to_list(len(all_article_ids) + 1)
        _evt_list = await db.events.find({"id": {"$in": all_event_ids}}).to_list(len(all_event_ids) + 1)
        articles_cache = {a["id"]: a for a in _art_list}
        events_cache = {e["id"]: e for e in _evt_list}

        # Check against stock levels
        alerts = []

        for date_key, article_demands in date_article_demand.items():
            for article_id, total_demand in article_demands.items():
                article = articles_cache.get(article_id)

                if not article:
                    continue

                stock = article.get("current_stock", 0)
                if total_demand > stock:
                    shortage = total_demand - stock

                    conflicting_bookings = [b for b in bookings
                                           if b.get("article_id") == article_id]
                    event_ids = list(set(b.get("event_id") for b in conflicting_bookings))
                    event_names = [events_cache[eid].get("event_name", "?")
                                   for eid in event_ids if eid in events_cache]
                    
                    alerts.append({
                        "date": date_key,
                        "article_id": article_id,
                        "article_name": article.get("name"),
                        "inventory_code": article.get("inventory_code"),
                        "current_stock": stock,
                        "total_demand": total_demand,
                        "shortage": shortage,
                        "conflicting_events": event_names[:3],  # Max 3 event names
                        "severity": "critical" if shortage >= stock else "warning"
                    })
        
        # Sort by date, then severity
        alerts.sort(key=lambda x: (x["date"], x["severity"] == "warning"))
        
        # Remove duplicate alerts (same article, same shortage)
        seen = set()
        unique_alerts = []
        for alert in alerts:
            key = (alert["article_id"], alert["shortage"])
            if key not in seen:
                seen.add(key)
                unique_alerts.append(alert)
        
        return {
            "total_alerts": len(unique_alerts),
            "critical_count": len([a for a in unique_alerts if a["severity"] == "critical"]),
            "warning_count": len([a for a in unique_alerts if a["severity"] == "warning"]),
            "alerts": unique_alerts[:50],  # Limit to 50 alerts
            "period": {"start": today.isoformat(), "end": end_date.isoformat()}
        }
        
    except Exception as e:
        logging.error(f"Error checking overbooking: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


@api_router.get("/events/{event_id}/loading-calculation")
async def calculate_loading_meters(
    event_id: str,
    truck_width_m: float = 2.4,
    truck_height_m: float = 2.7,
    current_user: User = Depends(get_current_user)
):
    """Calculate loading meters (Lademeter) for LKW planning"""
    try:
        event = await db.events.find_one({"id": event_id})
        if not event:
            raise HTTPException(status_code=404, detail="Event nicht gefunden")
        
        bookings = await db.bookings.find({"event_id": event_id, "status": "booked"}).to_list(500)

        # F5: Bulk-fetch articles via $in query
        b_article_ids = list({b.get("article_id") for b in bookings if b.get("article_id")})
        b_articles_list = await db.articles.find({"id": {"$in": b_article_ids}}).to_list(len(b_article_ids) + 1)
        b_articles_map = {a["id"]: a for a in b_articles_list}

        total_weight_kg = 0.0
        total_volume_m3 = 0.0
        items_detail = []

        for booking in bookings:
            article = b_articles_map.get(booking.get("article_id") or "")
            if article:
                qty = booking.get("quantity", 1)
                weight = (article.get("weight_kg") or 0) * qty
                
                # Estimate volume (if not stored, use rough estimate based on weight)
                # Typical equipment density: ~200-400 kg/m³
                estimated_volume = weight / 300 if weight > 0 else 0.01 * qty
                
                total_weight_kg += weight
                total_volume_m3 += estimated_volume
                
                items_detail.append({
                    "article_name": article.get("name"),
                    "quantity": qty,
                    "weight_kg": weight,
                    "estimated_volume_m3": round(estimated_volume, 3)
                })
        
        # Lademeter calculation
        # 1 Lademeter = truck_width * truck_height * 1m length
        lademeter_volume = truck_width_m * truck_height_m  # m³ per Lademeter
        loading_meters = total_volume_m3 / lademeter_volume if lademeter_volume > 0 else 0
        
        # Standard truck types
        truck_recommendations = []
        if loading_meters <= 2:
            truck_recommendations.append({"type": "Sprinter/Transporter", "capacity": "2-3 LDM", "suitable": True})
        if loading_meters <= 5:
            truck_recommendations.append({"type": "7.5t LKW", "capacity": "5-6 LDM", "suitable": loading_meters <= 5})
        if loading_meters <= 10:
            truck_recommendations.append({"type": "12t LKW", "capacity": "8-10 LDM", "suitable": loading_meters <= 10})
        if loading_meters <= 13.6:
            truck_recommendations.append({"type": "Sattelzug", "capacity": "13.6 LDM", "suitable": True})
        
        return {
            "event_id": event_id,
            "event_name": event.get("event_name"),
            "calculation": {
                "total_weight_kg": round(total_weight_kg, 1),
                "total_volume_m3": round(total_volume_m3, 2),
                "loading_meters": round(loading_meters, 2),
                "truck_dimensions": {
                    "width_m": truck_width_m,
                    "height_m": truck_height_m
                }
            },
            "truck_recommendations": truck_recommendations,
            "items_count": len(items_detail),
            "items": items_detail
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error calculating loading meters: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


@api_router.get("/reports/inventory-excel")
async def export_inventory_excel(current_user: User = Depends(get_current_user)):
    """Export inventory as Excel file"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    try:
        articles = await db.articles.find({"deleted": {"$ne": True}}).to_list(MAX_EXPORT_ROWS)
        categories = await db.categories.find().to_list(1000)
        category_map = {cat["id"]: cat["name"] for cat in categories}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventar"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="007AFF", end_color="007AFF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["Inventar-Nr.", "Name", "Kategorie", "Bestand", "Min-Bestand", "Status", "Mietpreis/Tag", "Gewicht (kg)", "Leistung (W)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data
        for row, article in enumerate(articles, 2):
            ws.cell(row=row, column=1, value=article.get("inventory_code", "")).border = thin_border
            ws.cell(row=row, column=2, value=article.get("name", "")).border = thin_border
            ws.cell(row=row, column=3, value=category_map.get(article.get("category_id"), "")).border = thin_border
            ws.cell(row=row, column=4, value=article.get("current_stock", 0)).border = thin_border
            ws.cell(row=row, column=5, value=article.get("min_stock_level", 0)).border = thin_border
            ws.cell(row=row, column=6, value=article.get("status", "OK")).border = thin_border
            ws.cell(row=row, column=7, value=article.get("rental_price") or article.get("rental_price_day") or 0).border = thin_border
            ws.cell(row=row, column=8, value=article.get("weight_kg") or 0).border = thin_border
            ws.cell(row=row, column=9, value=article.get("power_watt") or 0).border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=inventar_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"}
        )
        
    except Exception as e:
        logging.error(f"Error generating Excel export: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


@api_router.get("/reports/events-excel")
async def export_events_excel(current_user: User = Depends(get_current_user)):
    """Export events with bookings as Excel file"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    try:
        events = await db.events.find().sort("start_date", -1).to_list(1000)
        customers = await db.customers.find({"is_active": True}).to_list(1000)
        customer_map = {c["id"]: c.get("company_name") for c in customers}
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Events"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="5856D6", end_color="5856D6", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        headers = ["Event-Nr.", "Name", "Kunde", "Typ", "Datum Start", "Datum Ende", "Ort", "Status", "Wert (€)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        for row, event in enumerate(events, 2):
            start_date = event.get("start_date", "")
            end_date = event.get("end_date", "")
            if hasattr(start_date, 'strftime'):
                start_date = start_date.strftime('%d.%m.%Y')
            elif isinstance(start_date, str) and len(start_date) >= 10:
                start_date = start_date[:10]
            if hasattr(end_date, 'strftime'):
                end_date = end_date.strftime('%d.%m.%Y')
            elif isinstance(end_date, str) and len(end_date) >= 10:
                end_date = end_date[:10]
            
            ws.cell(row=row, column=1, value=event.get("event_number", "")).border = thin_border
            ws.cell(row=row, column=2, value=event.get("event_name", "")).border = thin_border
            ws.cell(row=row, column=3, value=customer_map.get(event.get("customer_id"), "")).border = thin_border
            ws.cell(row=row, column=4, value=event.get("event_type", "")).border = thin_border
            ws.cell(row=row, column=5, value=start_date).border = thin_border
            ws.cell(row=row, column=6, value=end_date).border = thin_border
            ws.cell(row=row, column=7, value=event.get("location", "")).border = thin_border
            ws.cell(row=row, column=8, value=event.get("status", "")).border = thin_border
            ws.cell(row=row, column=9, value=event.get("total_value") or 0).border = thin_border
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=events_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"}
        )
        
    except Exception as e:
        logging.error(f"Error generating events Excel: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


# ============================================================
# PUSH NOTIFICATIONS & SECURITY
# ============================================================

class PushTokenCreate(BaseModel):
    token: str
    platform: str = "ios"

@api_router.post("/push-tokens")
async def register_push_token(
    token_data: PushTokenCreate,
    current_user: User = Depends(get_current_user)
):
    """Register push notification token for user"""
    try:
        existing = await db.push_tokens.find_one({
            "user_id": current_user.id,
            "token": token_data.token
        })
        
        if not existing:
            await db.push_tokens.insert_one({
                "id": str(uuid.uuid4()),
                "user_id": current_user.id,
                "token": token_data.token,
                "platform": token_data.platform,
                "created_at": datetime.now(timezone.utc),
                "is_active": True
            })
        
        return {"status": "registered"}
    except Exception as e:
        logging.error(f"Error registering push token: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


@api_router.get("/notifications/dguv-reminders")
async def get_dguv_reminders(
    days_ahead: int = 30,
    current_user: User = Depends(get_current_user)
):
    """Get upcoming DGUV test reminders"""
    try:
        cutoff_date = datetime.now(timezone.utc) + timedelta(days=days_ahead)
        
        # Find articles with upcoming DGUV tests
        articles = await db.articles.find({
            "next_dguv_test": {"$lte": cutoff_date, "$gte": datetime.now(timezone.utc)}
        }).to_list(100)
        
        reminders = []
        for article in articles:
            test_date = article.get("next_dguv_test")
            days_until = (test_date - datetime.now(timezone.utc)).days if test_date else 0
            
            reminders.append({
                "article_id": article.get("id"),
                "article_name": article.get("name"),
                "inventory_code": article.get("inventory_code"),
                "test_date": test_date.isoformat() if test_date else None,
                "days_until": days_until,
                "urgency": "high" if days_until <= 7 else "medium" if days_until <= 14 else "low"
            })
        
        # Sort by urgency
        reminders.sort(key=lambda x: x["days_until"])
        
        return reminders
    except Exception as e:
        logging.error(f"Error getting DGUV reminders: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


@api_router.get("/notifications/maintenance-reminders")
async def get_maintenance_reminders(
    days_ahead: int = 14,
    current_user: User = Depends(get_current_user)
):
    """Get upcoming maintenance reminders"""
    try:
        cutoff_date = datetime.now(timezone.utc) + timedelta(days=days_ahead)
        
        tasks = await db.maintenance_tasks.find({
            "scheduled_date": {"$lte": cutoff_date, "$gte": datetime.now(timezone.utc)},
            "status": {"$in": ["scheduled", "pending"]}
        }).to_list(100)
        
        reminders = []
        for task in tasks:
            article = await db.articles.find_one({"id": task.get("article_id")})
            scheduled_date = task.get("scheduled_date")
            days_until = (scheduled_date - datetime.now(timezone.utc)).days if scheduled_date else 0
            
            reminders.append({
                "task_id": task.get("id"),
                "article_id": task.get("article_id"),
                "article_name": article.get("name") if article else "Unbekannt",
                "task_type": task.get("task_type"),
                "scheduled_date": scheduled_date.isoformat() if scheduled_date else None,
                "days_until": days_until,
                "urgency": "high" if days_until <= 3 else "medium" if days_until <= 7 else "low"
            })
        
        reminders.sort(key=lambda x: x["days_until"])
        
        return reminders
    except Exception as e:
        logging.error(f"Error getting maintenance reminders: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


# Password Change
class PasswordChange(BaseModel):
    current_password: str
    new_password: str

@api_router.post("/users/change-password")
@limiter.limit("5/minute")
async def change_password(
    request: Request,
    password_data: PasswordChange,
    current_user: User = Depends(get_current_user)
):
    """Change user password"""
    try:
        # Verify current password
        user_doc = await db.users.find_one({"id": current_user.id})
        if not user_doc:
            raise HTTPException(status_code=404, detail="Benutzer nicht gefunden")
        
        if not pwd_context.verify(password_data.current_password, user_doc.get("hashed_password", "")):
            raise HTTPException(status_code=400, detail="Aktuelles Passwort ist falsch")
        
        # Validate new password
        if len(password_data.new_password) < 8:
            raise HTTPException(status_code=400, detail="Passwort muss mindestens 8 Zeichen lang sein")
        
        # Hash and update
        new_hash = pwd_context.hash(password_data.new_password)
        await db.users.update_one(
            {"id": current_user.id},
            {"$set": {
                "hashed_password": new_hash,
                "password_changed_at": datetime.now(timezone.utc)
            }}
        )
        
        return {"status": "success", "message": "Passwort erfolgreich geändert"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error changing password: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


# Session Management
@api_router.get("/users/sessions")
async def get_user_sessions(current_user: User = Depends(get_current_user)):
    """Get active sessions for current user"""
    try:
        sessions = await db.user_sessions.find({
            "user_id": current_user.id,
            "is_active": True
        }).to_list(20)
        
        result = []
        for session in sessions:
            session.pop("_id", None)
            result.append({
                "id": session.get("id"),
                "device": session.get("device", "Unbekannt"),
                "ip_address": session.get("ip_address"),
                "created_at": session.get("created_at"),
                "last_activity": session.get("last_activity"),
                "is_current": session.get("token") == current_user.id  # Simplified check
            })
        
        return result
    except Exception as e:
        logging.error(f"Error getting sessions: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


@api_router.delete("/users/sessions/{session_id}")
async def revoke_session(
    session_id: str,
    current_user: User = Depends(get_current_user)
):
    """Revoke a specific session"""
    try:
        result = await db.user_sessions.update_one(
            {"id": session_id, "user_id": current_user.id},
            {"$set": {"is_active": False, "revoked_at": datetime.now(timezone.utc)}}
        )
        
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Session nicht gefunden")
        
        return {"status": "success", "message": "Session beendet"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error revoking session: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


@api_router.post("/users/logout-all")
async def logout_all_sessions(current_user: User = Depends(get_current_user)):
    """Logout from all sessions"""
    try:
        await db.user_sessions.update_many(
            {"user_id": current_user.id, "is_active": True},
            {"$set": {"is_active": False, "revoked_at": datetime.now(timezone.utc)}}
        )
        
        return {"status": "success", "message": "Von allen Geräten abgemeldet"}
    except Exception as e:
        logging.error(f"Error logging out all sessions: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


# Admin: Reset Database Endpoint
class DatabaseResetRequest(BaseModel):
    confirmation_phrase: str  # H1: Must equal "DATENBANK LÖSCHEN" to confirm wipe

@api_router.post("/admin/reset-database")
async def reset_database(
    body: DatabaseResetRequest,
    current_user: User = Depends(require_permission(Permission.BACKUP_DATABASE))
):
    """Reset entire database - Admin only. Requires explicit confirmation phrase."""
    # H1: Two-step confirmation — prevents accidental total data loss
    if body.confirmation_phrase != "DATENBANK LÖSCHEN":
        raise HTTPException(
            status_code=400,
            detail="Bestätigungsphrase falsch. Sende {\"confirmation_phrase\": \"DATENBANK LÖSCHEN\"} um fortzufahren."
        )

    try:
        # List all collections to delete data from
        collections_to_clear = [
            'articles',
            'categories',
            'suppliers',
            'customers',
            'events',
            'bookings',
            'invoices',
            'storage_locations',
            'storage_zones',
            'inventory_movements',
            'maintenance_tasks',
            'bom',
            'teams',
            'audit_logs',
            'messages',
            'notifications'
        ]
        
        deleted_counts = {}
        
        # Delete all documents from each collection
        for collection_name in collections_to_clear:
            collection = db[collection_name]
            result = await collection.delete_many({})
            deleted_counts[collection_name] = result.deleted_count
        
        # Log the reset action in audit log
        audit_entry = {
            "id": str(uuid.uuid4()),
            "timestamp": datetime.now(timezone.utc).isoformat() + "Z",
            "user_id": current_user.id,
            "username": current_user.username,
            "action": "database_reset",
            "entity_type": "system",
            "entity_id": "all",
            "details": "Entire database reset - all data deleted"
        }
        await db.audit_logs.insert_one(audit_entry)
        
        return {
            "message": "Database reset successful",
            "deleted_counts": deleted_counts,
            "total_deleted": sum(deleted_counts.values())
        }
        
    except Exception as e:
        logging.error(f"Database reset error: {str(e)}")
        raise HTTPException(status_code=500, detail="Datenbankoperation fehlgeschlagen. Details wurden protokolliert.")

# Admin: Get Database Statistics
@api_router.get("/admin/database-stats")
async def get_database_stats(current_user: User = Depends(require_permission(Permission.ADMIN_ACCESS))):
    """Get database statistics - Admin only"""
    
    try:
        stats = {
            "articles": await db.articles.count_documents({}),
            "categories": await db.categories.count_documents({}),
            "suppliers": await db.suppliers.count_documents({}),
            "customers": await db.customers.count_documents({}),
            "events": await db.events.count_documents({}),
            "bookings": await db.bookings.count_documents({}),
            "invoices": await db.invoices.count_documents({}),
            "storage_locations": await db.storage_locations.count_documents({}),
            "storage_zones": await db.storage_zones.count_documents({}),
            "inventory_movements": await db.inventory_movements.count_documents({}),
            "maintenance_tasks": await db.maintenance_tasks.count_documents({}),
            "bom": await db.bom.count_documents({}),
            "teams": await db.teams.count_documents({}),
            "users": await db.users.count_documents({}),
            "audit_logs": await db.audit_logs.count_documents({}),
        }
        
        stats["total_documents"] = sum(stats.values())
        
        return stats
        
    except Exception as e:
        logging.error(f"Database stats error: {str(e)}")
        raise HTTPException(status_code=500, detail="Datenbankstatistiken konnten nicht abgerufen werden.")

# ============================================================
# PROJEKTPLANUNG - Artikel-Verfügbarkeit & Sub-Rental Tracking
# ============================================================

# Response Models for Availability
class ArticleAvailability(BaseModel):
    article_id: str
    article_name: str
    inventory_code: str
    total_stock: int
    available_stock: int
    booked_quantity: int
    bookings: List[dict] = []

class AvailabilityCalendarEntry(BaseModel):
    date: str
    articles_booked: int
    total_booked_quantity: int
    events: List[dict] = []

class SubRentalArticle(BaseModel):
    id: str
    name: str
    inventory_code: str
    supplier_name: str
    supplier_id: str
    sub_rental_cost: float
    current_stock: int
    status: str

class SubRentalCreate(BaseModel):
    article_id: str
    supplier_id: str
    cost: float
    quantity: int = 1
    rental_start: Optional[datetime] = None
    rental_end: Optional[datetime] = None
    notes: Optional[str] = None

class SubRental(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    article_id: str
    article_name: str
    supplier_id: str
    supplier_name: str
    cost: float
    quantity: int
    rental_start: Optional[datetime] = None
    rental_end: Optional[datetime] = None
    status: str = "active"  # active, returned, cancelled
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    created_by: str

# Artikel-Verfügbarkeit für einen Zeitraum prüfen
@api_router.get("/availability/articles")
async def get_articles_availability(
    start_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
    article_id: Optional[str] = Query(None, description="Filter by article ID"),
    current_user: User = Depends(get_current_user)
):
    """Get article availability for a date range"""
    try:
        # Parse dates
        if start_date:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        else:
            start_dt = datetime.now()
        
        if end_date:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        else:
            end_dt = start_dt + timedelta(days=30)
        
        # Build article query
        article_query = {}
        if article_id:
            article_query["id"] = article_id
        
        articles = await db.articles.find(article_query).to_list(1000)
        
        availability_list = []
        
        for article in articles:
            # Find active bookings that overlap with the date range
            booking_query = {
                "article_id": article["id"],
                "status": {"$ne": "cancelled"},
                "$or": [
                    {"pickup_date": {"$lte": end_dt}, "return_date": {"$gte": start_dt}},
                    {"pickup_date": {"$lte": end_dt, "$gte": start_dt}},
                    {"return_date": {"$lte": end_dt, "$gte": start_dt}}
                ]
            }
            
            bookings = await db.bookings.find(booking_query).to_list(100)
            
            # Calculate booked quantity
            booked_qty = sum(b.get("quantity", 0) for b in bookings if b.get("status") == "booked")
            
            # Get event info for each booking
            booking_details = []
            for booking in bookings:
                event = await db.events.find_one({"id": booking.get("event_id")})
                booking_details.append({
                    "booking_id": booking.get("id"),
                    "event_id": booking.get("event_id"),
                    "event_name": event.get("event_name") if event else "Unbekannt",
                    "quantity": booking.get("quantity"),
                    "pickup_date": booking.get("pickup_date").isoformat() if booking.get("pickup_date") else None,
                    "return_date": booking.get("return_date").isoformat() if booking.get("return_date") else None,
                    "status": booking.get("status")
                })
            
            availability_list.append({
                "article_id": article["id"],
                "article_name": article["name"],
                "inventory_code": article.get("inventory_code", ""),
                "total_stock": article.get("current_stock", 0),
                "available_stock": max(0, article.get("current_stock", 0) - booked_qty),
                "booked_quantity": booked_qty,
                "bookings": booking_details,
                "is_sub_rental": article.get("is_sub_rental", False),
                "weight_kg": article.get("weight_kg"),
                "power_watt": article.get("power_watt")
            })
        
        return {
            "start_date": start_date or start_dt.strftime("%Y-%m-%d"),
            "end_date": end_date or end_dt.strftime("%Y-%m-%d"),
            "total_articles": len(availability_list),
            "articles": availability_list
        }
        
    except Exception as e:
        logging.error(f"Error getting availability: {str(e)}")
        raise HTTPException(status_code=500, detail="Verfügbarkeit konnte nicht abgerufen werden.")

# Kalender-Ansicht der Verfügbarkeit
@api_router.get("/availability/calendar")
async def get_availability_calendar(
    month: Optional[int] = Query(None, description="Month 1-12"),
    year: Optional[int] = Query(None, description="Year YYYY"),
    current_user: User = Depends(get_current_user)
):
    """Get calendar view of bookings and availability for a month"""
    try:
        now = datetime.now()
        target_month = month or now.month
        target_year = year or now.year
        
        # Calculate start and end of month
        start_of_month = datetime(target_year, target_month, 1)
        if target_month == 12:
            end_of_month = datetime(target_year + 1, 1, 1) - timedelta(seconds=1)
        else:
            end_of_month = datetime(target_year, target_month + 1, 1) - timedelta(seconds=1)
        
        # Get all bookings for this month
        bookings = await db.bookings.find({
            "status": {"$ne": "cancelled"},
            "$or": [
                {"pickup_date": {"$lte": end_of_month}, "return_date": {"$gte": start_of_month}},
                {"pickup_date": {"$gte": start_of_month, "$lte": end_of_month}}
            ]
        }).to_list(500)
        
        # Get all events for this month
        events = await db.events.find({
            "$or": [
                {"start_date": {"$lte": end_of_month}, "end_date": {"$gte": start_of_month}},
                {"start_date": {"$gte": start_of_month, "$lte": end_of_month}}
            ]
        }).to_list(100)
        
        # Build calendar data
        calendar_data = {}
        
        # Process each day of the month
        current_day = start_of_month
        while current_day <= end_of_month:
            date_str = current_day.strftime("%Y-%m-%d")
            calendar_data[date_str] = {
                "date": date_str,
                "bookings_count": 0,
                "total_quantity": 0,
                "events": [],
                "articles": []
            }
            current_day += timedelta(days=1)
        
        # Add bookings to calendar
        for booking in bookings:
            pickup = booking.get("pickup_date")
            return_date = booking.get("return_date")
            
            if not pickup or not return_date:
                continue
            
            # Get article info
            article = await db.articles.find_one({"id": booking.get("article_id")})
            article_name = article.get("name") if article else "Unbekannt"
            
            # Mark all days of the booking
            current = max(pickup, start_of_month)
            end = min(return_date, end_of_month)
            
            while current <= end:
                date_str = current.strftime("%Y-%m-%d")
                if date_str in calendar_data:
                    calendar_data[date_str]["bookings_count"] += 1
                    calendar_data[date_str]["total_quantity"] += booking.get("quantity", 0)
                    calendar_data[date_str]["articles"].append({
                        "article_id": booking.get("article_id"),
                        "article_name": article_name,
                        "quantity": booking.get("quantity", 0)
                    })
                current += timedelta(days=1)
        
        # Add events to calendar
        for event in events:
            start = event.get("start_date")
            end = event.get("end_date")
            
            if not start or not end:
                continue
            
            current = max(start, start_of_month)
            end_event = min(end, end_of_month)
            
            while current <= end_event:
                date_str = current.strftime("%Y-%m-%d")
                if date_str in calendar_data:
                    calendar_data[date_str]["events"].append({
                        "event_id": event.get("id"),
                        "event_name": event.get("event_name"),
                        "event_number": event.get("event_number"),
                        "status": event.get("status")
                    })
                current += timedelta(days=1)
        
        return {
            "month": target_month,
            "year": target_year,
            "days": list(calendar_data.values())
        }
        
    except Exception as e:
        logging.error(f"Error getting availability calendar: {str(e)}")
        raise HTTPException(status_code=500, detail="Kalender konnte nicht abgerufen werden.")

# Sub-Rental Management
@api_router.get("/sub-rentals")
async def get_sub_rentals(
    status: Optional[str] = Query(None, description="Filter by status"),
    current_user: User = Depends(get_current_user)
):
    """Get all sub-rental articles and records"""
    try:
        # Get articles marked as sub-rental
        query = {"is_sub_rental": True}
        sub_rental_articles = await db.articles.find(query).to_list(100)
        
        # Get supplier info
        results = []
        for article in sub_rental_articles:
            supplier = None
            if article.get("sub_rental_supplier_id"):
                supplier = await db.suppliers.find_one({"id": article.get("sub_rental_supplier_id")})
            
            results.append({
                "id": article["id"],
                "name": article["name"],
                "inventory_code": article.get("inventory_code", ""),
                "supplier_id": article.get("sub_rental_supplier_id"),
                "supplier_name": supplier.get("name") if supplier else "Kein Lieferant",
                "sub_rental_cost": article.get("sub_rental_cost", 0),
                "current_stock": article.get("current_stock", 0),
                "status": article.get("status", "OK"),
                "weight_kg": article.get("weight_kg"),
                "power_watt": article.get("power_watt")
            })
        
        # Also get sub-rental records
        rental_query = {}
        if status:
            rental_query["status"] = status
        
        sub_rentals = await db.sub_rentals.find(rental_query).sort("created_at", -1).to_list(100)
        
        return {
            "sub_rental_articles": results,
            "sub_rental_records": sub_rentals,
            "total_articles": len(results),
            "total_records": len(sub_rentals)
        }
        
    except Exception as e:
        logging.error(f"Error getting sub-rentals: {str(e)}")
        raise HTTPException(status_code=500, detail="Fremdmietliste konnte nicht abgerufen werden.")

@api_router.post("/sub-rentals")
async def create_sub_rental(
    rental_data: SubRentalCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a new sub-rental record"""
    try:
        # Get article
        article = await db.articles.find_one({"id": rental_data.article_id})
        if not article:
            raise HTTPException(status_code=404, detail="Article not found")
        
        # Get supplier
        supplier = await db.suppliers.find_one({"id": rental_data.supplier_id})
        if not supplier:
            raise HTTPException(status_code=404, detail="Supplier not found")
        
        # Create sub-rental record
        sub_rental = {
            "id": str(uuid.uuid4()),
            "article_id": rental_data.article_id,
            "article_name": article["name"],
            "supplier_id": rental_data.supplier_id,
            "supplier_name": supplier["name"],
            "cost": rental_data.cost,
            "quantity": rental_data.quantity,
            "rental_start": rental_data.rental_start,
            "rental_end": rental_data.rental_end,
            "status": "active",
            "notes": rental_data.notes,
            "created_at": datetime.now(timezone.utc),
            "created_by": current_user.id
        }
        
        await db.sub_rentals.insert_one(sub_rental)
        
        # Update article atomically — $inc prevents race conditions with concurrent sub-rentals
        await db.articles.update_one(
            {"id": rental_data.article_id},
            {
                "$set": {
                    "is_sub_rental": True,
                    "sub_rental_supplier_id": rental_data.supplier_id,
                    "sub_rental_cost": rental_data.cost,
                    "updated_at": datetime.now(timezone.utc)
                },
                "$inc": {"current_stock": rental_data.quantity}
            }
        )
        
        return sub_rental
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating sub-rental: {str(e)}")
        raise HTTPException(status_code=500, detail="Fremdmiete konnte nicht erstellt werden.")

@api_router.put("/sub-rentals/{rental_id}/return")
async def return_sub_rental(
    rental_id: str,
    current_user: User = Depends(get_current_user)
):
    """Mark a sub-rental as returned"""
    try:
        rental = await db.sub_rentals.find_one({"id": rental_id})
        if not rental:
            raise HTTPException(status_code=404, detail="Sub-rental not found")
        
        # Update status
        await db.sub_rentals.update_one(
            {"id": rental_id},
            {"$set": {
                "status": "returned",
                "rental_end": datetime.now(timezone.utc)
            }}
        )
        
        # Reduce article stock atomically — prevents race condition on concurrent returns
        qty = rental.get("quantity", 0)
        if qty > 0:
            await db.articles.update_one(
                {"id": rental.get("article_id"), "current_stock": {"$gte": qty}},
                {"$inc": {"current_stock": -qty}, "$set": {"updated_at": datetime.now(timezone.utc)}}
            )
            # If guard fails (stock already 0), still proceed — sub-rental may have been partial
        
        return {"message": "Sub-rental returned successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error returning sub-rental: {str(e)}")
        raise HTTPException(status_code=500, detail="Fremdmiete-Rückgabe fehlgeschlagen.")

# Event Gewichts- und Strombedarf berechnen
@api_router.get("/events/{event_id}/requirements")
async def get_event_requirements(
    event_id: str,
    current_user: User = Depends(get_current_user)
):
    """Calculate total weight and power requirements for an event based on bookings"""
    try:
        # Get event
        event = await db.events.find_one({"id": event_id})
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        
        # Get all bookings for this event
        bookings = await db.bookings.find({
            "event_id": event_id,
            "status": {"$ne": "cancelled"}
        }).to_list(100)
        
        total_weight = 0.0
        total_power_230v = 0
        total_power_400v = 0
        total_rental_cost = 0.0
        articles_list = []
        
        for booking in bookings:
            article = await db.articles.find_one({"id": booking.get("article_id")})
            if not article:
                continue
            
            quantity = booking.get("quantity", 1)
            weight = (article.get("weight_kg") or 0) * quantity
            power = (article.get("power_watt") or 0) * quantity
            rental = (article.get("rental_price") or 0) * quantity
            
            total_weight += weight
            total_rental_cost += rental
            
            power_type = article.get("power_type", "230V")
            if power_type == "400V":
                total_power_400v += power
            else:
                total_power_230v += power
            
            articles_list.append({
                "article_id": article["id"],
                "article_name": article["name"],
                "quantity": quantity,
                "weight_kg": article.get("weight_kg"),
                "power_watt": article.get("power_watt"),
                "power_type": power_type,
                "rental_price": article.get("rental_price"),
                "is_sub_rental": article.get("is_sub_rental", False),
                "sub_rental_cost": article.get("sub_rental_cost") if article.get("is_sub_rental") else None
            })
        
        # Calculate recommendations
        recommendations = {
            "vehicle_type": "PKW" if total_weight < 500 else ("Sprinter" if total_weight < 1500 else "LKW"),
            "schuko_16a_needed": int((total_power_230v / 3500) + 1) if total_power_230v > 0 else 0,
            "cee_32a_needed": int((total_power_400v / 22000) + 1) if total_power_400v > 0 else 0
        }
        
        # Calculate event duration
        start = event.get("start_date")
        end = event.get("end_date")
        days = (end - start).days + 1 if start and end else 1
        
        return {
            "event_id": event_id,
            "event_name": event.get("event_name"),
            "event_dates": {
                "start": start.isoformat() if start else None,
                "end": end.isoformat() if end else None,
                "duration_days": days
            },
            "weight": {
                "total_kg": round(total_weight, 2),
                "recommendations": recommendations["vehicle_type"]
            },
            "power": {
                "total_230v_watt": total_power_230v,
                "total_400v_watt": total_power_400v,
                "total_watt": total_power_230v + total_power_400v,
                "ampere_230v": round(total_power_230v / 230, 1) if total_power_230v > 0 else 0,
                "ampere_400v": round(total_power_400v / 400, 1) if total_power_400v > 0 else 0,
                "schuko_16a_needed": recommendations["schuko_16a_needed"],
                "cee_32a_needed": recommendations["cee_32a_needed"]
            },
            "rental": {
                "daily_cost": round(total_rental_cost, 2),
                "total_cost": round(total_rental_cost * days, 2),
                "duration_days": days
            },
            "articles": articles_list,
            "total_bookings": len(bookings)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error calculating event requirements: {str(e)}")
        raise HTTPException(status_code=500, detail="Bedarfsberechnung fehlgeschlagen.")

# ============================================================
# FINANZ-LOGIK - Mietpreis-Kalkulator & PDF-Generierung
# ============================================================

# F6 — Shared helper: tiered rental price for a single article over N days
def _calculate_rental_price(article: dict, days: int) -> float:
    """Return the total rental price for `days` days using tiered pricing."""
    days = max(1, days)
    if days <= 3:
        unit = article.get("rental_price_day") or article.get("rental_price") or 0.0
        return float(unit) * days
    elif days <= 7:
        unit = article.get("rental_price_week") or (float(article.get("rental_price") or 0.0) * 5)
        return float(unit)
    else:
        unit = article.get("rental_price_month") or (float(article.get("rental_price") or 0.0) * 20)
        return float(unit)


# F6 — Return all pricing tiers for a single article
@api_router.get("/articles/{article_id}/rental-tiers")
async def get_rental_tiers(article_id: str, current_user: User = Depends(get_current_user)):
    article = await db.articles.find_one({"id": article_id, "deleted": {"$ne": True}})
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
    base = float(article.get("rental_price") or 0.0)
    return {
        "article_id": article["id"],
        "article_name": article["name"],
        "tiers": {
            "day": {
                "label": "Tagespreis (1–3 Tage)",
                "price_per_period": float(article.get("rental_price_day") or base),
                "example_3_days": _calculate_rental_price(article, 3),
            },
            "week": {
                "label": "Wochenpreis (4–7 Tage)",
                "price_per_period": float(article.get("rental_price_week") or base * 5),
                "example_7_days": _calculate_rental_price(article, 7),
            },
            "month": {
                "label": "Monatspreis (>7 Tage)",
                "price_per_period": float(article.get("rental_price_month") or base * 20),
                "example_30_days": _calculate_rental_price(article, 30),
            },
        },
    }


# Mietpreis-Kalkulator - Berechnung basierend auf Dauer
@api_router.post("/rental-calculator")
async def rental_calculator_endpoint(
    article_ids: List[str] = [],
    quantities: List[int] = [],
    start_date: str = None,
    end_date: str = None,
    current_user: User = Depends(get_current_user)
):
    """Calculate rental prices based on duration with tiered pricing"""
    try:
        if not start_date or not end_date:
            raise HTTPException(status_code=400, detail="Start and end date required")
        
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        duration_days = (end_dt - start_dt).days + 1
        
        if duration_days < 1:
            raise HTTPException(status_code=400, detail="Invalid date range")
        
        # Determine pricing tier
        if duration_days <= 3:
            pricing_tier = "day"
            tier_name = "Tagespreis"
        elif duration_days <= 7:
            pricing_tier = "week"
            tier_name = "Wochenpreis"
        else:
            pricing_tier = "month"
            tier_name = "Monatspreis"
        
        items = []
        subtotal = 0.0
        
        for i, article_id in enumerate(article_ids):
            article = await db.articles.find_one({"id": article_id})
            if not article:
                continue
            
            qty = quantities[i] if i < len(quantities) else 1
            
            # Get appropriate price based on tier
            if pricing_tier == "day":
                unit_price = article.get("rental_price_day") or article.get("rental_price") or 0
            elif pricing_tier == "week":
                unit_price = article.get("rental_price_week") or (article.get("rental_price") or 0) * 5
            else:  # month
                unit_price = article.get("rental_price_month") or (article.get("rental_price") or 0) * 20
            
            # Apply weekend factor if applicable
            weekend_factor = 1.0
            if start_dt.weekday() >= 4:  # Friday or later
                weekend_factor = article.get("rental_factor_weekend", 1.0) or 1.0
            
            # Calculate item total
            if pricing_tier == "day":
                item_total = unit_price * qty * duration_days * weekend_factor
            else:
                # For week/month pricing, price is for the period, not per day
                item_total = unit_price * qty * weekend_factor
            
            items.append({
                "article_id": article["id"],
                "article_name": article["name"],
                "inventory_code": article.get("inventory_code"),
                "quantity": qty,
                "unit_price": unit_price,
                "weekend_factor": weekend_factor,
                "item_total": round(item_total, 2),
                "prices": {
                    "day": article.get("rental_price_day") or article.get("rental_price"),
                    "week": article.get("rental_price_week"),
                    "month": article.get("rental_price_month")
                }
            })
            
            subtotal += item_total
        
        tax_rate = 19.0
        tax_amount = subtotal * (tax_rate / 100)
        total = subtotal + tax_amount
        
        return {
            "duration_days": duration_days,
            "pricing_tier": pricing_tier,
            "tier_name": tier_name,
            "start_date": start_date,
            "end_date": end_date,
            "items": items,
            "subtotal": round(subtotal, 2),
            "tax_rate": tax_rate,
            "tax_amount": round(tax_amount, 2),
            "total": round(total, 2),
            "discounts": {
                "week_discount": "~30% Ersparnis gegenüber Tagespreis" if pricing_tier == "week" else None,
                "month_discount": "~50% Ersparnis gegenüber Tagespreis" if pricing_tier == "month" else None
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error calculating rental: {str(e)}")
        raise HTTPException(status_code=500, detail="Mietberechnung fehlgeschlagen.")

# Mietvertrag Model
class RentalContract(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contract_number: str
    event_id: str
    customer_id: str
    items: List[dict]
    start_date: datetime
    end_date: datetime
    subtotal: float
    tax_rate: float = 19.0
    tax_amount: float
    total_amount: float
    deposit_amount: float = 0.0
    status: str = "draft"  # draft, signed, active, completed, cancelled
    terms_accepted: bool = False
    signature_customer: Optional[str] = None  # Base64 signature
    signature_date: Optional[datetime] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    created_by: str

async def generate_contract_number():
    """Generate unique contract number: MV-YYYY-NNNN"""
    year = datetime.now().year
    count = await db.rental_contracts.count_documents({
        "contract_number": {"$regex": f"^MV-{year}"}
    })
    return f"MV-{year}-{str(count + 1).zfill(4)}"

# Mietvertrag erstellen
@api_router.post("/rental-contracts")
async def create_rental_contract(
    event_id: str,
    deposit_amount: float = 0.0,
    notes: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Create a rental contract for an event"""
    try:
        # Get event
        event = await db.events.find_one({"id": event_id})
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        
        # Get bookings
        bookings = await db.bookings.find({
            "event_id": event_id,
            "status": {"$ne": "cancelled"}
        }).to_list(100)
        
        if not bookings:
            raise HTTPException(status_code=400, detail="No bookings for this event")
        
        # Calculate items and totals
        items = []
        subtotal = 0.0
        
        start_date = event.get("start_date", datetime.now(timezone.utc))
        end_date = event.get("end_date", datetime.now(timezone.utc))
        duration_days = (end_date - start_date).days + 1 if start_date and end_date else 1
        
        for booking in bookings:
            article = await db.articles.find_one({"id": booking.get("article_id")})
            if not article:
                continue
            
            qty = booking.get("quantity", 1)
            
            # Determine price based on duration
            if duration_days <= 3:
                unit_price = article.get("rental_price_day") or article.get("rental_price") or 0
                price_label = "Tagespreis"
            elif duration_days <= 7:
                unit_price = article.get("rental_price_week") or (article.get("rental_price") or 0) * 5
                price_label = "Wochenpreis"
            else:
                unit_price = article.get("rental_price_month") or (article.get("rental_price") or 0) * 20
                price_label = "Monatspreis"
            
            item_total = unit_price * qty
            
            items.append({
                "article_id": article["id"],
                "article_name": article["name"],
                "inventory_code": article.get("inventory_code"),
                "serial_number": article.get("serial_number"),
                "quantity": qty,
                "unit_price": unit_price,
                "price_label": price_label,
                "item_total": round(item_total, 2)
            })
            
            subtotal += item_total
        
        tax_rate = 19.0
        tax_amount = subtotal * (tax_rate / 100)
        total = subtotal + tax_amount
        
        contract_number = await generate_contract_number()
        
        contract = RentalContract(
            contract_number=contract_number,
            event_id=event_id,
            customer_id=event.get("customer_id"),
            items=items,
            start_date=start_date,
            end_date=end_date,
            subtotal=round(subtotal, 2),
            tax_rate=tax_rate,
            tax_amount=round(tax_amount, 2),
            total_amount=round(total, 2),
            deposit_amount=deposit_amount,
            notes=notes,
            created_by=current_user.id
        )
        
        await db.rental_contracts.insert_one(contract.dict())
        
        return contract
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating contract: {str(e)}")
        raise HTTPException(status_code=500, detail="Vertrag konnte nicht erstellt werden.")

# Mietverträge abrufen
@api_router.get("/rental-contracts")
async def get_rental_contracts(
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all rental contracts"""
    try:
        query = {}
        if status:
            query["status"] = status
        
        contracts = await db.rental_contracts.find(query).sort("created_at", -1).to_list(100)
        return contracts
        
    except Exception as e:
        logging.error(f"Error getting contracts: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")

# Einzelnen Mietvertrag abrufen
@api_router.get("/rental-contracts/{contract_id}")
async def get_rental_contract(
    contract_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get a single rental contract"""
    contract = await db.rental_contracts.find_one({"id": contract_id})
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    return contract

# Mietvertrag Status aktualisieren
@api_router.put("/rental-contracts/{contract_id}/status")
async def update_contract_status(
    contract_id: str,
    status: str,
    current_user: User = Depends(get_current_user)
):
    """Update contract status"""
    contract = await db.rental_contracts.find_one({"id": contract_id})
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    valid_statuses = ["draft", "signed", "active", "completed", "cancelled"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")
    
    await db.rental_contracts.update_one(
        {"id": contract_id},
        {"$set": {"status": status}}
    )
    
    return {"message": f"Contract status updated to {status}"}

# Mietvertrag unterschreiben (digitale Unterschrift)
@api_router.put("/rental-contracts/{contract_id}/sign")
async def sign_rental_contract(
    contract_id: str,
    body: dict,
    current_user: User = Depends(get_current_user)
):
    """Save customer signature and mark contract as signed"""
    contract = await db.rental_contracts.find_one({"id": contract_id})
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    await db.rental_contracts.update_one(
        {"id": contract_id},
        {"$set": {
            "signature_customer": body.get("signature_customer", ""),
            "signature_date": datetime.now(timezone.utc),
            "signed_by": body.get("signed_by", ""),
            "status": "signed",
        }}
    )
    return {"message": "Unterschrift gespeichert"}

# ── Briefpapier Hilfsfunktionen ────────────────────────────────────────────────

def get_letterhead_header_html(settings: dict) -> str:
    """Build the document header block using company/letterhead settings."""
    logo_html = ""
    if settings.get("letterhead_show_logo", True) and settings.get("letterhead_logo_url"):
        logo_html = f'<img src="{settings["letterhead_logo_url"]}" style="max-height:60px;max-width:200px;object-fit:contain;margin-bottom:6px;">'
    slogan = settings.get("letterhead_slogan", "")
    slogan_html = f'<div class="slogan">{slogan}</div>' if slogan else ""
    addr_parts = [
        settings.get("company_address", ""),
        settings.get("company_phone", ""),
        settings.get("company_email", ""),
    ]
    company_info = " · ".join(p for p in addr_parts if p)
    return f"""
    <div class="lh-header">
      {logo_html}
      <div class="lh-company">{settings.get("company_name", "")}</div>
      {slogan_html}
      {f'<div class="lh-info">{company_info}</div>' if company_info else ""}
    </div>"""

def get_letterhead_footer_html(settings: dict) -> str:
    """Build the document footer block if configured."""
    text = settings.get("letterhead_footer_text", "")
    if not text:
        return ""
    return f'<div class="lh-footer">{text}</div>'

def get_letterhead_styles(settings: dict) -> str:
    """Return CSS variables/overrides for the letterhead primary color."""
    color = settings.get("letterhead_primary_color", "#FF9500")
    return f"""
    .lh-header {{ border-bottom: 3px solid {color}; padding-bottom: 16px; margin-bottom: 28px; }}
    .lh-company {{ font-size: 22px; font-weight: bold; color: {color}; }}
    .slogan {{ font-size: 13px; color: #888; margin-top: 2px; }}
    .lh-info {{ font-size: 12px; color: #666; margin-top: 4px; }}
    .lh-footer {{ margin-top: 40px; padding-top: 12px; border-top: 1px solid #ddd; font-size: 11px; color: #888; text-align: center; }}
    .primary-color {{ color: {color}; }}
    .primary-bg {{ background: {color}; }}
    .primary-border {{ border-color: {color}; }}"""

# PDF Template für Mietvertrag HTML
def generate_rental_contract_html(contract: dict, event: dict, customer: dict, settings: dict = None) -> str:
    """Generate HTML for rental contract PDF"""
    items_html = ""
    for item in contract.get("items", []):
        items_html += f"""
        <tr>
            <td>{item.get('article_name', '')}</td>
            <td>{item.get('inventory_code', '')}</td>
            <td>{item.get('quantity', 1)}</td>
            <td>{item.get('price_label', 'Tagespreis')}</td>
            <td>€{item.get('unit_price', 0):.2f}</td>
            <td>€{item.get('item_total', 0):.2f}</td>
        </tr>
        """
    
    start_date = contract.get("start_date")
    end_date = contract.get("end_date")
    if isinstance(start_date, str):
        start_date = parse_iso_date(start_date, "start_date")
    if isinstance(end_date, str):
        end_date = parse_iso_date(end_date, "end_date")
    
    s = settings or {}
    color = s.get("letterhead_primary_color", "#FF9500")
    lh_header = get_letterhead_header_html(s)
    lh_footer = get_letterhead_footer_html(s)
    lh_styles = get_letterhead_styles(s)

    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{ font-family: 'Helvetica', sans-serif; padding: 40px; color: #333; }}
            .contract-number {{ font-size: 18px; color: #666; margin-top: 6px; }}
            .section {{ margin: 25px 0; padding: 15px; background: #f8f9fa; border-radius: 8px; }}
            .section-title {{ color: {color}; font-weight: bold; margin-bottom: 10px; }}
            table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            th {{ background: {color}; color: white; padding: 12px; text-align: left; }}
            td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
            .totals {{ text-align: right; margin-top: 20px; }}
            .total-row {{ padding: 8px 0; }}
            .total-final {{ font-size: 20px; font-weight: bold; color: {color}; border-top: 2px solid {color}; padding-top: 10px; }}
            .terms {{ font-size: 11px; margin-top: 30px; padding: 15px; border: 1px solid #ddd; }}
            .signature {{ margin-top: 50px; display: flex; justify-content: space-between; }}
            .signature-box {{ width: 45%; padding-top: 50px; border-top: 1px solid #333; text-align: center; }}
            {lh_styles}
        </style>
    </head>
    <body>
        {lh_header}
        <div class="contract-number">MIETVERTRAG · Vertrag-Nr: {contract.get('contract_number', '')} · Datum: {datetime.now().strftime('%d.%m.%Y')}</div>

        <div class="section">
            <div class="section-title">Vermieter</div>
            <div><strong>{s.get("company_name", "Ihr Unternehmen")}</strong></div>
            <div>{s.get("company_address", "")}</div>
            <div>Tel: {s.get("company_phone", "")} | E-Mail: {s.get("company_email", "")}</div>
        </div>
        
        <div class="section">
            <div class="section-title">Mieter</div>
            <div><strong>{customer.get('company_name', customer.get('name', 'N/A'))}</strong></div>
            <div>{customer.get('contact_person', '')}</div>
            <div>{customer.get('address', '')}</div>
            <div>Tel: {customer.get('phone', '')} | E-Mail: {customer.get('email', '')}</div>
        </div>
        
        <div class="section">
            <div class="section-title">Mietgegenstand für Event: {event.get('event_name', '')}</div>
            <div><strong>Mietzeitraum:</strong> {start_date.strftime('%d.%m.%Y') if start_date else ''} bis {end_date.strftime('%d.%m.%Y') if end_date else ''}</div>
            <div><strong>Veranstaltungsort:</strong> {event.get('location', '')}</div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Artikel</th>
                    <th>Art.-Nr.</th>
                    <th>Menge</th>
                    <th>Preisart</th>
                    <th>Einzelpreis</th>
                    <th>Gesamt</th>
                </tr>
            </thead>
            <tbody>
                {items_html}
            </tbody>
        </table>
        
        <div class="totals">
            <div class="total-row">Zwischensumme: €{contract.get('subtotal', 0):.2f}</div>
            <div class="total-row">MwSt ({contract.get('tax_rate', 19)}%): €{contract.get('tax_amount', 0):.2f}</div>
            <div class="total-row total-final">Gesamtbetrag: €{contract.get('total_amount', 0):.2f}</div>
            {f'<div class="total-row">Kaution: €{contract.get("deposit_amount", 0):.2f}</div>' if contract.get('deposit_amount', 0) > 0 else ''}
        </div>
        
        <div class="terms">
            <strong>Mietbedingungen:</strong><br>
            1. Der Mieter verpflichtet sich, die Mietgegenstände pfleglich zu behandeln.<br>
            2. Bei Beschädigung oder Verlust haftet der Mieter.<br>
            3. Die Rückgabe erfolgt im ursprünglichen Zustand.<br>
            4. Zahlungsziel: 14 Tage nach Rechnungserhalt.<br>
            {f'5. Hinweis: {contract.get("notes", "")}' if contract.get('notes') else ''}
        </div>
        
        <div class="signature">
            <div class="signature-box">
                Vermieter<br>
                Ort, Datum
            </div>
            <div class="signature-box">
                Mieter<br>
                Ort, Datum
            </div>
        </div>
        {lh_footer}
    </body>
    </html>
    """

# PDF-Daten für Mietvertrag generieren
@api_router.get("/rental-contracts/{contract_id}/pdf-data")
async def get_contract_pdf_data(
    contract_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get all data needed to generate contract PDF on frontend"""
    try:
        contract = await db.rental_contracts.find_one({"id": contract_id})
        if not contract:
            raise HTTPException(status_code=404, detail="Contract not found")
        
        event = await db.events.find_one({"id": contract.get("event_id")})
        customer = await db.customers.find_one({"id": contract.get("customer_id")})
        
        # Load settings for letterhead
        app_settings = await db.app_settings.find_one({"_id": "main"}) or {}

        # Generate HTML
        html = generate_rental_contract_html(
            contract,
            event or {},
            customer or {},
            app_settings
        )
        
        return {
            "contract": contract,
            "event": event,
            "customer": customer,
            "html": html
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error generating PDF data: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")

# Verbesserte Rechnungsdaten mit detaillierten Positionen
@api_router.get("/invoices/{invoice_id}/pdf-data")
async def get_invoice_pdf_data(
    invoice_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get all data needed to generate invoice PDF with detailed line items"""
    try:
        invoice = await db.invoices.find_one({"id": invoice_id})
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        event = await db.events.find_one({"id": invoice.get("event_id")})
        customer = await db.customers.find_one({"id": invoice.get("customer_id")})
        
        # Get booking details for line items
        bookings = await db.bookings.find({
            "event_id": invoice.get("event_id"),
            "status": {"$ne": "cancelled"}
        }).to_list(100)
        
        items = []
        for booking in bookings:
            article = await db.articles.find_one({"id": booking.get("article_id")})
            if article:
                items.append({
                    "article_name": article.get("name"),
                    "inventory_code": article.get("inventory_code"),
                    "quantity": booking.get("quantity", 1),
                    "unit_price": booking.get("unit_price", article.get("rental_price", 0)),
                    "total": booking.get("total_price", 0)
                })
        
        return {
            "invoice": invoice,
            "event": event,
            "customer": customer,
            "items": items
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error getting invoice PDF data: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


# ============================================================
# CREW & FUHRPARK - Mitarbeiter und Fahrzeuge verwalten
# ============================================================

class CrewMember(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    role: str
    phone: Optional[str] = None
    email: Optional[str] = None
    hourly_rate: float = 0
    skills: List[str] = []
    is_available: bool = True
    created_at: datetime = Field(default_factory=datetime.utcnow)

class Vehicle(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    type: str
    license_plate: str
    capacity_kg: float = 0
    loading_meters: float = 0
    is_available: bool = True
    created_at: datetime = Field(default_factory=datetime.utcnow)

class CrewAssignment(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    event_id: str
    crew_ids: List[str] = []
    vehicle_id: Optional[str] = None
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

# Crew Endpoints
@api_router.get("/crew")
async def get_crew(current_user: User = Depends(get_current_user)):
    """Get all crew members"""
    crew = await db.crew.find().sort("name", 1).to_list(500)
    for member in crew:
        member.pop("_id", None)
    return crew

@api_router.post("/crew")
async def create_crew_member(
    member_data: CrewMember,
    current_user: User = Depends(get_current_user)
):
    """Create a new crew member"""
    try:
        member = CrewMember(**member_data.dict())
        await db.crew.insert_one(member.dict())
        return member
    except Exception as e:
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")

@api_router.put("/crew/{member_id}")
async def update_crew_member(
    member_id: str,
    update_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Update a crew member"""
    result = await db.crew.update_one(
        {"id": member_id},
        {"$set": update_data}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Mitarbeiter nicht gefunden")
    return {"status": "success"}

@api_router.delete("/crew/{member_id}")
async def delete_crew_member(
    member_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete a crew member"""
    result = await db.crew.delete_one({"id": member_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Mitarbeiter nicht gefunden")
    return {"status": "deleted"}

# Vehicle Endpoints
@api_router.get("/vehicles")
async def get_vehicles(current_user: User = Depends(get_current_user)):
    """Get all vehicles"""
    vehicles = await db.vehicles.find().sort("name", 1).to_list(100)
    for vehicle in vehicles:
        vehicle.pop("_id", None)
    return vehicles

@api_router.post("/vehicles")
async def create_vehicle(
    vehicle_data: VehicleCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a new vehicle"""
    try:
        vehicle = Vehicle(**vehicle_data.dict())
        await db.vehicles.insert_one(vehicle.dict())
        return vehicle
    except Exception as e:
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")

@api_router.put("/vehicles/{vehicle_id}")
async def update_vehicle(
    vehicle_id: str,
    update_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Update a vehicle"""
    result = await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": update_data}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Fahrzeug nicht gefunden")
    return {"status": "success"}

@api_router.delete("/vehicles/{vehicle_id}")
async def delete_vehicle(
    vehicle_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete a vehicle"""
    result = await db.vehicles.delete_one({"id": vehicle_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Fahrzeug nicht gefunden")
    return {"status": "deleted"}

# Crew Assignment Endpoints
@api_router.get("/crew-assignments")
async def get_crew_assignments(
    event_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get crew assignments"""
    query = {}
    if event_id:
        query["event_id"] = event_id
    
    assignments = await db.crew_assignments.find(query).to_list(500)
    
    result = []
    for assignment in assignments:
        assignment.pop("_id", None)
        event = await db.events.find_one({"id": assignment.get("event_id")})
        assignment["event_name"] = event.get("event_name") if event else "Unbekannt"
        result.append(assignment)
    
    return result

@api_router.post("/crew-assignments")
async def create_crew_assignment(
    assignment_data: CrewAssignment,
    current_user: User = Depends(get_current_user)
):
    """Create a new crew assignment"""
    try:
        assignment = CrewAssignment(**assignment_data.dict())
        await db.crew_assignments.insert_one(assignment.dict())
        return assignment
    except Exception as e:
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")

@api_router.put("/events/{event_id}/assign-crew")
async def assign_crew_to_event(
    event_id: str,
    crew_ids: List[str],
    vehicle_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Assign crew and vehicle to an event"""
    try:
        # Check if assignment exists
        existing = await db.crew_assignments.find_one({"event_id": event_id})
        
        if existing:
            await db.crew_assignments.update_one(
                {"event_id": event_id},
                {"$set": {"crew_ids": crew_ids, "vehicle_id": vehicle_id}}
            )
        else:
            assignment = CrewAssignment(
                event_id=event_id,
                crew_ids=crew_ids,
                vehicle_id=vehicle_id
            )
            await db.crew_assignments.insert_one(assignment.dict())
        
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")


# ============================================================
# SETS & BUNDLES - Artikel-Pakete verwalten
# ============================================================

async def calculate_bundle_totals(items: List[dict]) -> dict:
    """Calculate total weight, power, and prices for bundle items"""
    total_weight = 0.0
    total_power = 0
    total_price_day = 0.0
    total_price_week = 0.0
    total_price_month = 0.0
    total_items = 0
    
    for item in items:
        article = await db.articles.find_one({"id": item.get("article_id")})
        if article:
            qty = item.get("quantity", 1)
            total_items += qty
            total_weight += (article.get("weight_kg") or 0) * qty
            total_power += (article.get("power_watt") or 0) * qty
            
            day_price = article.get("rental_price_day") or article.get("rental_price") or 0
            week_price = article.get("rental_price_week") or (day_price * 5)
            month_price = article.get("rental_price_month") or (day_price * 20)
            
            total_price_day += day_price * qty
            total_price_week += week_price * qty
            total_price_month += month_price * qty
    
    return {
        "total_items": total_items,
        "total_weight_kg": round(total_weight, 2),
        "total_power_watt": total_power,
        "rental_price_day": round(total_price_day, 2),
        "rental_price_week": round(total_price_week, 2),
        "rental_price_month": round(total_price_month, 2)
    }

async def generate_bundle_code(category: str) -> str:
    """Generate unique bundle code"""
    prefix_map = {
        "PA": "PA",
        "Licht": "LT",
        "Video": "VD",
        "Rigging": "RG",
        "Strom": "PWR",
        "Standard": "SET",
        "Sonstige": "MIX"
    }
    prefix = prefix_map.get(category, "SET")
    count = await db.bundles.count_documents({"category": category})
    return f"{prefix}-{str(count + 1).zfill(3)}"

@api_router.get("/bundles")
async def get_bundles(
    category: Optional[str] = None,
    active_only: bool = True,
    current_user: User = Depends(get_current_user)
):
    """Get all bundles/sets"""
    query = {}
    if category:
        query["category"] = category
    if active_only:
        query["is_active"] = True
    
    bundles = await db.bundles.find(query).sort("name", 1).to_list(100)
    
    # Enrich with article details
    result = []
    for bundle in bundles:
        # Remove MongoDB _id
        bundle.pop("_id", None)
        
        items_with_details = []
        for item in bundle.get("items", []):
            article = await db.articles.find_one({"id": item.get("article_id")})
            item_data = {
                "article_id": item.get("article_id"),
                "quantity": item.get("quantity", 1),
                "is_optional": item.get("is_optional", False)
            }
            if article:
                item_data["article_name"] = article.get("name")
                item_data["inventory_code"] = article.get("inventory_code")
                item_data["available_stock"] = article.get("current_stock", 0)
            items_with_details.append(item_data)
        
        bundle["items"] = items_with_details
        result.append(bundle)
    
    return result

@api_router.get("/bundles/{bundle_id}")
async def get_bundle(
    bundle_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get a single bundle with full details"""
    bundle = await db.bundles.find_one({"id": bundle_id})
    if not bundle:
        raise HTTPException(status_code=404, detail="Bundle nicht gefunden")
    
    # Enrich with article details
    items_with_details = []
    for item in bundle.get("items", []):
        article = await db.articles.find_one({"id": item.get("article_id")})
        if article:
            items_with_details.append({
                **item,
                "article_name": article.get("name"),
                "inventory_code": article.get("inventory_code"),
                "serial_number": article.get("serial_number"),
                "available_stock": article.get("current_stock", 0),
                "weight_kg": article.get("weight_kg"),
                "power_watt": article.get("power_watt"),
                "rental_price_day": article.get("rental_price_day") or article.get("rental_price"),
                "image": article.get("image_base64")
            })
    
    return {
        **bundle,
        "items": items_with_details
    }

@api_router.post("/bundles")
async def create_bundle(
    bundle_data: BundleCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a new bundle/set"""
    try:
        # Generate bundle code if not provided
        bundle_code = bundle_data.bundle_code
        if not bundle_code:
            bundle_code = await generate_bundle_code(bundle_data.category)
        
        # Check if code already exists
        existing = await db.bundles.find_one({"bundle_code": bundle_code})
        if existing:
            raise HTTPException(status_code=400, detail="Bundle-Code existiert bereits")
        
        # Calculate totals
        items_dict = [item.dict() for item in bundle_data.items]
        totals = await calculate_bundle_totals(items_dict)
        
        # Create bundle data dict and override bundle_code
        bundle_dict = bundle_data.dict()
        bundle_dict['bundle_code'] = bundle_code  # Use generated/validated code
        
        bundle = Bundle(
            **bundle_dict,
            **totals,
            created_by=current_user.id
        )
        
        await db.bundles.insert_one(bundle.dict())
        
        return bundle
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating bundle: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")

@api_router.put("/bundles/{bundle_id}")
async def update_bundle(
    bundle_id: str,
    bundle_data: BundleUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update an existing bundle"""
    bundle = await db.bundles.find_one({"id": bundle_id})
    if not bundle:
        raise HTTPException(status_code=404, detail="Bundle nicht gefunden")
    
    update_data = {k: v for k, v in bundle_data.dict().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    # Recalculate totals if items changed
    if "items" in update_data:
        items_dict = [item.dict() if hasattr(item, 'dict') else item for item in update_data["items"]]
        totals = await calculate_bundle_totals(items_dict)
        update_data.update(totals)
    
    await db.bundles.update_one({"id": bundle_id}, {"$set": update_data})
    
    return {"message": "Bundle aktualisiert", "id": bundle_id}

@api_router.delete("/bundles/{bundle_id}")
async def delete_bundle(
    bundle_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete a bundle (soft delete - sets inactive)"""
    bundle = await db.bundles.find_one({"id": bundle_id})
    if not bundle:
        raise HTTPException(status_code=404, detail="Bundle nicht gefunden")
    
    # Soft delete - just deactivate
    await db.bundles.update_one(
        {"id": bundle_id},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"message": "Bundle deaktiviert"}

@api_router.post("/bundles/{bundle_id}/book")
async def book_bundle(
    bundle_id: str,
    event_id: str,
    exclude_items: List[str] = [],  # article_ids to exclude (for optional items)
    current_user: User = Depends(get_current_user)
):
    """Book all items of a bundle for an event"""
    try:
        bundle = await db.bundles.find_one({"id": bundle_id})
        if not bundle:
            raise HTTPException(status_code=404, detail="Bundle nicht gefunden")
        
        event = await db.events.find_one({"id": event_id})
        if not event:
            raise HTTPException(status_code=404, detail="Event nicht gefunden")
        
        bookings_created = []
        errors = []
        
        for item in bundle.get("items", []):
            article_id = item.get("article_id")
            
            # Skip excluded items (optional items the user doesn't want)
            if article_id in exclude_items:
                continue
            
            article = await db.articles.find_one({"id": article_id})
            if not article:
                errors.append(f"Artikel {article_id} nicht gefunden")
                continue
            
            quantity = item.get("quantity", 1)
            
            # Check availability
            available = article.get("current_stock", 0)
            if available < quantity:
                errors.append(f"{article.get('name')}: Nur {available} verfügbar, {quantity} benötigt")
                continue
            
            # Get price (with bundle discount)
            base_price = article.get("rental_price_day") or article.get("rental_price") or 0
            discount = bundle.get("bundle_discount_percent", 0) / 100
            unit_price = base_price * (1 - discount)
            
            # Create booking
            booking = Booking(
                event_id=event_id,
                article_id=article_id,
                quantity=quantity,
                pickup_date=event.get("start_date"),
                return_date=event.get("end_date"),
                unit_price=round(unit_price, 2),
                total_price=round(unit_price * quantity, 2),
                status="booked",
                notes=f"Aus Bundle: {bundle.get('name')} ({bundle.get('bundle_code')})",
                created_by=current_user.id
            )
            
            await db.bookings.insert_one(booking.dict())
            bookings_created.append({
                "booking_id": booking.id,
                "article_name": article.get("name"),
                "quantity": quantity,
                "price": round(unit_price * quantity, 2)
            })
        
        # Calculate total
        total_price = sum(b["price"] for b in bookings_created)
        
        return {
            "message": f"{len(bookings_created)} Buchungen erstellt",
            "bundle_name": bundle.get("name"),
            "event_name": event.get("event_name"),
            "bookings": bookings_created,
            "total_price": round(total_price, 2),
            "discount_applied": f"{bundle.get('bundle_discount_percent', 0)}%",
            "errors": errors if errors else None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error booking bundle: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler. Details wurden protokolliert.")

@api_router.get("/bundles/{bundle_id}/availability")
async def check_bundle_availability(
    bundle_id: str,
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user)
):
    """Check if all items in a bundle are available for a date range"""
    bundle = await db.bundles.find_one({"id": bundle_id})
    if not bundle:
        raise HTTPException(status_code=404, detail="Bundle nicht gefunden")
    
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    
    availability = []
    all_available = True
    
    for item in bundle.get("items", []):
        article = await db.articles.find_one({"id": item.get("article_id")})
        if not article:
            continue
        
        required_qty = item.get("quantity", 1)
        total_stock = article.get("current_stock", 0)
        
        # Check overlapping bookings
        bookings = await db.bookings.find({
            "article_id": item.get("article_id"),
            "status": {"$ne": "cancelled"},
            "$or": [
                {"pickup_date": {"$lte": end_dt}, "return_date": {"$gte": start_dt}},
            ]
        }).to_list(100)
        
        booked_qty = sum(b.get("quantity", 0) for b in bookings)
        available_qty = total_stock - booked_qty
        
        is_available = available_qty >= required_qty
        if not is_available:
            all_available = False
        
        availability.append({
            "article_id": article["id"],
            "article_name": article["name"],
            "required": required_qty,
            "available": max(0, available_qty),
            "is_available": is_available,
            "is_optional": item.get("is_optional", False)
        })
    
    return {
        "bundle_id": bundle_id,
        "bundle_name": bundle.get("name"),
        "date_range": {"start": start_date, "end": end_date},
        "all_available": all_available,
        "items": availability
    }

# Include router AFTER all endpoints are defined
app.add_api_websocket_route("/ws", websocket_endpoint)
app.include_router(api_router)

# v1 router aliases are registered at the END of this file (after all @app routes)

_allowed_origins_env = os.environ.get("ALLOWED_ORIGINS", "http://localhost:8002,http://localhost:8081")
ALLOWED_ORIGINS = [o.strip() for o in _allowed_origins_env.split(",") if o.strip()]
if "*" in ALLOWED_ORIGINS:
    logging.critical("⛔ ALLOWED_ORIGINS enthält '*' – das ist ein Sicherheitsproblem! Bitte .env anpassen.")
    ALLOWED_ORIGINS = ["http://localhost:8002"]

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=ALLOWED_ORIGINS,
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "Accept"],
)

# M2 — Security response headers (clickjacking, MIME sniffing, referrer leakage)
class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        response.headers.setdefault("X-Content-Type-Options", "nosniff")
        response.headers.setdefault("X-Frame-Options", "DENY")
        response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
        response.headers.setdefault("X-XSS-Protection", "1; mode=block")
        # K3: Content-Security-Policy — blocks inline scripts, limits resource origins
        response.headers.setdefault(
            "Content-Security-Policy",
            "default-src 'self'; "
            "script-src 'self'; "
            "style-src 'self' 'unsafe-inline'; "
            "img-src 'self' data:; "
            "font-src 'self'; "
            "connect-src 'self'; "
            "frame-ancestors 'none';"
        )
        return response

app.add_middleware(SecurityHeadersMiddleware)


class RequestLoggingMiddleware(BaseHTTPMiddleware):
    """V5: Log every request with method, path, status, duration, and client IP."""
    async def dispatch(self, request, call_next):
        import time
        start = time.perf_counter()
        # Resolve real IP behind reverse proxies
        forwarded_for = request.headers.get("X-Forwarded-For")
        client_ip = forwarded_for.split(",")[0].strip() if forwarded_for else (
            request.client.host if request.client else "unknown"
        )
        response = await call_next(request)
        duration_ms = (time.perf_counter() - start) * 1000
        logging.info(
            "%s %s %s %.1fms %s",
            request.method,
            request.url.path,
            response.status_code,
            duration_ms,
            client_ip,
        )
        return response

app.add_middleware(RequestLoggingMiddleware)


# V3: Global exception handler — converts unhandled DB / runtime errors to HTTP 503
# so clients always get a structured JSON response instead of a 500 traceback.
@app.exception_handler(Exception)
async def global_exception_handler(request, exc):
    # Re-raise HTTPException so FastAPI handles it normally (404, 401, etc.)
    if isinstance(exc, HTTPException):
        raise exc
    logging.error("Unhandled exception on %s %s: %s", request.method, request.url.path, exc, exc_info=True)
    return JSONResponse(
        status_code=503,
        content={"detail": "Service temporarily unavailable. Please try again later."}
    )


# Backup & Restore Functions
async def create_database_backup():
    """Create backup of all database collections"""
    try:
        backup_data = {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "database": DB_NAME,
            "collections": {}
        }
        
        # Get all collection names
        collection_names = await db.list_collection_names()
        
        # Backup each collection
        for collection_name in collection_names:
            collection = db[collection_name]
            documents = await collection.find().to_list(None)
            
            # Convert ObjectId and datetime to string
            serializable_docs = []
            for doc in documents:
                if '_id' in doc:
                    doc['_id'] = str(doc['_id'])
                # Convert datetime fields to ISO format
                for key, value in doc.items():
                    if isinstance(value, datetime):
                        doc[key] = value.isoformat()
                serializable_docs.append(doc)
            
            backup_data["collections"][collection_name] = serializable_docs
            logging.info(f"Backed up {len(serializable_docs)} documents from {collection_name}")
        
        # Write to latest_backup.json (canonical file for restore)
        with open(BACKUP_FILE, 'w') as f:
            json.dump(backup_data, f, indent=2)

        # M3 — Keep a timestamped copy and retain only the 7 most recent backups
        ts = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
        timestamped = BACKUP_DIR / f"backup_{ts}.json"
        shutil.copy2(BACKUP_FILE, timestamped)
        all_backups = sorted(BACKUP_DIR.glob("backup_*.json"))
        for old_backup in all_backups[:-7]:
            try:
                old_backup.unlink()
            except Exception as cleanup_err:
                logging.warning(f"Could not delete old backup {old_backup}: {cleanup_err}")

        logging.info(f"Database backup completed. Saved to {BACKUP_FILE} + {timestamped.name}")
        return True
        
    except Exception as e:
        logging.error(f"Backup failed: {str(e)}")
        return False

async def restore_database_from_backup():
    """Restore database from backup file"""
    try:
        if not BACKUP_FILE.exists():
            logging.error("No backup file found")
            return False, "No backup file found"
        
        # Read backup file
        with open(BACKUP_FILE, 'r') as f:
            backup_data = json.load(f)
        
        restored_collections = []
        
        # Restore each collection
        for collection_name, documents in backup_data["collections"].items():
            if not documents:
                continue
                
            collection = db[collection_name]
            
            # Clear existing data
            await collection.delete_many({})
            
            # Convert datetime strings back to datetime objects
            for doc in documents:
                for key, value in doc.items():
                    if isinstance(value, str) and 'T' in value:
                        try:
                            doc[key] = datetime.fromisoformat(value.replace('Z', '+00:00'))
                        except (ValueError, AttributeError):
                            pass
            
            # Insert backup data
            if documents:
                await collection.insert_many(documents)
                restored_collections.append(f"{collection_name}: {len(documents)} docs")
                logging.info(f"Restored {len(documents)} documents to {collection_name}")
        
        logging.info("Database restore completed successfully")
        return True, f"Restored {len(restored_collections)} collections"
        
    except Exception as e:
        logging.error(f"Restore failed: {str(e)}")
        return False, str(e)

# Scheduled backup job
def scheduled_backup_job():
    """Job to run daily backup"""
    loop = asyncio.new_event_loop()
    try:
        asyncio.set_event_loop(loop)
        result = loop.run_until_complete(create_database_backup())
        if result:
            logging.info("Scheduled backup completed successfully")
        else:
            logging.error("Scheduled backup FAILED — create_database_backup returned False")
    except Exception as e:
        logging.error(f"Scheduled backup job error: {e}", exc_info=True)
    finally:
        loop.close()

# ===========================================
# RATE LIMITER ERROR HANDLER
# ===========================================
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# ===========================================
# GLOBALER EXCEPTION HANDLER
# Verhindert, dass interne Fehlermeldungen an den Client gelangen
# ===========================================
from fastapi.responses import JSONResponse

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    logging.error(f"Unbehandelter Fehler bei {request.method} {request.url.path}: {exc}", exc_info=True)
    return JSONResponse(
        status_code=500,
        content={"detail": "Interner Serverfehler. Details wurden protokolliert."}
    )

# ===========================
# APP SETTINGS
# ===========================

@app.get("/api/settings/app")
async def get_app_settings():
    doc = await db.app_settings.find_one({"_id": "main"})
    if not doc:
        defaults = AppSettings()
        await db.app_settings.insert_one({"_id": "main", **defaults.dict()})
        return defaults.dict()
    doc.pop("_id", None)
    return doc

@app.put("/api/settings/app")
async def update_app_settings(settings: AppSettings, current_user: dict = Depends(get_current_user)):
    # Get current settings to compare changes
    old_doc = await db.app_settings.find_one({"_id": "main"}) or {}
    old_settings = AppSettings(**{k: v for k, v in old_doc.items() if k != "_id"}) if old_doc else AppSettings()

    # Find what changed
    changes = {}
    for field, new_value in settings.dict().items():
        if field in ['settings_version', 'version_history']:
            continue
        old_value = getattr(old_settings, field, None)
        if old_value != new_value:
            changes[field] = {"old": old_value, "new": new_value}

    # Increment version and save history if there are changes
    if changes:
        new_version = old_settings.settings_version + 1
        settings.settings_version = new_version

        # Create version history entry
        version_entry = {
            "version": new_version,
            "changed_at": datetime.utcnow(),
            "changed_by": current_user.get("username", "unknown"),
            "changes": changes
        }

        # Save to version history collection
        await db.settings_versions.insert_one(version_entry)

    await db.app_settings.update_one(
        {"_id": "main"},
        {"$set": settings.dict()},
        upsert=True
    )
    return settings.dict()

@app.get("/api/settings/versions")
async def get_settings_versions(current_user: dict = Depends(get_current_user)):
    """Get settings version history"""
    versions = await db.settings_versions.find().sort("version", -1).to_list(50)
    for v in versions:
        v.pop("_id", None)
    return versions

@app.post("/api/settings/restore/{version}")
async def restore_settings_version(version: int, current_user: dict = Depends(get_current_user)):
    """Restore settings to a specific version"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Einstellungen wiederherstellen")

    version_entry = await db.settings_versions.find_one({"version": version})
    if not version_entry:
        raise HTTPException(status_code=404, detail="Version nicht gefunden")

    # Get current settings
    current_doc = await db.app_settings.find_one({"_id": "main"}) or {}
    current_settings = AppSettings(**{k: v for k, v in current_doc.items() if k != "_id"}) if current_doc else AppSettings()

    # Restore the old values
    restored_settings = current_settings.dict()
    for field, values in version_entry["changes"].items():
        restored_settings[field] = values["old"]

    # Increment version
    restored_settings["settings_version"] = current_settings.settings_version + 1

    # Save
    await db.app_settings.update_one(
        {"_id": "main"},
        {"$set": restored_settings},
        upsert=True
    )

    return {"message": f"Einstellungen auf Version {version} zurückgesetzt", "version": restored_settings["settings_version"]}

# ── Online-Angebote: öffentlicher Endpunkt (KEIN Auth) ────────────────────────
@app.get("/api/quotes/public/{token}")
@limiter.limit("20/minute")  # H4: Prevent token brute-force
async def get_public_quote(request: Request, token: str):
    """Public quote view — no authentication required"""
    app_settings_doc = await db.app_settings.find_one({"_id": "main"}) or {}
    if not app_settings_doc.get("online_quotes_enabled", True):
        raise HTTPException(status_code=403, detail="Online-Angebote sind deaktiviert")

    quote = await db.quotes.find_one({"public_token": token})
    if not quote:
        raise HTTPException(status_code=404, detail="Angebot nicht gefunden oder Link ungültig")

    # Check expiry
    expiry_days = int(app_settings_doc.get("online_quotes_expiry_days", 30))
    created_at = quote.get("public_token_created_at")
    if created_at:
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
        age_days = (datetime.now(timezone.utc) - created_at.replace(tzinfo=None)).days
        if age_days > expiry_days:
            raise HTTPException(status_code=410, detail="Dieser Link ist abgelaufen")

    return {
        "quote_number": quote.get("quote_number"),
        "customer_name": quote.get("customer_name"),
        "event_name": quote.get("event_name"),
        "event_date": quote.get("event_date"),
        "valid_until": quote.get("valid_until"),
        "items": quote.get("items", []),
        "discount_percent": quote.get("discount_percent", 0),
        "total_net": quote.get("total_net", 0),
        "status": quote.get("status"),
        "notes": quote.get("notes"),
        "company_name": app_settings_doc.get("company_name", ""),
        "company_email": app_settings_doc.get("company_email", ""),
        "company_phone": app_settings_doc.get("company_phone", ""),
        "letterhead_primary_color": app_settings_doc.get("letterhead_primary_color", "#FF9500"),
        "letterhead_logo_url": app_settings_doc.get("letterhead_logo_url", ""),
        "letterhead_slogan": app_settings_doc.get("letterhead_slogan", ""),
    }

# ── Einladungen: öffentlicher RSVP-Endpunkt (KEIN Auth) ───────────────────────
@app.put("/api/invitations/public/{token}/respond")
async def respond_to_invitation(token: str, body: dict, request: Request):
    # M2: Basic CSRF protection — verify Origin header is from a known origin
    origin = request.headers.get("Origin") or request.headers.get("Referer", "")
    if origin and not any(origin.startswith(o) for o in ALLOWED_ORIGINS):
        raise HTTPException(status_code=403, detail="Ungültige Anfragequelle")
    status = body.get("status")  # accepted | declined
    if status not in ("accepted", "declined"):
        raise HTTPException(status_code=400, detail="Status muss 'accepted' oder 'declined' sein")
    result = await db.event_invitations.update_one(
        {"token": token},
        {"$set": {"status": status, "responded_at": datetime.now(timezone.utc)}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Einladung nicht gefunden")
    inv = await db.event_invitations.find_one({"token": token})
    return {"status": inv.get("status"), "name": inv.get("name")}

# ── Online-Rechnungen: öffentlicher Endpunkt (KEIN Auth) ──────────────────────
@app.get("/api/invoices/public/{token}")
@limiter.limit("20/minute")  # H4: Prevent token brute-force
async def get_public_invoice(request: Request, token: str):
    """Public invoice view — no authentication required"""
    app_settings_doc = await db.app_settings.find_one({"_id": "main"}) or {}
    if not app_settings_doc.get("online_invoices_enabled", False):
        raise HTTPException(status_code=403, detail="Online-Rechnungen sind deaktiviert")

    invoice = await db.invoices.find_one({"public_token": token})
    if not invoice:
        raise HTTPException(status_code=404, detail="Rechnung nicht gefunden oder Link ungültig")

    # H2: Expiry check — same pattern as quotes
    expiry_days = int(app_settings_doc.get("online_invoices_expiry_days", 30))
    token_created = invoice.get("public_token_created_at")
    if token_created:
        if isinstance(token_created, str):
            token_created = datetime.fromisoformat(token_created.replace("Z", "+00:00"))
        age_days = (datetime.now(timezone.utc) - token_created.replace(tzinfo=None)).days
        if age_days > expiry_days:
            raise HTTPException(status_code=410, detail="Dieser Rechnungslink ist abgelaufen")

    customer = await db.customers.find_one({"id": invoice.get("customer_id")}) or {}
    event = await db.events.find_one({"id": invoice.get("event_id")}) or {}

    return {
        "invoice_number": invoice.get("invoice_number"),
        "status": invoice.get("status"),
        "payment_status": invoice.get("payment_status"),
        "issue_date": invoice.get("issue_date"),
        "due_date": invoice.get("due_date"),
        "items": invoice.get("items", []),
        "subtotal": invoice.get("subtotal", 0),
        "tax_rate": invoice.get("tax_rate", 19),
        "tax_amount": invoice.get("tax_amount", 0),
        "total_amount": invoice.get("total_amount", 0),
        "notes": invoice.get("notes"),
        "customer_name": customer.get("name", invoice.get("customer_id", "")),
        "event_name": event.get("title", event.get("name", "")),
        "company_name": app_settings_doc.get("company_name", ""),
        "company_email": app_settings_doc.get("company_email", ""),
        "company_phone": app_settings_doc.get("company_phone", ""),
        "letterhead_primary_color": app_settings_doc.get("letterhead_primary_color", "#FF9500"),
        "letterhead_logo_url": app_settings_doc.get("letterhead_logo_url", ""),
        "payment_text": app_settings_doc.get("online_invoices_payment_text", ""),
    }

# ===========================
# FAHRZEUGE
# ===========================

@app.get("/api/vehicles")
async def get_vehicles(current_user: dict = Depends(get_current_user)):
    vehicles = await db.vehicles.find({}).to_list(500)
    for v in vehicles:
        v["id"] = str(v.get("_id", v.get("id", "")))
        v.pop("_id", None)
    return vehicles

@app.post("/api/vehicles")
async def create_vehicle(vehicle: Vehicle, current_user: dict = Depends(get_current_user)):
    doc = vehicle.dict()
    doc["_id"] = doc["id"]
    await db.vehicles.insert_one(doc)
    return doc

@app.put("/api/vehicles/{vehicle_id}")
async def update_vehicle(vehicle_id: str, vehicle: VehicleCreate, current_user: dict = Depends(get_current_user)):
    await db.vehicles.update_one({"_id": vehicle_id}, {"$set": vehicle.dict()})
    return {"id": vehicle_id, **vehicle.dict()}

@app.delete("/api/vehicles/{vehicle_id}")
async def delete_vehicle(vehicle_id: str, current_user: dict = Depends(get_current_user)):
    await db.vehicles.delete_one({"_id": vehicle_id})
    return {"message": "Fahrzeug gelöscht"}

# ===========================
# AUFGABEN
# ===========================

@app.get("/api/tasks")
async def get_tasks(status: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if status:
        query["status"] = status
    tasks = await db.tasks.find(query).sort("created_at", -1).to_list(500)
    for t in tasks:
        t["id"] = str(t.get("_id", t.get("id", "")))
        t.pop("_id", None)
    return tasks

@app.post("/api/tasks")
async def create_task(task: TaskCreate, current_user: User = Depends(get_current_user)):
    import uuid as _uuid, traceback as _tb
    from datetime import datetime as _dt
    try:
        task_id = str(_uuid.uuid4())
        doc = {
            "_id": task_id,
            "id": task_id,
            "created_at": _dt.utcnow().isoformat(),
            "created_by": getattr(current_user, "username", str(current_user)),
            **task.dict(),
        }
        await db.tasks.insert_one(doc)
        doc.pop("_id", None)
        return doc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"ERR_TASK: {_tb.format_exc()}")

@app.put("/api/tasks/{task_id}")
async def update_task(task_id: str, task: TaskCreate, current_user: dict = Depends(get_current_user)):
    await db.tasks.update_one({"_id": task_id}, {"$set": task.dict()})
    return {"id": task_id, **task.dict()}

@app.delete("/api/tasks/{task_id}")
async def delete_task(task_id: str, current_user: dict = Depends(get_current_user)):
    await db.tasks.delete_one({"_id": task_id})
    return {"message": "Aufgabe gelöscht"}

# ===========================
# SERIENNUMMERN
# ===========================

@app.get("/api/serial-numbers")
async def get_serial_numbers(article_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if article_id:
        query["article_id"] = article_id
    items = await db.serial_numbers.find(query).to_list(1000)
    for item in items:
        item["id"] = str(item.get("_id", item.get("id", "")))
        item.pop("_id", None)
    return items

@app.post("/api/serial-numbers")
async def create_serial_number(sn: SerialNumber, current_user: dict = Depends(get_current_user)):
    doc = sn.dict()
    doc["_id"] = doc["id"]
    await db.serial_numbers.insert_one(doc)
    return doc

@app.put("/api/serial-numbers/{sn_id}")
async def update_serial_number(sn_id: str, sn: SerialNumberCreate, current_user: dict = Depends(get_current_user)):
    await db.serial_numbers.update_one({"_id": sn_id}, {"$set": sn.dict()})
    return {"id": sn_id, **sn.dict()}

@app.delete("/api/serial-numbers/{sn_id}")
async def delete_serial_number(sn_id: str, current_user: dict = Depends(get_current_user)):
    await db.serial_numbers.delete_one({"_id": sn_id})
    return {"message": "Seriennummer gelöscht"}

# ===========================
# ABWESENHEITSANTRÄGE
# ===========================

@app.get("/api/absence-requests")
async def get_absence_requests(current_user: dict = Depends(get_current_user)):
    items = await db.absence_requests.find({}).sort("created_at", -1).to_list(500)
    for item in items:
        item["id"] = str(item.get("_id", item.get("id", "")))
        item.pop("_id", None)
    return items

@app.post("/api/absence-requests")
async def create_absence_request(req: AbsenceRequest, current_user: dict = Depends(get_current_user)):
    doc = req.dict()
    doc["_id"] = doc["id"]
    await db.absence_requests.insert_one(doc)
    return doc

@app.put("/api/absence-requests/{req_id}/status")
async def update_absence_status(req_id: str, body: dict, current_user: dict = Depends(get_current_user)):
    await db.absence_requests.update_one(
        {"_id": req_id},
        {"$set": {"status": body.get("status", "ausstehend")}}
    )
    return {"message": "Status aktualisiert"}

@app.put("/api/absence-requests/{req_id}/sign")
async def sign_absence_request(req_id: str, body: dict, current_user: dict = Depends(get_current_user)):
    """Save manager approval signature"""
    new_status = body.get("status", "genehmigt")
    await db.absence_requests.update_one(
        {"_id": req_id},
        {"$set": {
            "approval_signature": body.get("signature", ""),
            "approved_by": body.get("signed_by", ""),
            "approved_at": datetime.now(timezone.utc),
            "status": new_status,
        }}
    )
    return {"message": "Antrag bearbeitet"}

@app.put("/api/absence-requests/{req_id}")
async def update_absence_request(req_id: str, req: AbsenceRequestCreate, current_user: dict = Depends(get_current_user)):
    await db.absence_requests.update_one({"_id": req_id}, {"$set": req.dict()})
    return {"id": req_id, **req.dict()}

@app.delete("/api/absence-requests/{req_id}")
async def delete_absence_request(req_id: str, current_user: dict = Depends(get_current_user)):
    await db.absence_requests.delete_one({"_id": req_id})
    return {"message": "Antrag gelöscht"}

# ===========================
# BESTANDSZÄHLUNGEN
# ===========================

@app.get("/api/stock-counts")
async def get_stock_counts(current_user: dict = Depends(get_current_user)):
    items = await db.stock_counts.find({}).sort("created_at", -1).to_list(200)
    for item in items:
        item["id"] = str(item.get("_id", item.get("id", "")))
        item.pop("_id", None)
    return items

@app.post("/api/stock-counts/start")
async def start_stock_count(current_user: dict = Depends(get_current_user)):
    articles = await db.articles.find({}).to_list(10000)
    items = [
        {"article_id": str(a.get("_id", a.get("id",""))),
         "article_name": a.get("name",""),
         "expected_quantity": a.get("quantity", 0),
         "counted_quantity": 0}
        for a in articles
    ]
    from datetime import datetime as dt
    import uuid as _uuid
    sc_id = str(_uuid.uuid4())
    doc = {
        "_id": sc_id,
        "id": sc_id,
        "name": f"Bestandszählung {dt.now().strftime('%d.%m.%Y')}",
        "status": "offen",
        "items": items,
        "created_at": dt.utcnow(),
        "created_by": current_user.username,
        "notes": ""
    }
    await db.stock_counts.insert_one(doc)
    doc.pop("_id", None)
    return doc

@app.post("/api/stock-counts")
async def create_stock_count(sc: StockCount, current_user: User = Depends(get_current_user)):
    doc = sc.dict()
    doc["_id"] = doc["id"]
    doc["created_by"] = current_user.username
    await db.stock_counts.insert_one(doc)
    return doc

@app.get("/api/stock-counts/{sc_id}")
async def get_stock_count(sc_id: str, current_user: dict = Depends(get_current_user)):
    doc = await db.stock_counts.find_one({"_id": sc_id})
    if not doc:
        raise HTTPException(404, "Nicht gefunden")
    doc["id"] = str(doc.get("_id", doc.get("id","")))
    doc.pop("_id", None)
    return doc

@app.put("/api/stock-counts/{sc_id}")
async def update_stock_count(sc_id: str, sc: StockCountCreate, current_user: dict = Depends(get_current_user)):
    await db.stock_counts.update_one({"_id": sc_id}, {"$set": sc.dict()})
    return {"id": sc_id, **sc.dict()}

@app.delete("/api/stock-counts/{sc_id}")
async def delete_stock_count(sc_id: str, current_user: dict = Depends(get_current_user)):
    await db.stock_counts.delete_one({"_id": sc_id})
    return {"message": "Zählung gelöscht"}

# ===========================
# PRÜFUNGEN
# ===========================

@app.get("/api/inspections")
async def get_inspections(current_user: dict = Depends(get_current_user)):
    items = await db.inspections.find({}).sort("due_date", 1).to_list(1000)
    for item in items:
        item["id"] = str(item.get("_id", item.get("id","")))
        item.pop("_id", None)
    return items

@app.post("/api/inspections")
async def create_inspection(insp: Inspection, current_user: dict = Depends(get_current_user)):
    doc = insp.dict()
    doc["_id"] = doc["id"]
    await db.inspections.insert_one(doc)
    return doc

@app.put("/api/inspections/{insp_id}")
async def update_inspection(insp_id: str, insp: InspectionCreate, current_user: dict = Depends(get_current_user)):
    await db.inspections.update_one({"_id": insp_id}, {"$set": insp.dict()})
    return {"id": insp_id, **insp.dict()}

@app.delete("/api/inspections/{insp_id}")
async def delete_inspection(insp_id: str, current_user: dict = Depends(get_current_user)):
    await db.inspections.delete_one({"_id": insp_id})
    return {"message": "Prüfung gelöscht"}

# ===========================
# DASHBOARD FINANCIAL STATS
# ===========================

@app.get("/api/dashboard/financial")
async def get_financial_stats(current_user: dict = Depends(get_current_user)):
    invoices = await db.invoices.find({}).to_list(10000)
    outstanding_total = sum(inv.get("total_amount", 0) for inv in invoices
                            if inv.get("payment_status", "offen") != "bezahlt")
    outstanding_offen = sum(inv.get("total_amount", 0) for inv in invoices
                            if inv.get("payment_status", "offen") == "offen")
    outstanding_ueberfaellig = sum(inv.get("total_amount", 0) for inv in invoices
                                   if inv.get("payment_status") == "überfällig")
    outstanding_bezahlt = sum(inv.get("total_amount", 0) for inv in invoices
                              if inv.get("payment_status") == "bezahlt")
    events = await db.events.find({}).to_list(5000)
    revenue_confirmed = sum(e.get("total_value", 0) for e in events
                            if e.get("status") in ["confirmed", "aktiv", "bestätigt"])
    revenue_option = sum(e.get("total_value", 0) for e in events
                         if e.get("status") in ["option", "angebot"])
    revenue_cancelled = sum(e.get("total_value", 0) for e in events
                            if e.get("status") in ["cancelled", "storniert", "abgesagt"])
    quotes = await db.quotes.find({"status": {"$in": ["entwurf", "gesendet"]}}).to_list(5000)
    open_quotes_count = len(quotes)
    open_quotes_total = sum(q.get("total_net", 0) for q in quotes)
    return {
        "outstanding_total": round(outstanding_total, 2),
        "outstanding_offen": round(outstanding_offen, 2),
        "outstanding_ueberfaellig": round(outstanding_ueberfaellig, 2),
        "outstanding_bezahlt": round(outstanding_bezahlt, 2),
        "revenue_confirmed": round(revenue_confirmed, 2),
        "revenue_option": round(revenue_option, 2),
        "revenue_cancelled": round(revenue_cancelled, 2),
        "open_quotes_count": open_quotes_count,
        "open_quotes_total": round(open_quotes_total, 2),
    }

# ===========================
# BESTELLUNGEN (PURCHASE ORDERS)
# ===========================

class PurchaseOrderItem(BaseModel):
    article_id: str = ""
    article_name: str = ""
    quantity: int = 1
    unit_price: float = 0.0
    total_price: float = 0.0

class PurchaseOrderCreate(BaseModel):
    supplier_id: Optional[str] = None
    supplier_name: str = ""
    order_date: Optional[str] = None
    expected_delivery: Optional[str] = None
    status: str = "offen"  # offen | bestellt | teilweise_geliefert | geliefert | storniert
    items: List[PurchaseOrderItem] = []
    notes: Optional[str] = None
    total_amount: float = 0.0

class PurchaseOrder(PurchaseOrderCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    order_number: str = ""
    created_at: datetime = Field(default_factory=datetime.utcnow)
    created_by: str = ""

@app.get("/api/purchase-orders")
async def get_purchase_orders(status: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if status:
        query["status"] = status
    items = await db.purchase_orders.find(query).sort("created_at", -1).to_list(500)
    for item in items:
        item["id"] = str(item.get("_id", item.get("id", "")))
        item.pop("_id", None)
    return items

@app.post("/api/purchase-orders")
async def create_purchase_order(po: PurchaseOrderCreate, current_user: User = Depends(get_current_user)):
    import uuid as _uuid
    from datetime import datetime as _dt
    po_id = str(_uuid.uuid4())
    year = _dt.utcnow().year
    count = await db.purchase_orders.count_documents({"order_number": {"$regex": f"^PO-{year}-"}})
    doc = {
        "_id": po_id,
        "id": po_id,
        "created_at": _dt.utcnow().isoformat(),
        "created_by": current_user.username,
        "order_number": f"PO-{year}-{count + 1:04d}",
        **po.dict(),
    }
    await db.purchase_orders.insert_one(doc)
    doc.pop("_id", None)
    return doc

@app.get("/api/purchase-orders/{po_id}")
async def get_purchase_order(po_id: str, current_user: User = Depends(get_current_user)):
    doc = await db.purchase_orders.find_one({"_id": po_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Bestellung nicht gefunden")
    doc["id"] = str(doc.get("_id", doc.get("id", "")))
    doc.pop("_id", None)
    return doc

@app.put("/api/purchase-orders/{po_id}")
async def update_purchase_order(po_id: str, po: PurchaseOrderCreate, current_user: User = Depends(get_current_user)):
    await db.purchase_orders.update_one({"_id": po_id}, {"$set": po.dict()})
    return {"id": po_id, **po.dict()}

@app.delete("/api/purchase-orders/{po_id}")
async def delete_purchase_order(po_id: str, current_user: User = Depends(get_current_user)):
    await db.purchase_orders.delete_one({"_id": po_id})
    return {"message": "Bestellung gelöscht"}

# ===========================
# AKTIVITÄTEN (TIME TRACKING)
# ===========================

class ActivityCreate(BaseModel):
    title: str
    description: Optional[str] = None
    crew_member_id: Optional[str] = None
    crew_member_name: str = ""
    event_id: Optional[str] = None
    event_name: str = ""
    date: str  # YYYY-MM-DD
    duration_hours: float = 0.0
    activity_type: str = "allgemein"  # allgemein | auf_und_abbau | transport | probe | veranstaltung
    status: str = "geplant"  # geplant | in_bearbeitung | abgeschlossen

class Activity(ActivityCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)

@app.get("/api/activities")
async def get_activities(
    crew_member_id: Optional[str] = None,
    event_id: Optional[str] = None,
    date: Optional[str] = None,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    current_user: User = Depends(get_current_user)
):
    query = {}
    if crew_member_id:
        query["crew_member_id"] = crew_member_id
    if event_id:
        query["event_id"] = event_id
    if date:
        query["date"] = date
    skip = (page - 1) * page_size
    items = await db.activities.find(query).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    for item in items:
        item["id"] = str(item.get("_id", item.get("id", "")))
        item.pop("_id", None)
    return items

@app.post("/api/activities")
async def create_activity(activity: Activity, current_user: User = Depends(get_current_user)):
    doc = activity.dict()
    doc["_id"] = doc["id"]
    await db.activities.insert_one(doc)
    doc.pop("_id", None)
    return doc

@app.put("/api/activities/{activity_id}")
async def update_activity(activity_id: str, activity: ActivityCreate, current_user: User = Depends(get_current_user)):
    await db.activities.update_one({"_id": activity_id}, {"$set": activity.dict()})
    return {"id": activity_id, **activity.dict()}

@app.delete("/api/activities/{activity_id}")
async def delete_activity(activity_id: str, current_user: User = Depends(get_current_user)):
    await db.activities.delete_one({"_id": activity_id})
    return {"message": "Aktivität gelöscht"}

# ===========================
# CROSS-DOCKING
# ===========================

class CrossDockingCreate(BaseModel):
    article_id: str
    article_name: str = ""
    quantity: int = 1
    source_event_id: Optional[str] = None
    source_event_name: str = ""
    target_event_id: Optional[str] = None
    target_event_name: str = ""
    transfer_date: Optional[str] = None
    status: str = "geplant"  # geplant | bereit | übertragen | abgebrochen
    notes: Optional[str] = None

class CrossDocking(CrossDockingCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)

@app.get("/api/cross-docking")
async def get_cross_docking(current_user: User = Depends(get_current_user)):
    items = await db.cross_docking.find({}).sort("created_at", -1).to_list(200)
    for item in items:
        item["id"] = str(item.get("_id", item.get("id", "")))
        item.pop("_id", None)
    return items

@app.post("/api/cross-docking")
async def create_cross_docking(cd: CrossDocking, current_user: User = Depends(get_current_user)):
    doc = cd.dict()
    doc["_id"] = doc["id"]
    await db.cross_docking.insert_one(doc)
    doc.pop("_id", None)
    return doc

@app.put("/api/cross-docking/{cd_id}")
async def update_cross_docking(cd_id: str, cd: CrossDockingCreate, current_user: User = Depends(get_current_user)):
    await db.cross_docking.update_one({"_id": cd_id}, {"$set": cd.dict()})
    return {"id": cd_id, **cd.dict()}

@app.delete("/api/cross-docking/{cd_id}")
async def delete_cross_docking(cd_id: str, current_user: User = Depends(get_current_user)):
    await db.cross_docking.delete_one({"_id": cd_id})
    return {"message": "Cross-Docking Eintrag gelöscht"}

# ===========================
# JOB BOARD
# ===========================

class JobBoardEntryCreate(BaseModel):
    title: str
    description: Optional[str] = None
    event_id: Optional[str] = None
    event_name: str = ""
    assigned_to_id: Optional[str] = None
    assigned_to_name: str = ""
    job_type: str = "aufbau"  # aufbau | abbau | transport | technik | sonstiges
    date: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    location: str = ""
    status: str = "offen"  # offen | besetzt | abgeschlossen

class JobBoardEntry(JobBoardEntryCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)

@app.get("/api/job-board")
async def get_job_board(
    date: Optional[str] = None,
    status: Optional[str] = None,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    current_user: User = Depends(get_current_user)
):
    query = {}
    if date:
        query["date"] = date
    if status:
        query["status"] = status
    skip = (page - 1) * page_size
    items = await db.job_board.find(query).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    for item in items:
        item["id"] = str(item.get("_id", item.get("id", "")))
        item.pop("_id", None)
    return items

@app.post("/api/job-board")
async def create_job_board_entry(entry: JobBoardEntry, current_user: User = Depends(get_current_user)):
    doc = entry.dict()
    doc["_id"] = doc["id"]
    await db.job_board.insert_one(doc)
    doc.pop("_id", None)
    return doc

@app.put("/api/job-board/{entry_id}")
async def update_job_board_entry(entry_id: str, entry: JobBoardEntryCreate, current_user: User = Depends(get_current_user)):
    # H1: IDOR fix — verify record exists before modifying
    existing = await db.job_board.find_one({"_id": entry_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Job Board Eintrag nicht gefunden")
    if existing.get("created_by") and existing["created_by"] != current_user.username and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Keine Berechtigung diesen Eintrag zu bearbeiten")
    await db.job_board.update_one({"_id": entry_id}, {"$set": entry.dict()})
    return {"id": entry_id, **entry.dict()}

@app.delete("/api/job-board/{entry_id}")
async def delete_job_board_entry(entry_id: str, current_user: User = Depends(get_current_user)):
    # H1: IDOR fix — verify record exists before deleting
    existing = await db.job_board.find_one({"_id": entry_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Job Board Eintrag nicht gefunden")
    if existing.get("created_by") and existing["created_by"] != current_user.username and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Keine Berechtigung diesen Eintrag zu löschen")
    await db.job_board.delete_one({"_id": entry_id})
    await create_audit_log("DELETE", "job_board", current_user, entity_id=entry_id)
    return {"message": "Job Board Eintrag gelöscht"}

# ===========================
# KOMMUNIKATIONSPROTOKOLL
# ===========================

class CommunicationLogCreate(BaseModel):
    type: str = "email"  # email | note | sms | anruf
    direction: str = "ausgehend"  # ausgehend | eingehend
    subject: str = ""
    body: str = ""
    recipient: str = ""
    sender: str = ""
    customer_id: Optional[str] = None
    customer_name: str = ""
    event_id: Optional[str] = None
    event_name: str = ""
    sent_at: Optional[str] = None
    status: str = "gesendet"  # gesendet | empfangen | entwurf | fehler

class CommunicationLog(CommunicationLogCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)
    created_by: str = ""

@app.get("/api/communication-log")
async def get_communication_log(
    type: Optional[str] = None,
    direction: Optional[str] = None,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    current_user: User = Depends(get_current_user)
):
    query = {}
    if type:
        query["type"] = type
    if direction:
        query["direction"] = direction
    skip = (page - 1) * page_size
    items = await db.communication_log.find(query).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    for item in items:
        item["id"] = str(item.get("_id", item.get("id", "")))
        item.pop("_id", None)
    return items

@app.post("/api/communication-log")
async def create_communication_log(log: CommunicationLogCreate, current_user: User = Depends(get_current_user)):
    import uuid as _uuid
    from datetime import datetime as _dt
    log_id = str(_uuid.uuid4())
    doc = {
        "_id": log_id,
        "id": log_id,
        "created_at": _dt.utcnow().isoformat(),
        "created_by": current_user.username,
        **log.dict(),
    }
    await db.communication_log.insert_one(doc)
    doc.pop("_id", None)
    return doc

@app.put("/api/communication-log/{log_id}")
async def update_communication_log(log_id: str, log: CommunicationLogCreate, current_user: User = Depends(get_current_user)):
    # H1: IDOR fix — verify record exists before modifying
    existing = await db.communication_log.find_one({"_id": log_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Kommunikationseintrag nicht gefunden")
    if existing.get("created_by") and existing["created_by"] != current_user.username and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Keine Berechtigung diesen Eintrag zu bearbeiten")
    await db.communication_log.update_one({"_id": log_id}, {"$set": log.dict()})
    return {"id": log_id, **log.dict()}

@app.delete("/api/communication-log/{log_id}")
async def delete_communication_log(log_id: str, current_user: User = Depends(get_current_user)):
    # H1: IDOR fix — verify record exists before deleting
    existing = await db.communication_log.find_one({"_id": log_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Kommunikationseintrag nicht gefunden")
    if existing.get("created_by") and existing["created_by"] != current_user.username and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Keine Berechtigung diesen Eintrag zu löschen")
    await db.communication_log.delete_one({"_id": log_id})
    await create_audit_log("DELETE", "communication_log", current_user, entity_id=log_id)
    return {"message": "Kommunikationseintrag gelöscht"}

# ===========================
# ARTIKEL ARCHIV-ENDPUNKTE
# ===========================

@app.get("/api/articles/archived")
async def get_archived_articles(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    current_user: User = Depends(get_current_user)
):
    skip = (page - 1) * page_size
    items = await db.articles.find({"archived": True}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    for item in items:
        item["id"] = str(item.get("_id", item.get("id", "")))
        item.pop("_id", None)
    return items

@app.post("/api/articles/{article_id}/archive")
async def archive_article(article_id: str, current_user: User = Depends(get_current_user)):
    result = await db.articles.update_one({"_id": article_id}, {"$set": {"archived": True}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Artikel nicht gefunden")
    return {"message": "Artikel archiviert"}

@app.post("/api/articles/{article_id}/unarchive")
async def unarchive_article(article_id: str, current_user: User = Depends(get_current_user)):
    result = await db.articles.update_one({"_id": article_id}, {"$set": {"archived": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Artikel nicht gefunden")
    return {"message": "Artikel aus Archiv entfernt"}

# ===========================
# LAGERORT ARCHIV-ENDPUNKTE
# ===========================

@app.get("/api/storage-locations/archived")
async def get_archived_storage_locations(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    current_user: User = Depends(get_current_user)
):
    skip = (page - 1) * page_size
    items = await db.storage_locations.find({"archived": True}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    for item in items:
        item["id"] = str(item.get("_id", item.get("id", "")))
        item.pop("_id", None)
    return items

@app.post("/api/storage-locations/{location_id}/archive")
async def archive_storage_location(location_id: str, current_user: User = Depends(get_current_user)):
    result = await db.storage_locations.update_one({"_id": location_id}, {"$set": {"archived": True}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lagerort nicht gefunden")
    return {"message": "Lagerort archiviert"}

@app.post("/api/storage-locations/{location_id}/unarchive")
async def unarchive_storage_location(location_id: str, current_user: User = Depends(get_current_user)):
    result = await db.storage_locations.update_one({"_id": location_id}, {"$set": {"archived": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lagerort nicht gefunden")
    return {"message": "Lagerort aus Archiv entfernt"}

# ===========================
# ZU FAKTURIEREN (BILLING QUEUE)
# ===========================

@app.get("/api/billing-queue")
async def get_billing_queue(current_user: User = Depends(get_current_user)):
    events = await db.events.find({"status": {"$in": ["confirmed", "aktiv", "abgeschlossen"]}}).to_list(5000)
    invoices = await db.invoices.find({}).to_list(10000)
    billed_event_ids = {inv.get("event_id") for inv in invoices if inv.get("event_id")}
    queue = []
    for event in events:
        if event.get("id") not in billed_event_ids:
            queue.append({
                "event_id": event.get("id"),
                "event_name": event.get("name") or event.get("title", ""),
                "customer_name": event.get("customer_name", ""),
                "start_date": event.get("start_date", ""),
                "end_date": event.get("end_date", ""),
                "total_value": event.get("total_value", 0),
                "status": event.get("status", ""),
            })
    return queue

# ─── Rental Requests ────────────────────────────────────────────────────────

def serialize_doc(doc: dict) -> dict:
    if not doc:
        return doc
    doc = dict(doc)
    if "_id" in doc:
        doc["id"] = str(doc["_id"])
        del doc["_id"]
    elif "id" not in doc:
        doc["id"] = ""
    # Convert datetime to ISO string
    from datetime import datetime as _dt
    for k, v in doc.items():
        if isinstance(v, _dt):
            doc[k] = v.isoformat()
    return doc

class RentalRequestCreate(BaseModel):
    customer_name: str
    customer_email: str = ""
    customer_phone: str = ""
    event_name: str = ""
    event_date: Optional[str] = None
    event_location: str = ""
    description: str = ""
    status: str = "neu"  # neu | in_bearbeitung | angebot_gesendet | bestaetigt | abgelehnt
    notes: str = ""

class RentalRequest(RentalRequestCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)
    created_by: str = ""

@app.get("/api/rental-requests")
async def get_rental_requests(status: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if status:
        query["status"] = status
    items = await db.rental_requests.find(query).sort("created_at", -1).to_list(500)
    return [serialize_doc(i) for i in items]

@app.post("/api/rental-requests")
async def create_rental_request(data: RentalRequestCreate, current_user: User = Depends(get_current_user)):
    item = RentalRequest(**data.dict(), created_by=current_user.username)
    doc = item.dict()
    doc["_id"] = doc["id"]
    await db.rental_requests.insert_one(doc)
    doc.pop("_id", None)
    return doc

@app.put("/api/rental-requests/{item_id}")
async def update_rental_request(item_id: str, data: RentalRequestCreate, current_user: User = Depends(get_current_user)):
    result = await db.rental_requests.update_one({"id": item_id}, {"$set": data.dict()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Rental request not found")
    updated = await db.rental_requests.find_one({"id": item_id})
    return serialize_doc(updated)

@app.delete("/api/rental-requests/{item_id}")
async def delete_rental_request(item_id: str, current_user: User = Depends(get_current_user)):
    result = await db.rental_requests.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Rental request not found")
    return {"message": "Deleted"}


# Startup event to ensure admin user exists
@app.on_event("startup")
async def startup_event():
    await ensure_admin_user()
    
    # ===========================================
    # DATABASE INDEXES - Performance Optimization
    # ===========================================
    try:
        # Articles indexes
        await db.articles.create_index("id", unique=True)
        await db.articles.create_index("inventory_code", unique=True)
        await db.articles.create_index("category_id")
        await db.articles.create_index("name")
        await db.articles.create_index("status")
        
        # Users indexes
        await db.users.create_index("id", unique=True)
        await db.users.create_index("username", unique=True)
        await db.users.create_index("email", unique=True)
        
        # Events indexes
        await db.events.create_index("id", unique=True)
        await db.events.create_index("event_number", unique=True)
        await db.events.create_index("customer_id")
        await db.events.create_index("start_date")
        await db.events.create_index("status")
        
        # Bookings indexes
        await db.bookings.create_index("id", unique=True)
        await db.bookings.create_index("event_id")
        await db.bookings.create_index("article_id")
        await db.bookings.create_index("status")
        # Korrekter Feldname: pickup_date/return_date (nicht start_date/end_date)
        await db.bookings.create_index([("pickup_date", 1), ("return_date", 1)])
        # Zusammengesetzter Index für Konflikt-Checks (article_id + status + Datum)
        await db.bookings.create_index([("article_id", 1), ("status", 1), ("pickup_date", 1), ("return_date", 1)])

        # Customers indexes
        await db.customers.create_index("id", unique=True)
        await db.customers.create_index("company_name")

        # Categories indexes
        await db.categories.create_index("id", unique=True)

        # Storage locations indexes
        await db.storage_locations.create_index("id", unique=True)

        # Maintenance indexes
        await db.maintenance_tasks.create_index("id", unique=True)
        await db.maintenance_tasks.create_index("article_id")
        await db.maintenance_tasks.create_index("due_date")
        await db.maintenance_tasks.create_index("status")
        # Compound index for overdue alerts query
        await db.maintenance_tasks.create_index([("due_date", 1), ("status", 1)])

        # Messages indexes (for conversation queries)
        await db.messages.create_index([("sender_id", 1), ("recipient_id", 1), ("created_at", -1)])
        await db.messages.create_index([("recipient_id", 1), ("is_read", 1)])

        # Consumable stock alert compound index
        await db.articles.create_index([("is_consumable", 1), ("current_stock", 1)])

        # Bundles indexes
        await db.bundles.create_index("id", unique=True)
        await db.bundles.create_index("bundle_code", unique=True)

        # Refresh tokens indexes
        await db.refresh_tokens.create_index("token")
        await db.refresh_tokens.create_index("user_id")
        await db.refresh_tokens.create_index("expires_at", expireAfterSeconds=0)  # TTL index

        # F2: Password reset tokens indexes (TTL auto-expires tokens after 1 hour)
        await db.password_reset_tokens.create_index("token", unique=True)
        await db.password_reset_tokens.create_index("expires_at", expireAfterSeconds=0)

        # Soft-delete filter indexes (V7: deleted flag on all major entities)
        await db.articles.create_index("deleted")
        await db.articles.create_index([("deleted", 1), ("status", 1)])
        await db.events.create_index("deleted")
        await db.events.create_index([("deleted", 1), ("status", 1), ("start_date", 1)])
        await db.customers.create_index("is_active")
        await db.customers.create_index([("is_active", 1), ("company_name", 1)])

        # Invoices indexes
        await db.invoices.create_index("id", unique=True)
        await db.invoices.create_index("event_id")
        await db.invoices.create_index("status")
        await db.invoices.create_index("due_date")
        await db.invoices.create_index([("status", 1), ("due_date", 1)])

        # Quotes indexes
        await db.quotes.create_index("id", unique=True)
        await db.quotes.create_index("customer_id")
        await db.quotes.create_index("status")
        await db.quotes.create_index([("customer_id", 1), ("status", 1)])

        # Time entries indexes
        await db.time_entries.create_index("id", unique=True)
        await db.time_entries.create_index("user_id")
        await db.time_entries.create_index("date")
        await db.time_entries.create_index([("user_id", 1), ("date", -1)])

        # Tasks indexes
        await db.tasks.create_index("id", unique=True)
        await db.tasks.create_index("status")
        await db.tasks.create_index("assigned_to")
        await db.tasks.create_index("due_date")
        await db.tasks.create_index([("assigned_to", 1), ("status", 1)])

        # Audit log indexes (for user activity queries)
        await db.audit_logs.create_index("user_id")
        await db.audit_logs.create_index("action")
        await db.audit_logs.create_index([("user_id", 1), ("timestamp", -1)])
        await db.audit_logs.create_index("timestamp")

        # Purchase orders indexes
        await db.purchase_orders.create_index("id", unique=True)
        await db.purchase_orders.create_index("status")
        await db.purchase_orders.create_index([("status", 1), ("created_at", -1)])

        logging.info("Database indexes created/verified successfully")
    except Exception as e:
        logging.error(f"Error creating indexes: {str(e)}")
    
    # Start backup scheduler
    scheduler.add_job(
        scheduled_backup_job,
        'cron',
        hour=2,  # Run daily at 2 AM
        minute=0,
        id='daily_backup',
        replace_existing=True,
        misfire_grace_time=600,
        coalesce=True
    )
    scheduler.start()
    
    logging.info("Application startup complete - Admin user initialized")
    logging.info("Daily backup scheduler started (runs at 2:00 AM)")
    logging.info("Rate limiting enabled: 10 requests/minute for login")

# Sync service initialization removed for now

@app.on_event("shutdown")
async def shutdown_db_client():
    scheduler.shutdown()
    client.close()

# =============================================================================
# /api/v1 Route Aliases — registered AFTER all @app routes above
# Covers both @api_router routes and direct @app routes
# =============================================================================
from fastapi.routing import APIRoute as _APIRoute

_api_router_v1 = APIRouter(prefix="/api/v1")
_seen_v1_paths: set = set()

# 1) Copy all routes from api_router (registered with @api_router decorator)
for _route in api_router.routes:
    if isinstance(_route, _APIRoute):
        _rel = _route.path[len(api_router.prefix):]
        _key = (frozenset(_route.methods or []), _rel)
        if _key not in _seen_v1_paths:
            _seen_v1_paths.add(_key)
            _api_router_v1.add_api_route(
                _rel,
                _route.endpoint,
                methods=list(_route.methods) if _route.methods else None,
                response_model=_route.response_model,
                tags=_route.tags,
                dependencies=_route.dependencies,
            )

# 2) Also copy all direct @app routes starting with /api/ (but not /api/v1/ already)
for _route in list(app.routes):
    if (
        isinstance(_route, _APIRoute)
        and _route.path.startswith("/api/")
        and not _route.path.startswith("/api/v1/")
    ):
        _rel = _route.path[len("/api"):]  # e.g. "/settings/app"
        _key = (frozenset(_route.methods or []), _rel)
        if _key not in _seen_v1_paths:
            _seen_v1_paths.add(_key)
            _api_router_v1.add_api_route(
                _rel,
                _route.endpoint,
                methods=list(_route.methods) if _route.methods else None,
                response_model=_route.response_model,
                tags=_route.tags,
                dependencies=_route.dependencies,
            )

app.include_router(_api_router_v1)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8002)
